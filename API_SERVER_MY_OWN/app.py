"""
FastAPI Backend for RAN KPI Data Retrieval - Universal Query System
Performance Engineering - Telco Company

Architecture:
- Custom query templates for all use cases
- Dynamic aggregation level selection
- Flexible date/time filtering (ranges or multiple values)
- Complex KPI calculations support
- Modular and extensible
"""

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any, Union
from datetime import datetime, date
import mysql.connector
from mysql.connector import Error
from mysql.connector import pooling
from enum import Enum
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

app = FastAPI(title="RAN KPI Data Service - Universal Query System")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================================
# DATABASE CONFIGURATION
# ============================================================================

DB_CONFIG = {
    "host": "10.22.33.116",
    "port": 3306,
    "database": "performanceroute",
    "user": "root",
    "password": "performance",
    #"charset": "utf8mb4"
}

# DB_CONFIG = {
#     "host": "l92.168.145.214",
#     "port": 3306,
#     "database": "local_perf",
#     "user": "root",
#     "password": "performance",
#     #"charset": "utf8mb4"
# }

# ============================================================================
# REQUEST MODELS
# ============================================================================

class QueryRequest(BaseModel):
    query_category: str  # "2G", "3G", "4G", "GENERAL"
    query_subcategory: Optional[str] = None  # Vendor name (e.g., "HUAWEI", "ERICSSON")
    query_name: str  # e.g., "KPI_MONITORING"
    time_granularity: str = "DAILY"  # "HOURLY" or "DAILY"
    
    # Date/Time filters
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    multiple_date: Optional[List[str]] = None
    start_hour: Optional[int] = None
    end_hour: Optional[int] = None
    multiple_hour: Optional[List[int]] = None
    
    # Aggregation
    aggregation_level: Optional[str] = None  # e.g., "cell_name", "site_name"
    include_date: bool = True
    include_time: bool = False
    
    # KPIs to retrieve
    selected_kpis: Optional[List[str]] = None  # If None, returns all KPIs

    # Additional filters
    filters: Optional[Dict[str, List[str]]] = None  # IN filter
    not_in_filters: Optional[Dict[str, List[str]]] = None  # NOT IN filter (exclude values)

# ============================================================================
# DATABASE CONNECTION POOL (Better Performance)
# ============================================================================

# Create connection pool
try:
    connection_pool = pooling.MySQLConnectionPool(
        pool_name="mypool",
        pool_size=5,  # Number of connections to keep in pool
        pool_reset_session=True,
        **DB_CONFIG
    )
    print("✅ Database connection pool created successfully")
except Error as err:
    print(f"❌ Error creating connection pool: {err}")
    connection_pool = None

class DatabaseManager:
    """Handles all database operations with connection pooling"""

    def __init__(self):
        self.pool = connection_pool

    def connect(self):
        """Test database connection"""
        try:
            if self.pool:
                connection = self.pool.get_connection()
                connection.close()
                return True
            else:
                # Fallback to direct connection if pool failed
                connection = mysql.connector.connect(**DB_CONFIG)
                connection.close()
                return True
        except Error as err:
            raise HTTPException(status_code=500, detail=f"Database connection error: {err}")

    def execute_query(self, query: str, params: tuple = None):
        """Execute query using connection pool"""
        connection = None
        cursor = None
        try:
            # Get connection from pool (much faster than creating new one)
            if self.pool:
                connection = self.pool.get_connection()
            else:
                # Fallback to direct connection
                connection = mysql.connector.connect(**DB_CONFIG)

            cursor = connection.cursor(dictionary=True)

            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)

            results = cursor.fetchall()
            return results

        except Error as err:
            raise HTTPException(status_code=500, detail=f"Query execution error: {err}")

        finally:
            if cursor:
                try:
                    cursor.close()
                except:
                    pass
            if connection and connection.is_connected():
                connection.close()  # Returns connection to pool, doesn't actually close it

db_manager = DatabaseManager()

# ============================================================================
# GLOBAL GRANULARITY TRANSFORMS (for raw_sql query type)
# ============================================================================

GRANULARITY_TRANSFORMS = {
    "HOURLY": {
        "date_transform": "{col}",
        "time_transform": "{col}",
        "include_time": True
    },
    "DAILY": {
        "date_transform": "{col}",
        "time_transform": None,
        "include_time": False
    },
    "WEEKLY": {
        "date_transform": "YEARWEEK({col}, 1)",
        "time_transform": None,
        "include_time": False
    },
    "MONTHLY": {
        "date_transform": "DATE_FORMAT({col}, '%Y-%m')",
        "time_transform": None,
        "include_time": False
    }
}

# ============================================================================
# QUERY CONFIGURATIONS
# ============================================================================

QUERY_CONFIG = {

    "2G": {
        "HUAWEI": {

            "KPI_MONITORING": {
                "query_type": "multi_cte",
                "description": "2G  KPI monitoring - Combined Vendors",
                "date_time_filters": ["start_date", "end_date", "multiple_date" , 'start_hour', 'end_hour', 'multiple_hour'],

                # Subqueries configuration
                "subqueries": [
                    {
                        "name": "HUAWEI2G_KPI_1",
                        "table": "hourly_arcep_huawei_2g",
                        "alias": "t1",
                        "time_granularities": {
                            "HOURLY": {"is_available": True, "date_field": "DATE", "time_field": "TIME"},
                            "DAILY": {"is_available": True, "date_field": "DATE", "time_field": None},
                            "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(DATE, 1)", "time_field": None},
                            "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(DATE, '%Y-%m')", "time_field": None}
                        },
                        "aggregation_levels": {
                            "cell_name": "SUBSTRING_INDEX(SUBSTRING_INDEX(UCELL, '=', -3), ',', 1)",
                            "site_name": "S.site",
                            "arrondissement": "s.ARRONDISSEMENT",
                            "commune": "s.COMMUNE",
                            "controller_name": "NE_Name"
                        },
                        "kpi_fields": {
                            "CSSR_HUAWEI": """100 * (SUM(Suc_SDCCH_Seiz) / NULLIF(SUM(SDCCH_Seiz_Req), 0)) *
                                (1 - (SUM(SDCCHDrop) / NULLIF(SUM(Suc_SDCCH_Seiz), 0))) *
                                (SUM(Suc_TCH_Seiz_SG) + SUM(Suc_TCH_Seiz_TrafChann) + SUM(Suc_TCH_Seiz_handTrafChan)) /
                                NULLIF((SUM(TCH_Seizure_Req_SC) + SUM(TCH_Seiz_Req_TrafChan) + SUM(TCH_Seiz_Req_HandTrafChan)), 0) *
                                (1 - (SUM(TCHDrop_Nume_ARCEP) / NULLIF(SUM(TCHDrop_Deno_ARCEP), 0)))""",
                            "CDR_HUAWEI": "(100 * SUM(TCHDrop_Nume_ARCEP) / NULLIF(SUM(TCHDrop_Deno_ARCEP), 0))",
                            "CBR_HUAWEI": "100 * (SUM(TCHCONG_Nume_ARCEP) / NULLIF(SUM(TCHCONG_Deno_ARCEP), 0))"
                        },
                        "sql_template": """
                            SELECT {fields}
                            FROM {table} {alias}
                            INNER JOIN
                            sites_and_info_2g_huawei AS S ON S.Gcell = UCELL
                            WHERE {where_conditions}
                            {group_by}
                        """
                    },
                    {
                        "name": "HUAWEI2G_KPI_2",
                        "table": "hourly_huawei_trafic_per_cell_site_2g",
                        "alias": "e",
                        "time_granularities": {
                            "HOURLY": {"is_available": True, "date_field": "DATE", "time_field": "TIME"},
                            "DAILY": {"is_available": True, "date_field": "DATE", "time_field": None},
                            "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(DATE, 1)", "time_field": None},
                            "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(DATE, '%Y-%m')", "time_field": None}
                        },
                        "aggregation_levels": {
                            "cell_name": "cell_name",
                            "site_name": "S.site",
                            "arrondissement": "s.ARRONDISSEMENT",
                            "commune": "s.COMMUNE",
                            "controller_name": "NE_Name"
                        },
                        "kpi_fields": {
                            "voix_2g_huawei": """ROUND(SUM(Total_Voice), 2)""",
                            "DATA2g": "ROUND(SUM(Total_Data), 2) AS DATA2g",
                        },
                        "sql_template": """
                            SELECT {fields}
                            FROM {table} {alias}
                            INNER JOIN
                            sites_and_info_2g_huawei AS S ON S.cell = H.CELL_NAME
                            WHERE {where_conditions}
                            {group_by}
                        """
                    }
                ],

                # How to join the CTEs
                "cte_joins": [
                    {
                        "type": "INNER JOIN",
                        "left": "HUAWEI2G_KPI_1",
                        "right": "HUAWEI2G_KPI_2",
                        "on": ["HUAWEI2G_KPI_1.date = HUAWEI2G_KPI_2.DATE",
                                "HUAWEI2G_KPI_1.time = HUAWEI2G_KPI_2.TIME",
                                "HUAWEI2G_KPI_1.cell_name = HUAWEI2G_KPI_2.cell_name"]
                    }
                ],

                # Outer query configuration (what the client controls)
                "time_granularities": {
                    "HOURLY": {"is_available": True, "date_field": "HUAWEI2G_KPI_1.date", "time_field": "HUAWEI2G_KPI_1.TIME"},
                    "DAILY": {"is_available": True, "date_field": "HUAWEI2G_KPI_1.date", "time_field": None},
                    "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(HUAWEI2G_KPI_1.date, 1)", "time_field": None},
                    "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(HUAWEI2G_KPI_1.date, '%Y-%m')", "time_field": None}
                },

                "aggregation_levels": {
                    "cell_name": "HUAWEI2G_KPI_1.cell_name",
                    "site_name": "HUAWEI2G_KPI_1.site_name",
                    "arrondissement": "HUAWEI2G_KPI_1.ARRONDISSEMENT",
                    "commune": "HUAWEI2G_KPI_1.COMMUNE",
                    "controller_name": "HUAWEI2G_KPI_1.controller_name"
                },

                "kpi_fields": {
                    "CSSR_HUAWEI": "HUAWEI2G_KPI_1.CSSR_HUAWEI",
                    "CDR_HUAWEI": "HUAWEI2G_KPI_1.CDR_HUAWEI",
                    "CBR_HUAWEI": "HUAWEI2G_KPI_1.CBR_HUAWEI",

                    "TRAFFIC_DATA": "HUAWEI2G_KPI_2.DATA2g",
                    "TRAFFIC_VOIX": "HUAWEI2G_KPI_2.voix_2g_huawei",
                },

                "sql_template": """
                    SELECT {fields}
                    FROM {cte_joins}
                    {group_by}
                    {order_by}
                """
            },
            
            
            "KPI_MONITORING2": {
                "description": "2G HUAWEI KPI Monitoring",
                "date_time_filters": ["start_date", "end_date", "multiple_date", "start_hour", "end_hour", "multiple_hour"],
                "parameters": [],
                "aggregation_levels": {
                   
                },
                "time_granularities": {
                    "HOURLY": {"is_available": True, "date_field": "DATE", "time_field": "TIME_FORMAT(TIME, '%H:%i:%s')"},
                    "DAILY": {"is_available": True, "date_field": "DATE", "time_field": None},
                    "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(DATE, 1)", "time_field": None},
                    "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(DATE, '%Y-%m')", "time_field": None}
                },
                "kpi_fields": {
                     "CSSR_HUAWEI": """100 * (SUM(Suc_SDCCH_Seiz) / NULLIF(SUM(SDCCH_Seiz_Req), 0)) *
                        (1 - (SUM(SDCCHDrop) / NULLIF(SUM(Suc_SDCCH_Seiz), 0))) *
                        (SUM(Suc_TCH_Seiz_SG) + SUM(Suc_TCH_Seiz_TrafChann) + SUM(Suc_TCH_Seiz_handTrafChan)) /
                        NULLIF((SUM(TCH_Seizure_Req_SC) + SUM(TCH_Seiz_Req_TrafChan) + SUM(TCH_Seiz_Req_HandTrafChan)), 0) *
                        (1 - (SUM(TCHDrop_Nume_ARCEP) / NULLIF(SUM(TCHDrop_Deno_ARCEP), 0)))""",

                    "CDR_HUAWEI": "(100 * SUM(TCHDrop_Nume_ARCEP) / NULLIF(SUM(TCHDrop_Deno_ARCEP), 0))",
                    "CBR_HUAWEI": "100 * (SUM(TCHCONG_Nume_ARCEP) / NULLIF(SUM(TCHCONG_Deno_ARCEP), 0))"
                },

                "chart_configs":{
                    
                    "CSSR_HUAWEI":{
                        "KPIS":['CSSR_HUAWEI'],
                        "title":"2G Huawei CSSR ",
                        "default_type": "line",
                        "y_axis_title":"CSSR (%)",
                        "threshold":99
                    },
                    "CSSR_CDR_HUAWEI":{
                        "KPIS":["CSSR_HUAWEI","CDR_HUAWEI"],
                        "title":"2G Huawei CSSR vs CDR",
                        "default_type_1": "line",
                        "default_type_2": "bar",
                        "is_dual_axis":True,
                        "y_axis_titles":["CSSR (%)", "CDR (%)"]
                    },
                    "CBR_HUAWEI":{
                        "KPIS":["CBR_HUAWEI"],
                        "title":"2G Huawei CBR ",
                        "default_type": "line",
                        "y_axis_title":"CBR (%)",
                        "threshold":3
                    }

                },

                "sql_template": """
                SELECT {fields}
                FROM hourly_arcep_huawei_2g t
                INNER JOIN
                sites_and_info_2g_huawei AS S ON S.Gcell = UCELL
                WHERE {where_conditions}
                {group_by}
                ORDER BY {order_by}
                """
            }

        },

        "ERICSSON": {

            "KPI_MONITORING": {
                "description": "2G Ericsson KPI Monitoring",
                "date_time_filters": ["start_date", "end_date", "multiple_date", "start_hour", "end_hour", "multiple_hour"],
                "parameters": [],
                "aggregation_levels": {
                    "cell_name": "m.CELL_NAME",
                    "site_name": "l.SITE_NAME",
                    "arrondissement": "l.ARRONDISSEMENT",
                    "commune": "l.COMMUNE",
                    "controller_name": "m.CONTROLLER_NAME"
                },
                "time_granularities": {
                    "HOURLY": {"is_available": True, "date_field": "m.DATE", "time_field": "m.TIME"},
                    "DAILY": {"is_available": True, "date_field": "m.DATE", "time_field": None},
                    "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(m.DATE, 1)", "time_field": None},
                    "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(m.DATE, '%Y-%m')", "time_field": None}
                },
                "kpi_fields": {
                    "Call_Success_Rate": "AVG(ARCEP_2G_CALL_SUCCESS_RATE)",
                    "Call_Drop_Rate": "AVG(ARCEP_2G_CALL_DROP_RATE)",
                    "TCH_Congestion": "AVG(Congestion_Tch_pct)",
                    "SDCCH_Congestion": "AVG(SDCCH_Congestion_Rate)",
                    "Availability": "AVG(`2G_Availability`)"
                },
                "sql_template": """
                SELECT {fields}
                FROM kpi_2g_ericsson m
                LEFT JOIN lookup_2g_ericsson l ON m.CELL_NAME = l.CELL_NAME
                WHERE {where_conditions}
                {group_by}
                ORDER BY {order_by}
                """
            }

        }

        ,

        
        "COMBINED": {
            "NETWORK": {
                "query_type": "raw_sql",
                "description": "2G Network KPIs HUAWEI ERICSSON",
                "date_time_filters": ["start_date", "end_date", "multiple_date", "start_hour", "end_hour", "multiple_hour"],

                "time_granularities": {
                    "HOURLY": {"is_available": True},
                    "DAILY": {"is_available": True},
                    "WEEKLY": {"is_available": True},
                    "MONTHLY": {"is_available": True}
                },

                "sources": {
                    "eric": {
                        "date_col": "e.DATE",
                        "time_col": "e.TIME",
                        "aggregations": {
                        }
                    },
                    "huawei": {
                        "date_col": "h.date",
                        "time_col": "h.time",
                        "aggregations": {
                        }
                    },
                    "combined":{
                        "date_col": "COALESCE(h.date, e.DATE)",
                        "time_col": "COALESCE(h.time, e.TIME)",
                        "aggregations": {
                        }
                    }
                },

                 # Joins config: auto-generate JOIN ON conditions based on granularity/aggregation
                "joins": [
                    {
                        "left": "eric",
                        "right": "huawei",
                        "include_date": True,       # Always include date
                        "include_time": True,       # Only if HOURLY granularity
                        "include_aggregation": True, # Only if aggregation is selected
                        "custom": []                 # Extra static conditions if needed
                    }
                ],

                  "chart_configs": {
                    "CSSR_ALL": {
                        "KPIS": ["CSSR_NETWORK", "CSSR_HUAWEI", "CSSR_ERICSSON"],
                        "title": "2G CSSR - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "CSSR (%)",
                        "threshold": 99
                    },
                    "CDR_ALL": {
                        "KPIS": ["CDR_NETWORK", "CDR_HUAWEI", "CDR_ERICSSON"],
                        "title": "2G CDR - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "CDR (%)"
                    },
                    "CBR_ALL": {
                        "KPIS": ["CBR_NETWORK", "CBR_HUAWEI", "CBR_ERICSSON"],
                        "title": "2G CBR - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "CBR (%)",
                        "threshold": 3
                    },
                    "TCH_CONGESTION_RATE_ALL": {
                        "KPIS": ["TCH_CONGESTION_RATE_NETWORK", "TCH_CONGESTION_RATE_HUAWEI", "TCH_CONGESTION_RATE_ERICSSON"],
                        "title": "2G TCH Congestion - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "TCH Congestion (%)",
                        "threshold": 2
                    },
                    "SDCCH_CONGESTION_RATE_ALL": {
                        "KPIS": ["SDCCH_CONGESTION_RATE_NETWORK", "SDCCH_CONGESTION_RATE_HUAWEI", "SDCCH_CONGESTION_RATE_ERICSSON"],
                        "title": "2G SDCCH Congestion - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "SDCCH Congestion (%)",
                        "threshold": 1
                    },
                    "SDCCH_DROP_RATE_ALL": {
                        "KPIS": ["SDCCH_DROP_RATE_NETWORK", "SDCCH_DROP_RATE_HUAWEI", "SDCCH_DROP_RATE_ERICSSON"],
                        "title": "2G SDCCH Drop Rate - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "SDCCH Drop Rate (%)"
                    },
                    "CELL_AVAILABILITY_ALL": {
                        "KPIS": ["CELL_AVAILABILITY_RATE_NETWORK", "CELL_AVAILABILITY_RATE_HUAWEI", "CELL_AVAILABILITY_RATE_ERICSSON"],
                        "title": "2G Cell Availability - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "Availability (%)",
                        "threshold": 99
                    },
                    "SDCCH_TRAFFIC_ALL": {
                        "KPIS": ["SDCCH_TRAFFIC_NETWORK", "SDCCH_TRAFFIC_HUAWEI", "SDCCH_TRAFFIC_ERICSSON"],
                        "title": "2G SDCCH Traffic - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "SDCCH Traffic (Erl)"
                    },

                    "CSSR_NETWORK_vs_CDR": {
                        "KPIS": ["CSSR_NETWORK", "CDR_NETWORK"],
                        "title": "2G Network CSSR vs CDR",
                        "default_type_1": "line",
                        "default_type_2": "bar",
                        "is_dual_axis": True,
                        "y_axis_titles": ["CSSR (%)", "CDR (%)"]
                    }

                },

                "sql_template": """
                WITH
  HUAWEI_2G_KPI AS (
    SELECT
      {huawei__select_fields}
      -- h.CELL_NAME,
      -- ept.SITE_NAME,
      -- ept.ARRONDISSEMENT,
      -- ept.COMMUNE,
      -- ept.DEPARTEMENT,

      
      -- CASE
      --   WHEN COALESCE(SUM(CAST(CELL_KPI_SD_REQ AS DOUBLE)), 0) = 0
      --     OR COALESCE(SUM(CAST(CELL_KPI_SD_SUCC AS DOUBLE)), 0) = 0
      --     OR (COALESCE(SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE)), 0)
      --          + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
      --          + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0)) = 0
      --     OR (COALESCE(SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE)), 0)
      --          + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE)), 0)
      --          + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)), 0)) = 0
      --   THEN 0
      --   ELSE
      --     100.0
      --     * (SUM(CAST(CELL_KPI_SD_SUCC AS DOUBLE))
      --         / SUM(CAST(CELL_KPI_SD_REQ AS DOUBLE)))
      --     * (1.0 - (SUM(CAST(CELL_SD_CALL_DROPS AS DOUBLE))
      --                / SUM(CAST(CELL_KPI_SD_SUCC AS DOUBLE))))
      --     * ((SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE))
      --          + SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE))
      --          + SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)))
      --         / (SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE))
      --             + SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE))
      --             + SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE))))
      --     * (1.0 - ((SUM(CAST(CELL_KPI_TCH_DROPS_SIG AS DOUBLE))
      --                 + SUM(CAST(CELL_TRAF_CH_CALL_DROPS AS DOUBLE))
      --                 + SUM(CAST(CELL_KPI_TCH_HO_DROPS_TRAF AS DOUBLE)))
      --                / (SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE))
      --                    + SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE))
      --                    + SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)))))
      -- END AS CSSR_HUAWEI
      

      ,100 * (SUM(CELL_KPI_SD_SUCC
      ) / SUM(CELL_KPI_SD_REQ)) * 
  (1 - (SUM(CELL_SD_CALL_DROPS) / SUM(CELL_KPI_SD_SUCC))) * 
  (SUM(CELL_KPI_TCH_SUCC_SIG) + SUM(CELL_KPI_TCH_ASS_SUCC_TRAF) + SUM(CELL_KPI_TCH_HO_SUCC_TRAF)) /
  (SUM(CELL_KPI_TCH_REQ_SIG) + SUM(CELL_KPI_TCH_ASS_REQ_TRAF) + SUM(CELL_KPI_TCH_HO_REQ_TRAF)) *
  (1 - (((sum(CELL_KPI_TCH_DROPS_SIG)+sum(CELL_KPI_TCH_STATIC_DROPS_TRAF)+sum(CELL_KPI_TCH_HO_DROPS_TRAF))) /
    ((sum(CELL_KPI_TCH_SUCC_SIG)+sum(CELL_KPI_TCH_ASS_SUCC_TRAF)+sum(CELL_KPI_TCH_HO_SUCC_TRAF))))) AS  CSSR_HUAWEI,

      -- ARCEP 2G CALL DROP RATE - HUAWEI
      100.0 * (
        COALESCE(SUM(CAST(CELL_KPI_TCH_DROPS_SIG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_TRAF_CH_CALL_DROPS AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_DROPS_TRAF AS DOUBLE)), 0)
      ) / NULLIF(
        COALESCE(SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)), 0),
        0
      ) AS CDR_HUAWEI,

      -- ARCEP 2G CALL BLOCKING RATE - HUAWEI
      100.0 * (
        COALESCE(SUM(CAST(CELL_KPI_TCH_CONG_SIG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_CONG_TRAF AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_CONGEST_TRAF AS DOUBLE)), 0)
      ) / NULLIF(
        COALESCE(SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0),
        0
      ) AS CBR_HUAWEI,

      -- TCH Congestion Rate - HUAWEI
      100.0 * (
        COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_CONG_TRAF AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_CONGEST_TRAF AS DOUBLE)), 0)
      ) / NULLIF(
        COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0),
        0
      ) AS TCH_CONGESTION_RATE_HUAWEI,

      -- SDCCH Congestion Rate - HUAWEI
      100.0 * COALESCE(SUM(CAST(CELL_KPI_SD_CONGEST AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(CELL_KPI_SD_REQ AS DOUBLE)), 0), 0) AS SDCCH_CONGESTION_RATE_HUAWEI,

      -- SDCCH Drop Rate - HUAWEI
      100.0 * COALESCE(SUM(CAST(CELL_SD_CALL_DROPS AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(CELL_IMM_ASS_SUCC_SD AS DOUBLE)), 0), 0) AS SDCCH_DROP_RATE_HUAWEI,

      -- ARCEP TCH Assignment Success Rate - HUAWEI (Vendor-specific)
      100.0 * (
        COALESCE(SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)), 0)
      ) / NULLIF(
        COALESCE(SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0),
        0
      ) AS TCH_ASSIGNMENT_SUCCESS_RATE_HUAWEI,

      -- SDCCH Traffic (Erlangs) - HUAWEI
      COALESCE(SUM(CAST(CELL_KPI_SD_TRAF_ERL AS DOUBLE)), 0) AS SDCCH_TRAFFIC_HUAWEI,

      -- 2G Availability Rate - HUAWEI
      100.0 * COALESCE(SUM(CAST(CELL_KPI_TCH_AVAIL_NUM AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(CELL_KPI_TCH_CFG_NUM AS DOUBLE)), 0), 0) AS CELL_AVAILABILITY_RATE_HUAWEI,


        COALESCE(SUM(CAST(CELL_KPI_TCH_TRAF_ERL_TRAF AS DOUBLE)), 0) TRAFFIC_VOIX_HUAWEI,
        -- handover sr
        100.0 * (
  (COALESCE(SUM(CAST(CELL_INTRABSC_OUTCELL_HO_SUCC AS DOUBLE)), 0)
   + COALESCE(SUM(CAST(CELL_INTERBSC_OUTCELL_HO_SUCC AS DOUBLE)), 0))
  / NULLIF(
      COALESCE(SUM(CAST(CELL_INTRABSC_OUTCELL_HO_CMD AS DOUBLE)), 0)
      + COALESCE(SUM(CAST(CELL_INTERBSC_OUTCELL_HO_CMD AS DOUBLE)), 0),
      0
    )
) AS HANDOVER_SUCCESS_RATE_HUAWEI



    FROM
      hourly_huawei_2g_all_counters h
    -- LEFT JOIN
    --   EPT_2G ept ON h.CELL_NAME = ept.CELL_NAME AND ept.VENDOR = 'HUAWEI'

    WHERE
      -- h.date BETWEEN '2026-01-26' AND '2026-01-26'  -- Adjust date range as needed
        {huawei__where_clause}
    GROUP BY
     {huawei__group_by}
  ),

  ERICSSON_2G_KPI AS (
    SELECT
     {eric__select_fields}
      -- e.CELL_NAME,
      -- ept.SITE_NAME,
      -- ept.ARRONDISSEMENT,
      -- ept.COMMUNE,
      -- ept.DEPARTEMENT,

      -- CSSR (Call Setup Success Rate) - ERICSSON
      ,CASE
        WHEN COALESCE(SUM(CAST(CCALLS AS DOUBLE)), 0) = 0
          OR COALESCE(SUM(CAST(CMSESTAB AS DOUBLE)), 0) = 0
          OR COALESCE(SUM(CAST(TASSALL AS DOUBLE)), 0) = 0
          OR (COALESCE(SUM(CAST(TFCASSALL AS DOUBLE)), 0)
               + COALESCE(SUM(CAST(THCASSALL AS DOUBLE)), 0)
               + COALESCE(SUM(CAST(TFCASSALLSUB AS DOUBLE)), 0)
               + COALESCE(SUM(CAST(THCASSALLSUB AS DOUBLE)), 0)
            ) = 0
        THEN 0
        ELSE
          100.0
          * (1.0 - (SUM(CAST(CCONGS AS DOUBLE))
                     / SUM(CAST(CCALLS AS DOUBLE))))
          * (1.0 - ((SUM(CAST(CNDROP AS DOUBLE))
                      - SUM(CAST(CNRELCONG AS DOUBLE)))
                     / SUM(CAST(CMSESTAB AS DOUBLE))))
          * (SUM(CAST(TCASSALL AS DOUBLE))
              / SUM(CAST(TASSALL AS DOUBLE)))
          * (1.0 - (
              (SUM(CAST(TFNDROP AS DOUBLE))
                + SUM(CAST(THNDROP AS DOUBLE))
                + SUM(CAST(TFNDROPSUB AS DOUBLE))
                + SUM(CAST(THNDROPSUB AS DOUBLE))
              )
              / (
                SUM(CAST(TFCASSALL AS DOUBLE))
                + SUM(CAST(THCASSALL AS DOUBLE))
                + SUM(CAST(TFCASSALLSUB AS DOUBLE))
                + SUM(CAST(THCASSALLSUB AS DOUBLE))
              )
            ))
      END AS CSSR_ERICSSON,

      -- Call Blocking Rate - ERICSSON (ARCEP)
      100.0 * (
        COALESCE(SUM(CAST(CNRELCONG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(TFNRELCONG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(TFNRELCONGSUB AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(THNRELCONG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(THNRELCONGSUB AS DOUBLE)), 0)
      ) / NULLIF(COALESCE(SUM(CAST(TASSALL AS DOUBLE)), 0), 0) AS CBR_ERICSSON,

      -- Call Drop Rate - ERICSSON (ARCEP)
      100.0 * (
        COALESCE(SUM(CAST(THNDROP AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(THNDROPSUB AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(TFNDROP AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(TFNDROPSUB AS DOUBLE)), 0)
      ) / NULLIF(
        COALESCE(SUM(CAST(TFCASSALL AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(TFCASSALLSUB AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(THCASSALL AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(THCASSALLSUB AS DOUBLE)), 0),
        0
      ) AS CDR_ERICSSON,

      -- 2G Cell Availability Rate - ERICSSON
      100.0 - (100.0 * COALESCE(SUM(CAST(TDWNACC AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(TDWNSCAN AS DOUBLE)), 0), 0)) AS CELL_AVAILABILITY_RATE_ERICSSON,

      -- TCH Channel Availability Rate - ERICSSON (Vendor-specific)
      100.0 * (COALESCE(SUM(CAST(TAVAACC AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(TAVASCAN AS DOUBLE)), 0), 0)) /
        (COALESCE(AVG(CAST(AVG_TNUCHCNT AS DOUBLE)), 1)) AS TCH_AVAILABILITY_RATE_ERICSSON,

      -- Downtime Manual - ERICSSON (Vendor-specific)
      COALESCE(SUM(CAST(HDWNACC AS DOUBLE)), 0) AS DOWNTIME_MANUAL,

      -- TCH Congestion Rate - ERICSSON
      100.0 * (
        COALESCE(SUM(CAST(CNRELCONG AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(TFNRELCONG AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(TFNRELCONGSUB AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(THNRELCONG AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(THNRELCONGSUB AS DOUBLE)), 0)
      ) / NULLIF(COALESCE(SUM(CAST(TASSALL AS DOUBLE)), 0), 0) AS TCH_CONGESTION_RATE_ERICSSON,

      -- SDCCH Drop Rate - ERICSSON
      100.0 * COALESCE(SUM(CAST(CNDROP AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(CMSESTAB AS DOUBLE)), 0), 0) AS SDCCH_DROP_RATE_ERICSSON,

      -- SDCCH Traffic - ERICSSON (Erlangs)
      COALESCE(SUM(CAST(CTRALACC AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(CNSCAN AS DOUBLE)), 0), 0) AS SDCCH_TRAFFIC_ERICSSON,

      -- SDCCH Blocking Rate - ERICSSON (Vendor-specific)
      100.0 * COALESCE(SUM(CAST(CCONGS AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(CCALLS AS DOUBLE)), 0), 0) AS SDCCH_BLOCKING_RATE_ERICSSON,

      -- SDCCH Congestion Rate - ERICSSON
      100.0 * (COALESCE(SUM(CAST(CCONGS AS DOUBLE)), 0) + COALESCE(SUM(CAST(CCONGSSUB AS DOUBLE)), 0)) /
        NULLIF(COALESCE(SUM(CAST(CCALLS AS DOUBLE)), 0), 0) AS SDCCH_CONGESTION_RATE_ERICSSON,

      COALESCE(SUM(CAST(TRAFFIC_DATA_GB AS DOUBLE)), 0) TRAFFIC_DATA_GB_ERICSSON,
      COALESCE(SUM(CAST(TRAFFIC_VOIX AS DOUBLE)), 0) TRAFFIC_VOIX_ERICSSON
    FROM
      hourly_ericsson_arcep_2g_counters e
    -- LEFT JOIN
    --   EPT_2G ept ON e.CELL_NAME = ept.CELL_NAME AND ept.VENDOR = 'ERICSSON'

    WHERE
      -- e.DATE BETWEEN '2026-01-28' AND '2026-01-29'  -- Adjust date range as needed
        {eric__where_clause}

    GROUP BY
     {eric__group_by}
  )



  

SELECT
  {combined__select_fields}

  -- Location Info
  -- COALESCE(h.SITE_NAME, e.SITE_NAME) AS SITE_NAME,
  -- COALESCE(h.ARRONDISSEMENT, e.ARRONDISSEMENT) AS ARRONDISSEMENT,
  -- COALESCE(h.COMMUNE, e.COMMUNE) AS COMMUNE,
  -- COALESCE(h.DEPARTEMENT, e.DEPARTEMENT) AS DEPARTEMENT,

  -- Vendor-specific KPIs
  ,h.CSSR_HUAWEI,
  -- h.CSSR_HUAWEI_2,
  e.CSSR_ERICSSON,
  h.CDR_HUAWEI,
  e.CDR_ERICSSON,
  h.CBR_HUAWEI,
  e.CBR_ERICSSON,
  h.TCH_CONGESTION_RATE_HUAWEI,
  e.TCH_CONGESTION_RATE_ERICSSON,
  h.SDCCH_CONGESTION_RATE_HUAWEI,
  e.SDCCH_CONGESTION_RATE_ERICSSON,
  h.SDCCH_DROP_RATE_HUAWEI,
  e.SDCCH_DROP_RATE_ERICSSON,
  h.SDCCH_TRAFFIC_HUAWEI,
  e.SDCCH_TRAFFIC_ERICSSON,
  h.CELL_AVAILABILITY_RATE_HUAWEI,
  e.CELL_AVAILABILITY_RATE_ERICSSON,

  h.TRAFFIC_VOIX_HUAWEI,
  e.TRAFFIC_VOIX_ERICSSON,

  -- Vendor-specific only
  h.TCH_ASSIGNMENT_SUCCESS_RATE_HUAWEI,
  e.TCH_AVAILABILITY_RATE_ERICSSON,
  e.DOWNTIME_MANUAL,
  e.SDCCH_BLOCKING_RATE_ERICSSON,

  h.HANDOVER_SUCCESS_RATE_HUAWEI,

  -- Network-level weighted averages (0.22 Huawei + 0.78 Ericsson)
  (COALESCE(h.CSSR_HUAWEI, 0) * 0.22) + (COALESCE(e.CSSR_ERICSSON, 0) * 0.78) AS CSSR_NETWORK,
  (COALESCE(h.CDR_HUAWEI, 0) * 0.22) + (COALESCE(e.CDR_ERICSSON, 0) * 0.78) AS CDR_NETWORK,
  (COALESCE(h.CBR_HUAWEI, 0) * 0.22) + (COALESCE(e.CBR_ERICSSON, 0) * 0.78) AS CBR_NETWORK,
  (COALESCE(h.TCH_CONGESTION_RATE_HUAWEI, 0) * 0.22) + (COALESCE(e.TCH_CONGESTION_RATE_ERICSSON, 0) * 0.78) AS TCH_CONGESTION_RATE_NETWORK,
  (COALESCE(h.SDCCH_CONGESTION_RATE_HUAWEI, 0) * 0.22) + (COALESCE(e.SDCCH_CONGESTION_RATE_ERICSSON, 0) * 0.78) AS SDCCH_CONGESTION_RATE_NETWORK,
  (COALESCE(h.SDCCH_DROP_RATE_HUAWEI, 0) * 0.22) + (COALESCE(e.SDCCH_DROP_RATE_ERICSSON, 0) * 0.78) AS SDCCH_DROP_RATE_NETWORK,
  (COALESCE(h.SDCCH_TRAFFIC_HUAWEI, 0) * 0.22) + (COALESCE(e.SDCCH_TRAFFIC_ERICSSON, 0) * 0.78) AS SDCCH_TRAFFIC_NETWORK,
  (COALESCE(h.CELL_AVAILABILITY_RATE_HUAWEI, 0) * 0.22) + (COALESCE(e.CELL_AVAILABILITY_RATE_ERICSSON, 0) * 0.78) AS CELL_AVAILABILITY_RATE_NETWORK,
  h.TRAFFIC_VOIX_HUAWEI + e.TRAFFIC_VOIX_ERICSSON AS TRAFFIC_NETWORK

FROM HUAWEI_2G_KPI h
LEFT JOIN ERICSSON_2G_KPI e
ON {eric__huawei__join_on}
                """
            }
        }
        

    },

    "3G": {
        "HUAWEI": {
            "KPI_MONITORING": {
                "description": "3G Huawei KPI Monitoring",
                "date_time_filters": ["start_date", "end_date", "multiple_date"],
                "parameters": [],
                "aggregation_levels": {
                    "cell_name": "m.CELL_NAME",
                    "site_name": "l.SITE_NAME",
                    "commune": "l.COMMUNE",
                    "controller_name": "m.CONTROLLER_NAME"
                },
                "time_granularities": {
                    "DAILY": {"is_available": True, "date_field": "m.DATE", "time_field": None},
                    "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(m.DATE, 1)", "time_field": None},
                    "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(m.DATE, '%Y-%m')", "time_field": None}
                },
                "kpi_fields": {
                    "CS_Call_Success_Rate": "AVG(VS_CSSR_3G)",
                    "PS_RAB_Success": "AVG(VS_RAB_PS_SR)"
                },
                "sql_template": """
                    SELECT {fields}
                    FROM kpi_3g_huawei m
                    LEFT JOIN lookup_3g_huawei l ON m.CELL_NAME = l.CELL_NAME
                    WHERE {where_conditions}
                    {group_by}
                    ORDER BY {order_by}
                """
            }
        },

        
        "COMBINED": {
            "NETWORK": {
                "query_type": "raw_sql",
                "description": "Sample raw_sql query - 2G Network KPIs",
                "date_time_filters": ["start_date", "end_date", "multiple_date", "start_hour", "end_hour", "multiple_hour"],

                "time_granularities": {
                    "HOURLY": {"is_available": True},
                    "DAILY": {"is_available": True},
                    "WEEKLY": {"is_available": True},
                    "MONTHLY": {"is_available": True}
                },

                "sources": {
                    "eric": {
                        "date_col": "e.DATE",
                        "time_col": "e.TIME",
                        "aggregations": {
                        }
                    },
                    "huawei": {
                        "date_col": "h.date",
                        "time_col": "h.time",
                        "aggregations": {
                        }
                    },

                    "combined":{
                        "date_col": "COALESCE(h.date, e.DATE)",
                        "time_col": "COALESCE(h.time, e.TIME)",
                        "aggregations": {
                        }
                    },
                    ##huawei sources
                    #being used for "date" and "time"
                    "huaweiSimple": {
                        "date_col": "date",
                        "time_col": "time",
                        "aggregations": {
                        }
                    },
                    "T1huawei": {
                        "date_col": "T1.date",
                        "time_col": "T1.time",
                        "aggregations": {
                        }
                    },
                    "T2huawei": {
                        "date_col": "T2.date",
                        "time_col": "T2.time",
                        "aggregations": {
                          
                        }
                    },
                    "T3huawei": {
                        "date_col": "T3.date",
                        "time_col": "T3.time",
                        "aggregations": {
                        }
                    },
                },

                # Joins config: auto-generate JOIN ON conditions based on granularity/aggregation
                "joins": [
                    {
                        "left": "T1huawei",
                        "right": "T2huawei",
                        "include_date": True,       # Always include date
                        "include_time": True,       # Only if HOURLY granularity
                        "include_aggregation": False, # Only if aggregation is selected
                        "custom": []                 # Extra static conditions if needed
                    },

                    {
                        "left": "T1huawei",
                        "right": "T3huawei",
                        "include_date": True,       # Always include date
                        "include_time": True,       # Only if HOURLY granularity
                        "include_aggregation": False, # Only if aggregation is selected
                        "custom": []                 # Extra static conditions if needed
                    },

                    {
                        "left": "eric",
                        "right": "huawei",
                        "include_date": True,       # Always include date
                        "include_time": True,       # Only if HOURLY granularity
                        "include_aggregation": True, # Only if aggregation is selected
                        "custom": []                 # Extra static conditions if needed
                    }
                ],

                "sql_template": """
                
WITH HUAWEI_3G_KPI AS (

SELECT
            {T1huawei__select_fields}
            -- T1.cell_name,
            ,T1.`PS_DROP_HUAWEI`,

            T3.`CS_CSSR_HUAWEI`,
            T3.`CALL_BLOCK_RATE_HUAWEI`,
            T3.`CS_DROP`,

            T3.`PS_CSSR_HUAWEI`,
            T3.`DEBIT_DL_HUAWEI`,
            T3.`DEBIT_UL_HUAWEI`,

            T1.CELL_AVAILABILITY_HUAWEI,
            T1.Total_PS_Traffic_GB_HUAWEI,
           
            T2.`TRAFFIC_VOIX_HUAWEI`,
            T2.IRAT_HO_SUCCESS_RATE_HUAWEI,
            T2.HO_SUCCESS_RATE_HUAWEI
        FROM
            (
                -- T1: From hourly_huawei_3g_all_counters_1
                SELECT
                    {huaweiSimple__select_fields}
                    -- cell_name,
                    ,CASE
                        WHEN (COALESCE(SUM(CAST(VS_RAB_NormRel_PS AS UNSIGNED)),0) + COALESCE(SUM(CAST(VS_RAB_AbnormRel_PS AS UNSIGNED)),0) -
                              COALESCE(SUM(CAST(VS_RAB_AbnormRel_PS_PCH AS UNSIGNED)),0) - COALESCE(SUM(CAST(VS_RAB_NormRel_PS_PCH AS UNSIGNED)),0) +
                              COALESCE(SUM(CAST(VS_DCCC_D2P_Succ AS UNSIGNED)),0) + COALESCE(SUM(CAST(VS_DCCC_Succ_F2P AS UNSIGNED)),0)) = 0
                        THEN NULL
                        ELSE ROUND((100 *
                            (COALESCE(SUM(CAST(VS_RAB_AbnormRel_PS AS UNSIGNED)),0) - COALESCE(SUM(CAST(VS_RAB_AbnormRel_PS_PCH AS UNSIGNED)),0) -
                             COALESCE(SUM(CAST(VS_RAB_AbnormRel_PS_D2P AS UNSIGNED)),0) - COALESCE(SUM(CAST(VS_RAB_AbnormRel_PS_F2P AS UNSIGNED)),0))
                        ) /
                        (
                            COALESCE(SUM(CAST(VS_RAB_NormRel_PS AS UNSIGNED)),0) + COALESCE(SUM(CAST(VS_RAB_AbnormRel_PS AS UNSIGNED)),0) -
                            COALESCE(SUM(CAST(VS_RAB_AbnormRel_PS_PCH AS UNSIGNED)),0) - COALESCE(SUM(CAST(VS_RAB_NormRel_PS_PCH AS UNSIGNED)),0) +
                            COALESCE(SUM(CAST(VS_DCCC_D2P_Succ AS UNSIGNED)),0) + COALESCE(SUM(CAST(VS_DCCC_Succ_F2P AS UNSIGNED)),0)
                        ), 3)
                    END AS `PS_DROP_HUAWEI`,
                    100 * (
                        (SUM(CAST(RRC_SuccConnEstab_CallReEst AS UNSIGNED)) + SUM(CAST(RRC_SuccConnEstab_EmgCall AS UNSIGNED)) +
                         SUM(CAST(RRC_SuccConnEstab_OrgConvCall AS UNSIGNED)) + SUM(CAST(RRC_SuccConnEstab_OrgSubCall AS UNSIGNED)) +
                         SUM(CAST(RRC_SuccConnEstab_TmConvCall AS UNSIGNED)) + SUM(CAST(RRC_SuccConnEstab_Unkown AS UNSIGNED))) /
                        NULLIF((SUM(CAST(RRC_AttConnEstab_CallReEst AS UNSIGNED)) + SUM(CAST(RRC_AttConnEstab_EmgCall AS UNSIGNED)) +
                         SUM(CAST(RRC_AttConnEstab_OrgSubCall AS UNSIGNED)) + SUM(CAST(RRC_AttConnEstab_OrgConvCall AS UNSIGNED)) +
                         SUM(CAST(RRC_AttConnEstab_TmConvCall AS UNSIGNED)) + SUM(CAST(RRC_AttConnEstab_Unknown AS UNSIGNED))), 0)
                    ) * (
                        (SUM(CAST(VS_HSPA_RAB_SuccEstab_CS_Conv AS UNSIGNED)) + SUM(CAST(VS_RAB_SuccEstabCS_Conv AS UNSIGNED)) +
                         SUM(CAST(VS_RAB_SuccEstabCS_Str AS UNSIGNED)) + SUM(CAST(VS_RAB_SuccEstabCS_AMR AS UNSIGNED)) +
                         SUM(CAST(VS_RAB_SuccEstabCS_AMRWB AS UNSIGNED))) /
                        NULLIF((SUM(CAST(VS_HSPA_RAB_AttEstab_CS_Conv AS UNSIGNED)) + SUM(CAST(VS_RAB_AttEstabCS_Conv AS UNSIGNED)) +
                         SUM(CAST(VS_RAB_AttEstabCS_Str AS UNSIGNED)) + SUM(CAST(VS_RAB_AttEstab_AMR AS UNSIGNED)) +
                         SUM(CAST(VS_RAB_AttEstabCS_AMRWB AS UNSIGNED))), 0)
                    ) * (
                        1 - (SUM(CAST(VS_RAB_AbnormRel_AMR AS UNSIGNED)) + SUM(CAST(VS_RAB_AbnormRel_AMRWB AS UNSIGNED)) +
                             SUM(CAST(VS_RAB_AbnormRel_CS_HSPA_Conv AS UNSIGNED))) /
                        NULLIF((SUM(CAST(VS_RAB_AbnormRel_AMR AS UNSIGNED)) + SUM(CAST(VS_RAB_AbnormRel_AMRWB AS UNSIGNED)) +
                         SUM(CAST(VS_RAB_AbnormRel_CS_HSPA_Conv AS UNSIGNED)) + SUM(CAST(VS_RAB_NormRel_CS AS UNSIGNED)) +
                         SUM(CAST(VS_RAB_NormRel_CS_HSPA_Conv AS UNSIGNED)) + SUM(CAST(VS_RAB_NormRel_AMRWB AS UNSIGNED))), 0)
                    ) AS `CS_CSSR_HUAWEI`,
                    100 * (
                        (SUM(CAST(VS_RAB_FailEstabCS_DLIUBBand_Cong AS UNSIGNED)) + SUM(CAST(VS_RAB_FailEstabCS_ULIUBBand_Cong AS UNSIGNED)) +
                         SUM(CAST(VS_RAB_FailEstabCS_ULCE_Cong AS UNSIGNED)) + SUM(CAST(VS_RAB_FailEstabCS_DLCE_Cong AS UNSIGNED)) +
                         SUM(CAST(VS_RAB_FailEstabCS_Code_Cong AS UNSIGNED)) + SUM(CAST(VS_RAB_FailEstabCS_ULPower_Cong AS UNSIGNED)) +
                         SUM(CAST(VS_RAB_FailEstabCS_DLPower_Cong AS UNSIGNED))) /
                        NULLIF((SUM(CAST(VS_RAB_AttEstabCS_Conv AS UNSIGNED)) + SUM(CAST(VS_RAB_AttEstabCS_Str AS UNSIGNED))), 0)
                    ) AS `CALL_BLOCK_RATE_HUAWEI`,
                    SUM(CAST(Total_PS_Traffic_GB AS UNSIGNED)) AS Total_PS_Traffic_GB_HUAWEI,
                    (1 - (SUM(CAST(VS_Cell_UnavailTime_Sys AS UNSIGNED))/(sum(60) * 60))) * 100 AS CELL_AVAILABILITY_HUAWEI

                FROM hourly_huawei_3g_all_counters_1 h1
                -- INNER JOIN EPT_3G  ON EPT_3G.`CELL_NAME` = h1.cell_name and `EPT_3G`.`VENDOR` = "HUAWEI"
                WHERE
                    {huaweiSimple__where_clause}
                GROUP BY
                    {huaweiSimple__group_by} 
                    -- , cell_name
            ) AS T1
        INNER JOIN
            (
                -- T2: From hourly_huawei_3g_all_counters_2
                SELECT
                    {huaweiSimple__select_fields}
                    -- cell_name,
                    ,100 * (
                        (SUM(CAST(RRC_SuccConnEstab_OrgInterCall AS double)) + SUM(CAST(RRC_SuccConnEstab_TmItrCall AS UNSIGNED)) +
                         SUM(CAST(RRC_SuccConnEstab_OrgBkgCall AS double)) + SUM(CAST(RRC_SuccConnEstab_TmBkgCall AS UNSIGNED)) +
                         SUM(CAST(RRC_SuccConnEstab_OrgSubCall AS double)) + SUM(CAST(RRC_SuccConnEstab_OrgStrCall AS UNSIGNED)) +
                         SUM(CAST(RRC_SuccConnEstab_TmStrCall AS double)) + SUM(CAST(VS_RRC_AttConnEstab_EDCH AS UNSIGNED)) +
                         SUM(CAST(VS_RRC_AttConnEstab_HSDSCH AS double))) /
                        NULLIF((SUM(CAST(RRC_AttConnEstab_OrgInterCall AS double)) + SUM(CAST(RRC_AttConnEstab_TmInterCall AS UNSIGNED)) +
                         SUM(CAST(RRC_AttConnEstab_TmBkgCall AS double)) + SUM(CAST(RRC_AttConnEstab_OrgBkgCall AS UNSIGNED)) +
                         SUM(CAST(RRC_AttConnEstab_OrgSubCall AS double)) + SUM(CAST(RRC_AttConnEstab_OrgStrCall AS UNSIGNED)) +
                         SUM(CAST(RRC_AttConnEstab_TmStrCall AS double)) + SUM(CAST(VS_RRC_SuccConnEstab_EDCH AS UNSIGNED)) +
                         SUM(CAST(VS_RRC_SuccConnEstab_HSDSCH AS double))), 0)
                    ) * (
                        (SUM(CAST(VS_RAB_SuccEstabPS_Conv AS double)) + SUM(CAST(VS_RAB_SuccEstabPS_Str AS UNSIGNED)) +
                         SUM(CAST(VS_RAB_SuccEstabPS_Int AS double)) + SUM(CAST(VS_RAB_SuccEstabPS_Bkg AS UNSIGNED))) /
                        NULLIF((SUM(CAST(VS_RAB_AttEstabPS_Conv AS double)) + SUM(CAST(VS_RAB_AttEstabPS_Str AS UNSIGNED)) +
                         SUM(CAST(VS_RAB_AttEstabPS_Int AS double)) + SUM(CAST(VS_RAB_AttEstabPS_Bkg AS UNSIGNED))), 0)
                    ) AS `PS_CSSR_HUAWEI`,
                    AVG(CAST(VS_HSDPA_MeanChThroughput AS double)) AS `DEBIT_DL_HUAWEI`,
                    AVG(CAST(VS_HSUPA_MeanChThroughput AS double)) AS `DEBIT_UL_HUAWEI`,
                    SUM(CAST(VS_AMR_Erlang_BestCell AS double)) AS `TRAFFIC_VOIX_HUAWEI`,
                    100.0 * (SUM(CAST(VS_SHO_SuccRLAdd AS double))+SUM(CAST(VS_SHO_SuccRLDel AS UNSIGNED))) / NULLIF((SUM(CAST(VS_SHO_AttRLAdd AS UNSIGNED))+SUM(CAST(VS_SHO_AttRLDel AS UNSIGNED))),0) AS HO_SUCCESS_RATE_HUAWEI,
                    100.0 * SUM(CAST(IRATHO_SuccOutCS AS double)) / NULLIF(SUM(CAST(IRATHO_AttOutCS AS UNSIGNED)), 0) AS IRAT_HO_SUCCESS_RATE_HUAWEI                    

                FROM hourly_huawei_3g_all_counters_2 h2
                -- INNER JOIN EPT_3G ON EPT_3G.`CELL_NAME` = h2.cell_name and `EPT_3G`.`VENDOR` = "HUAWEI"
                WHERE
                    {huaweiSimple__where_clause}
                GROUP BY
                    {huaweiSimple__group_by}
                    -- , cell_name
            ) AS T2
        ON  
            {T1huawei__T2huawei__join_on}
            
            -- T1.cell_name = T2.cell_name
        INNER JOIN  (
            SELECT 
  {huaweiSimple__select_fields}
 -- DATE_FORMAT(date, '%Y-%m') DATE
-- ,substring_index(substring_index(daily_arcep_huawei_3g.UCELL,'=',-3),',',1)
  -- ,NE_Name
  
  ,100 *
  (
    (
      SUM(CAST(NSRRCCallReest AS UNSIGNED)) +
      SUM(CAST(NSRRCCEmgCall AS UNSIGNED)) +
      SUM(CAST(OrgConvCall AS UNSIGNED)) +
      SUM(CAST(OrgSubCall AS UNSIGNED)) +
      SUM(CAST(TmConvCall AS UNSIGNED)) +
      SUM(CAST(Unknown AS UNSIGNED))
    )
    /
    NULLIF(
      SUM(CAST(AttCallReEst AS UNSIGNED)) +
      SUM(CAST(AttEmgCall AS UNSIGNED)) +
      SUM(CAST(AtOrgConvCall AS UNSIGNED)) +
      SUM(CAST(AtOrgSubCall AS UNSIGNED)) +
      SUM(CAST(AtTmConvCall AS UNSIGNED)) +
      SUM(CAST(AtUnknown AS UNSIGNED))
    , 0)
  )
  *
  (
    (
      SUM(CAST(SucHsCsConv AS UNSIGNED)) +
      SUM(CAST(SucRabEst AS UNSIGNED)) +
      SUM(CAST(SucCSStr AS UNSIGNED)) +
      SUM(CAST(SucAMR AS UNSIGNED)) +
      SUM(CAST(SucAMRWB AS UNSIGNED))
    )
    /
    NULLIF(
      SUM(CAST(AttHsCsConv AS UNSIGNED)) +
      SUM(CAST(CSConvRabEstReq AS UNSIGNED)) +
      SUM(CAST(CSStrRabEstReq AS UNSIGNED)) +
      SUM(CAST(AttAMR AS UNSIGNED)) +
      SUM(CAST(AttAMRWB AS UNSIGNED))
    , 0)
  )
  *
  (
    1 -
    (
      (SUM(CAST(NCsAbnAMR AS UNSIGNED)) +
       SUM(CAST(AtAbnAMRWB AS UNSIGNED)) +
       SUM(CAST(AtCsAbnHS AS UNSIGNED))
      )
      /
      NULLIF(
        SUM(CAST(NCsRabNor AS UNSIGNED)) +
        SUM(CAST(NCsAbnAMR AS UNSIGNED)) +
        (
          SUM(CAST(AtAbnAMRWB AS UNSIGNED)) +
          SUM(CAST(AtCsAbnHS AS UNSIGNED)) +
          SUM(CAST(AtCsNorHS AS UNSIGNED)) +
          SUM(CAST(AtNorAMRWB AS UNSIGNED))
        )
      , 0)
    )
  ) AS CS_CSSR_HUAWEI,
  100 *
  (
    (SUM(CAST(AtAbnAMRWB AS UNSIGNED)) + SUM(CAST(AtCsAbnHS AS UNSIGNED)))
    /
    NULLIF(
      (SUM(CAST(SucHsCsConv AS UNSIGNED)) + SUM(CAST(SucAMR AS UNSIGNED)) + SUM(CAST(SucAMRWB AS UNSIGNED))
      ), 0)
  ) AS CS_DROP ,
  100 *
  (
    (
      SUM(CAST(FailRabEstDlIub AS UNSIGNED)) 
      + SUM(CAST(FailRabestUlIub AS UNSIGNED)) 
      + SUM(CAST(FailRabEstUlCECong AS UNSIGNED)) 
      + SUM(CAST(FailRabEstDlCECong AS UNSIGNED)) 
      + SUM(CAST(FailRabEstCodeCong AS UNSIGNED)) 
      + SUM(CAST(FailRabEstUlPwrCong AS UNSIGNED)) 
      + SUM(CAST(FailRabEstDlPwrCong AS UNSIGNED))
    )
    /
    NULLIF(
      (SUM(CAST(CSConvRabEstReq AS UNSIGNED)) + SUM(CAST(CSStrRabEstReq AS UNSIGNED))
      ), 0)
  ) as CALL_BLOCK_RATE_HUAWEI,
 100 *
  (
    (
      (SUM(CAST(OrgInterCall AS UNSIGNED)) 
       + SUM(CAST(TmItrCall AS UNSIGNED)) 
       + SUM(CAST(OrgBkgCall AS UNSIGNED)) 
       + SUM(CAST(TmBkgCall AS UNSIGNED)) 
       + SUM(CAST(TmStrCall AS UNSIGNED)) 
       + SUM(CAST(OrgStrCall AS UNSIGNED)) 
       + SUM(CAST(OrgSubCall AS UNSIGNED)) 
       + SUM(CAST(AttHSDSCH AS UNSIGNED)) 
       + SUM(CAST(AttEDCH AS UNSIGNED))
      )
      /
      NULLIF(
        (SUM(CAST(AttOrgItrCall AS UNSIGNED)) 
         + SUM(CAST(AttTmItrCall AS UNSIGNED)) 
         + SUM(CAST(AttOrgBkgCall AS UNSIGNED)) 
         + SUM(CAST(AttTmBkgCall AS UNSIGNED)) 
         + SUM(CAST(AttOrgStrCall AS UNSIGNED)) 
         + SUM(CAST(AttTmStrCall AS UNSIGNED)) 
         + SUM(CAST(AtOrgSubCall AS UNSIGNED)) 
         + SUM(CAST(SuccEDCH AS UNSIGNED)) 
         + SUM(CAST(SuccHSDSCH AS UNSIGNED))
        ), 0)
    )
    *
    (
      (SUM(CAST(SucRabPSConv AS UNSIGNED)) 
       + SUM(CAST(SucRabPSStr AS UNSIGNED)) 
       + SUM(CAST(SucRabPSInter AS UNSIGNED)) 
       + SUM(CAST(SucRabPSBkg AS UNSIGNED)))
      /
      NULLIF(
        (SUM(CAST(AttRabPSConv AS UNSIGNED)) 
         + SUM(CAST(AttRabPSStr AS UNSIGNED)) 
         + SUM(CAST(AttRabPSInter AS UNSIGNED)) 
         + SUM(CAST(AttRabPSBkg AS UNSIGNED))
        ), 0)
    )
  ) AS PS_CSSR_HUAWEI,
  100 *
  (
    (SUM(CAST(AbnPSRabPCH AS UNSIGNED)) - SUM(CAST(AbnPSRabD2P AS UNSIGNED)) - SUM(CAST(CDRF2P AS UNSIGNED)))
    /
    NULLIF(
      (SUM(CAST(AbnPSRabPCH AS UNSIGNED)) - SUM(CAST(NormPSRabPCH AS UNSIGNED)) 
       + SUM(CAST(SuccD2P AS UNSIGNED)) + SUM(CAST(SuccFach AS UNSIGNED))
      ), 0)
  ) AS PS_DROP_HUAWEI,
      avg(CAST(DLHSTput AS DECIMAL)) AS DEBIT_DL_HUAWEI,
    avg(CAST(ULHSTput AS DECIMAL)) AS DEBIT_UL_HUAWEI
  --  round(AVG(CAST(NULLIF(DLHSTput,'') AS DECIMAL(10,2))),4) AS DLHSTput_v,

   --  round(AVG(CAST(NULLIF(ULHSTput,'') AS DECIMAL(10,2))),3) AS ULHSTput_v
FROM hourly_arcep_huawei_3g 
WHERE
{huaweiSimple__where_clause}
-- AND NE_Name='PKMBSC03'

 -- AND substring_index(substring_index(daily_arcep_huawei_3g.UCELL,'=',-3),',',1) NOT IN 
--  (
--  select cellule  from weekly_huawei_3g_disponibilite
--         WHERE DATE='2025-06-15' AND ( availibility_3g=0 OR availibility_3g='')
--  )
-- group by DATE_FORMAT(date, '%Y-%m')

GROUP BY 
   {huaweiSimple__group_by}
) AS T3
         ON
            {T1huawei__T3huawei__join_on}

),

ERICSSON_3G_KPI AS (
  SELECT
  {eric__select_fields}
  
  -- , CELL_NAME,
  -- SITE_NAME,
  -- CONTROLLER_NAME,

  -- CS_CSSR (Circuit-Switched Call Setup Success Rate)
  ,100.0 *
  (COALESCE(SUM(CAST(ReqCsSucc AS DOUBLE)), 0) /
   NULLIF(COALESCE(SUM(CAST(ReqCs AS DOUBLE)), 0), 0)) *
  (COALESCE(SUM(CAST(pmNoRabEstablishSuccessSpeech AS DOUBLE)), 0) /
   NULLIF(COALESCE(SUM(CAST(pmNoRabEstablishAttemptSpeech AS DOUBLE)), 0) -
          COALESCE(SUM(CAST(pmNoDirRetryAtt AS DOUBLE)), 0), 0)) *
  (COALESCE(SUM(CAST(pmNoNormalRabReleaseSpeech AS DOUBLE)), 0) /
   NULLIF(COALESCE(SUM(CAST(pmNoNormalRabReleaseSpeech AS DOUBLE)), 0) +
          COALESCE(SUM(CAST(pmNoSystemRabReleaseSpeech AS DOUBLE)), 0), 0)) AS CS_CSSR,

  -- CS_DROP (Circuit-Switched Drop Rate)
  100.0 * (COALESCE(SUM(CAST(pmNoSystemRabReleaseSpeech AS DOUBLE)), 0) /
   NULLIF(COALESCE(SUM(CAST(pmNoSystemRabReleaseSpeech AS DOUBLE)), 0) +
          COALESCE(SUM(CAST(pmNoNormalRabReleaseSpeech AS DOUBLE)), 0), 0)) AS CS_DROP,

  -- CS_CBR (Circuit-Switched Call Block Rate)
  100.0 * (COALESCE(SUM(CAST(pmNoRabEstBlockTnSpeech AS DOUBLE)), 0) +
           COALESCE(SUM(CAST(pmNoRabEstBlockNodeSpeechBest AS DOUBLE)), 0)) /
  NULLIF(COALESCE(SUM(CAST(pmNoRabEstablishAttemptSpeech AS DOUBLE)), 0) -
         COALESCE(SUM(CAST(pmNoRabEstablishSuccessSpeech AS DOUBLE)), 0), 0) AS CS_CBR,

  -- PS_CSSR (Packet-Switched Call Setup Success Rate)
  100.0 *
  (COALESCE(SUM(CAST(ReqPsSucc AS DOUBLE)), 0) /
   NULLIF(COALESCE(SUM(CAST(ReqPs AS DOUBLE)), 0) -
          COALESCE(SUM(CAST(pmNoLoadSharingRrcConnPs AS DOUBLE)), 0), 0)) *
  (COALESCE(SUM(CAST(pmNoRabEstablishSuccessPacketInteractive AS DOUBLE)), 0) /
   NULLIF(COALESCE(SUM(CAST(pmNoRabEstablishAttemptPacketInteractive AS DOUBLE)), 0), 0)) AS PS_CSSR,

  -- PS_DROP (Packet-Switched Drop Rate)
  100.0 * (COALESCE(SUM(CAST(pmNoSystemRabReleasePacket AS DOUBLE)), 0) /
   NULLIF(COALESCE(SUM(CAST(pmNoNormalRabReleasePacket AS DOUBLE)), 0) +
          COALESCE(SUM(CAST(pmPsIntHsToFachSucc AS DOUBLE)), 0) +
          COALESCE(SUM(CAST(pmNoSuccRbReconfOrigPsIntDch AS DOUBLE)), 0), 0)) AS PS_DROP,

  -- DEBIT_DL (Download Throughput in kbps)
  COALESCE(SUM(CAST(pmSumHsDlRlcUserPacketThp AS DOUBLE)), 0) /
  NULLIF(COALESCE(SUM(CAST(pmSamplesHsDlRlcUserPacketThp AS DOUBLE)), 0), 0) AS DEBIT_DL,

  -- DEBIT_UL (Upload Throughput in kbps)
  COALESCE(SUM(CAST(pmSumEulRlcUserPacketThp AS DOUBLE)), 0) /
  NULLIF(COALESCE(SUM(CAST(pmSamplesEulRlcUserPacketThp AS DOUBLE)), 0), 0) AS DEBIT_UL,

   SUM(CAST(pmSumEulRlcTotPacketThp AS DOUBLE)) / 
  NULLIF(SUM(CAST(pmSamplesEulRlcTotPacketThp AS DOUBLE)), 0) AS DEBIT_UL_CELL_ARCEP,

  -- CELL_AVAILABILITY (Cell Availability Rate)
  100.0 * (COALESCE(SUM(CAST(PERIOD_DURATION AS DOUBLE)), 0) * 60 -
           COALESCE(SUM(CAST(pmCellDowntimeAuto AS DOUBLE)), 0)) /
  NULLIF(COALESCE(SUM(CAST(PERIOD_DURATION AS DOUBLE)), 0) * 60, 0) AS CELL_AVAILABILITY,

  -- RL_ADD_SUCCESS_RATE (Radio Link Addition Success Rate)
  100.0 * COALESCE(SUM(CAST(pmRlAddSuccessBestCellSpeech AS DOUBLE)), 0) /
  NULLIF(COALESCE(SUM(CAST(pmRlAddAttemptsBestCellSpeech AS DOUBLE)), 0), 0) AS RL_ADD_SUCCESS_RATE,

  -- IRAT_HO_SUCCESS_RATE (Inter-RAT Handover Success Rate)
  100.0 * COALESCE(SUM(CAST(pmNoSuccessOutIratHoSpeech AS DOUBLE)), 0) /
  NULLIF(COALESCE(SUM(CAST(pmNoAttOutIratHoSpeech AS DOUBLE)), 0), 0) AS IRAT_HO_SUCCESS_RATE,

  COALESCE(SUM(CAST(TRAFFIC_VOIX AS DOUBLE)), 0)TRAFFIC_VOIX,
  COALESCE(SUM(CAST(TRAFFIC_DATA_GB AS DOUBLE)), 0) TRAFFIC_DATA_GB

FROM
  hourly_ericsson_arcep_3g_counters e

WHERE
{eric__where_clause}

GROUP BY
 {eric__group_by}
)

SELECT
 {combined__select_fields}

  -- Huawei KPIs
 , h.PS_DROP_HUAWEI,
  h.CS_CSSR_HUAWEI,
  h.CALL_BLOCK_RATE_HUAWEI AS CS_CBR_HUAWEI,
  h.CELL_AVAILABILITY_HUAWEI,
  h.Total_PS_Traffic_GB_HUAWEI,
  h.PS_CSSR_HUAWEI,
  h.DEBIT_DL_HUAWEI,
  h.DEBIT_UL_HUAWEI,
  h.TRAFFIC_VOIX_HUAWEI,
  h.IRAT_HO_SUCCESS_RATE_HUAWEI,
  h.HO_SUCCESS_RATE_HUAWEI AS RL_ADD_SUCCESS_RATE_HUAWEI,

  -- Ericsson KPIs
  e.CS_CSSR AS CS_CSSR_ERICSSON,
  e.CS_DROP AS CS_DROP_ERICSSON,
  e.CS_CBR AS CS_CBR_ERICSSON,
  e.PS_CSSR AS PS_CSSR_ERICSSON,
  e.PS_DROP AS PS_DROP_ERICSSON,
  e.DEBIT_DL AS DEBIT_DL_ERICSSON,
  e.DEBIT_UL AS DEBIT_UL_USER_ERICSSON,
  e.DEBIT_UL_CELL_ARCEP AS DEBIT_UL_ERICSSON,
  e.CELL_AVAILABILITY AS CELL_AVAILABILITY_ERICSSON,
  e.RL_ADD_SUCCESS_RATE AS HO_SUCESS_RATE,
  e.IRAT_HO_SUCCESS_RATE AS IRAT_HO_SUCCESS_RATE_ERICSSON,
  e.TRAFFIC_DATA_GB TRAFFIC_DATA_GB_ERICSSON,
  e.TRAFFIC_VOIX TRAFFIC_VOIX_ERICSSON, 

  -- Global Network KPIs (weighted: 22% Huawei, 78% Ericsson)
  (COALESCE(h.CS_CSSR_HUAWEI, 0) * 0.22) + (COALESCE(e.CS_CSSR, 0) * 0.78) AS CS_CSSR_GLOBAL,
  (COALESCE(h.PS_DROP_HUAWEI, 0) * 0.22) + (COALESCE(e.PS_DROP, 0) * 0.78) AS PS_DROP_GLOBAL,
  (COALESCE(h.CALL_BLOCK_RATE_HUAWEI, 0) * 0.22) + (COALESCE(e.CS_CBR, 0) * 0.78) AS CS_CBR_GLOBAL,
  (COALESCE(h.PS_CSSR_HUAWEI, 0) * 0.22) + (COALESCE(e.PS_CSSR, 0) * 0.78) AS PS_CSSR_GLOBAL,
  (COALESCE(h.DEBIT_DL_HUAWEI, 0) * 0.22) + (COALESCE(e.DEBIT_DL, 0) * 0.78) AS DEBIT_DL_GLOBAL,
  (COALESCE(h.DEBIT_UL_HUAWEI, 0) * 0.22) + (COALESCE(e.DEBIT_UL_CELL_ARCEP, 0) * 0.78) AS DEBIT_UL_GLOBAL,
  (COALESCE(h.CELL_AVAILABILITY_HUAWEI, 0) * 0.22) + (COALESCE(e.CELL_AVAILABILITY, 0) * 0.78) AS CELL_AVAILABILITY_GLOBAL,
  (COALESCE(h.HO_SUCCESS_RATE_HUAWEI, 0) * 0.22) + (COALESCE(e.RL_ADD_SUCCESS_RATE, 0) * 0.78) AS RL_ADD_SUCCESS_RATE_GLOBAL,
  (COALESCE(h.IRAT_HO_SUCCESS_RATE_HUAWEI, 0) * 0.22) + (COALESCE(e.IRAT_HO_SUCCESS_RATE, 0) * 0.78) AS IRAT_HO_SUCCESS_RATE_GLOBAL,
  e.TRAFFIC_VOIX + h.TRAFFIC_VOIX_HUAWEI AS TRAFFIC_VOIX_NETWORK,
  e.TRAFFIC_DATA_GB + H.Total_PS_Traffic_GB_HUAWEI AS TRAFFIC_DATA_NETWORK


FROM HUAWEI_3G_KPI h
left JOIN ERICSSON_3G_KPI e
 ON {eric__huawei__join_on}

                """
            }
        }

        
    },


    

    "GENERAL": {
        "NETWORK_HEALTH": {
            "description": "General network health across all technologies",
            "date_time_filters": ["start_date", "end_date"],
            "parameters": ["technology"],
            "aggregation_levels": {
                "technology": "m.technology",
                "vendor": "m.vendor",
                "controller": "m.controller"
            },
            "time_granularities": {
                "DAILY": {"is_available": True, "date_field": "m.date", "time_field": None},
                "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(m.date, 1)", "time_field": None},
                "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(m.date, '%Y-%m')", "time_field": None}
            },
            "kpi_fields": {
                "Total_Sites": "COUNT(DISTINCT site_name)",
                "Avg_Availability": "AVG(availability)"
            },
            "sql_template": """
                SELECT {fields}
                FROM network_health_summary m
                WHERE {where_conditions}
                {group_by}
                ORDER BY {order_by}
            """
        }
    },

    "2G_NETWORK": {

        "COMBINED": {

            "KPI_MONITORING": {
                "query_type": "multi_cte",
                "description": "2G Network KPI - Combined Vendors",
                "date_time_filters": ["start_date", "end_date", "multiple_date"],

                # Subqueries configuration
                "subqueries": [
                    {
                        "name": "HUAWEI2G_KPI",
                        "table": "daily_arcep_huawei_2g",
                        "alias": "h",
                        "time_granularities": {
                            "HOURLY": {"is_available": True, "date_field": "DATE", "time_field": "TIME"},
                            "DAILY": {"is_available": True, "date_field": "DATE", "time_field": None},
                            "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(DATE, 1)", "time_field": None},
                            "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(DATE, '%Y-%m')", "time_field": None}
                        },
                        "aggregation_levels": {},
                        "kpi_fields": {
                            "CSSR_HUAWEI": """100 * (SUM(Suc_SDCCH_Seiz) / NULLIF(SUM(SDCCH_Seiz_Req), 0)) *
                                (1 - (SUM(SDCCHDrop) / NULLIF(SUM(Suc_SDCCH_Seiz), 0))) *
                                (SUM(Suc_TCH_Seiz_SG) + SUM(Suc_TCH_Seiz_TrafChann) + SUM(Suc_TCH_Seiz_handTrafChan)) /
                                NULLIF((SUM(TCH_Seizure_Req_SC) + SUM(TCH_Seiz_Req_TrafChan) + SUM(TCH_Seiz_Req_HandTrafChan)), 0) *
                                (1 - (SUM(TCHDrop_Nume_ARCEP) / NULLIF(SUM(TCHDrop_Deno_ARCEP), 0)))""",
                            "CDR_HUAWEI": "(100 * SUM(TCHDrop_Nume_ARCEP) / NULLIF(SUM(TCHDrop_Deno_ARCEP), 0))",
                            "CBR_HUAWEI": "100 * (SUM(TCHCONG_Nume_ARCEP) / NULLIF(SUM(TCHCONG_Deno_ARCEP), 0))"
                        },
                        "sql_template": """
                            SELECT {fields}
                            FROM {table} {alias}
                            WHERE {where_conditions}
                            {group_by}
                        """
                    },
                    {
                        "name": "ERICSSON_2GKPI",
                        "table": "daily_arcep_ericsson_2g",
                        "alias": "e",
                        "time_granularities": {
                            "HOURLY": {"is_available": True, "date_field": "DATE", "time_field": "TIME"},
                            "DAILY": {"is_available": True, "date_field": "DATE_id", "time_field": None},
                            "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(DATE_id, 1)", "time_field": None},
                            "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(DATE_id, '%Y-%m')", "time_field": None}
                        },
                        "aggregation_levels": {},
                        "kpi_fields": {
                            "CSSR_ERICSSON": """(100 * (1 - (SUM(CCONGS) / NULLIF(SUM(CCALLS), 0)))) *
                                (1 - ((SUM(CNDROP) - SUM(CNRELCONG)) / NULLIF(SUM(CMSESTAB), 0))) *
                                (SUM(TCASSALL) / NULLIF(SUM(TASSALL), 0)) *
                                (1 - (((SUM(TFNDROP) + SUM(THNDROP) + SUM(TFNDROPSUB) + SUM(THNDROPSUB)) /
                                NULLIF((SUM(TFCASSALL) + SUM(THCASSALL) + SUM(TFCASSALLSUB) + SUM(THCASSALLSUB)), 0) * 100) / 100))""",
                            "CBR_ERICSSON": "(100 * (SUM(CNRELCONG) + SUM(THNRELCONG) + SUM(TFNRELCONG)) / NULLIF(SUM(TASSALL), 0))",
                            "CDR_ERICSSON": "(100 * (SUM(TFNDROP) + SUM(THNDROP) + SUM(TFNDROPSUB) + SUM(THNDROPSUB)) / NULLIF((SUM(TFCASSALL) + SUM(THCASSALL) + SUM(TFCASSALLSUB) + SUM(THCASSALLSUB)), 0))"
                        },
                        "sql_template": """
                            SELECT {fields}
                            FROM {table} {alias}
                            WHERE {where_conditions}
                            {group_by}
                        """
                    }
                ],

                # How to join the CTEs
                "cte_joins": [
                    {
                        "type": "LEFT JOIN",
                        "left": "HUAWEI2G_KPI",
                        "right": "ERICSSON_2GKPI",
                        "on": ["HUAWEI2G_KPI.date = ERICSSON_2GKPI.DATE"]
                    }
                ],

                # Outer query configuration (what the client controls)
                "time_granularities": {
                     "HOURLY": {"is_available": True, "date_field": "DATE", "time_field": "TIME"},
                    "DAILY": {"is_available": True, "date_field": "HUAWEI2G_KPI.date", "time_field": None},
                    "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(HUAWEI2G_KPI.date, 1)", "time_field": None},
                    "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(HUAWEI2G_KPI.date, '%Y-%m')", "time_field": None}
                },

                "aggregation_levels": {
                },

                "kpi_fields": {
                    "CSSR_HUAWEI": "HUAWEI2G_KPI.CSSR_HUAWEI",
                    "CDR_HUAWEI": "HUAWEI2G_KPI.CDR_HUAWEI",
                    "CBR_HUAWEI": "HUAWEI2G_KPI.CBR_HUAWEI",
                    "CSSR_ERICSSON": "ERICSSON_2GKPI.CSSR_ERICSSON",
                    "CDR_ERICSSON": "ERICSSON_2GKPI.CDR_ERICSSON",
                    "CBR_ERICSSON": "ERICSSON_2GKPI.CBR_ERICSSON",
                    "CSSR_NETWORK": "(HUAWEI2G_KPI.CSSR_HUAWEI * 0.22) + (ERICSSON_2GKPI.CSSR_ERICSSON * 0.78)",
                    "CDR_NETWORK": "(HUAWEI2G_KPI.CDR_HUAWEI * 0.22) + (ERICSSON_2GKPI.CDR_ERICSSON * 0.78)",
                    "CBR_NETWORK": "(HUAWEI2G_KPI.CBR_HUAWEI * 0.22) + (ERICSSON_2GKPI.CBR_ERICSSON * 0.78)"
                },

                "sql_template": """
                    SELECT {fields}
                    FROM {cte_joins}
                    {group_by}
                    {order_by}
                """
            }

        },

        "COMBINED_2": {

            "KPI_MONITORING": {
                "query_type": "multi_cte",
                "description": "2G Network KPI - Combined Vendors (Hourly with All KPIs)",
                "date_time_filters": ["start_date", "end_date", "multiple_date", "start_hour", "end_hour", "multiple_hour"],

                # Subqueries configuration
                "subqueries": [
                    {
                        "name": "HUAWEI_2G_KPI",
                        "table": "hourly_huawei_2g_all_counters",
                        "alias": "h",
                        "time_granularities": {
                            "HOURLY": {"is_available": True, "date_field": "date", "time_field": "TIME_FORMAT(time, '%H:%i:%s')"},
                            "DAILY": {"is_available": True, "date_field": "date", "time_field": None},
                            "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(date, 1)", "time_field": None},
                            "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(date, '%Y-%m')", "time_field": None}
                        },
                        "aggregation_levels": {},
                        "kpi_fields": {
                            "CSSR_HUAWEI": """CASE
                                WHEN COALESCE(SUM(CAST(CELL_KPI_SD_REQ AS DOUBLE)), 0) = 0
                                  OR COALESCE(SUM(CAST(CELL_KPI_SD_SUCC AS DOUBLE)), 0) = 0
                                  OR (COALESCE(SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE)), 0)
                                       + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
                                       + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0)) = 0
                                  OR (COALESCE(SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE)), 0)
                                       + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE)), 0)
                                       + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)), 0)) = 0
                                THEN 0
                                ELSE
                                  100.0
                                  * (SUM(CAST(CELL_KPI_SD_SUCC AS DOUBLE)) / SUM(CAST(CELL_KPI_SD_REQ AS DOUBLE)))
                                  * (1.0 - (SUM(CAST(CELL_SD_CALL_DROPS AS DOUBLE)) / SUM(CAST(CELL_KPI_SD_SUCC AS DOUBLE))))
                                  * ((SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE))
                                       + SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE))
                                       + SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)))
                                      / (SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE))
                                          + SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE))
                                          + SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE))))
                                  * (1.0 - ((SUM(CAST(CELL_KPI_TCH_DROPS_SIG AS DOUBLE))
                                              + SUM(CAST(CELL_TRAF_CH_CALL_DROPS AS DOUBLE))
                                              + SUM(CAST(CELL_KPI_TCH_HO_DROPS_TRAF AS DOUBLE)))
                                            / (SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE))
                                                + SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE))
                                                + SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)))))
                              END""",
                            "CDR_HUAWEI": """100.0 * (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_DROPS_SIG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_TRAF_CH_CALL_DROPS AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_DROPS_TRAF AS DOUBLE)), 0)
                              ) / (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)), 0)
                                + 1e-7
                              )""",
                            "CBR_HUAWEI": """100.0 * (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_CONG_SIG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_CONG_TRAF AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_CONGEST_TRAF AS DOUBLE)), 0)
                              ) / (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0)
                                + 1e-7
                              )""",
                            "TCH_CONGESTION_RATE_HUAWEI": """100.0 * (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_CONG_TRAF AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_CONGEST_TRAF AS DOUBLE)), 0)
                              ) / (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0)
                                + 1e-7
                              )""",
                            "SDCCH_CONGESTION_RATE_HUAWEI": """100.0 * COALESCE(SUM(CAST(CELL_KPI_SD_CONGEST AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(CELL_KPI_SD_REQ AS DOUBLE)), 0) + 1e-7)""",
                            "SDCCH_DROP_RATE_HUAWEI": """100.0 * COALESCE(SUM(CAST(CELL_SD_CALL_DROPS AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(CELL_IMM_ASS_SUCC_SD AS DOUBLE)), 0) + 1e-7)""",
                            "TCH_ASSIGNMENT_SUCCESS_RATE_HUAWEI": """100.0 * (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)), 0)
                              ) / (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0)
                                + 1e-7
                              )""",
                            "SDCCH_TRAFFIC_HUAWEI": """COALESCE(SUM(CAST(CELL_KPI_SD_TRAF_ERL AS DOUBLE)), 0)""",
                            "CELL_AVAILABILITY_RATE_HUAWEI": """100.0 * COALESCE(SUM(CAST(CELL_KPI_TCH_AVAIL_NUM AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(CELL_KPI_TCH_CFG_NUM AS DOUBLE)), 0) + 1e-7)"""
                        },
                        "sql_template": """
                            SELECT {fields}
                            FROM {table} {alias}
                            WHERE {where_conditions}
                            {group_by}
                        """
                    },
                    {
                        "name": "ERICSSON_2G_KPI",
                        "table": "hourly_ericsson_arcep_2g_counters",
                        "alias": "e",
                        "time_granularities": {
                            "HOURLY": {"is_available": True, "date_field": "DATE", "time_field": "TIME"},
                            "DAILY": {"is_available": True, "date_field": "DATE", "time_field": None},
                            "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(DATE, 1)", "time_field": None},
                            "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(DATE, '%Y-%m')", "time_field": None}
                        },
                        "aggregation_levels": {},
                        "kpi_fields": {
                            "CSSR_ERICSSON": """CASE
                                WHEN COALESCE(SUM(CAST(CCALLS AS DOUBLE)), 0) = 0
                                  OR COALESCE(SUM(CAST(CMSESTAB AS DOUBLE)), 0) = 0
                                  OR COALESCE(SUM(CAST(TASSALL AS DOUBLE)), 0) = 0
                                  OR (COALESCE(SUM(CAST(TFCASSALL AS DOUBLE)), 0)
                                       + COALESCE(SUM(CAST(THCASSALL AS DOUBLE)), 0)
                                       + COALESCE(SUM(CAST(TFCASSALLSUB AS DOUBLE)), 0)
                                       + COALESCE(SUM(CAST(THCASSALLSUB AS DOUBLE)), 0)) = 0
                                THEN 0
                                ELSE
                                  100.0
                                  * (1.0 - (SUM(CAST(CCONGS AS DOUBLE)) / SUM(CAST(CCALLS AS DOUBLE))))
                                  * (1.0 - ((SUM(CAST(CNDROP AS DOUBLE)) - SUM(CAST(CNRELCONG AS DOUBLE))) / SUM(CAST(CMSESTAB AS DOUBLE))))
                                  * (SUM(CAST(TCASSALL AS DOUBLE)) / SUM(CAST(TASSALL AS DOUBLE)))
                                  * (1.0 - ((SUM(CAST(TFNDROP AS DOUBLE)) + SUM(CAST(THNDROP AS DOUBLE))
                                              + SUM(CAST(TFNDROPSUB AS DOUBLE)) + SUM(CAST(THNDROPSUB AS DOUBLE)))
                                            / (SUM(CAST(TFCASSALL AS DOUBLE)) + SUM(CAST(THCASSALL AS DOUBLE))
                                                + SUM(CAST(TFCASSALLSUB AS DOUBLE)) + SUM(CAST(THCASSALLSUB AS DOUBLE)))))
                              END""",
                            "CBR_ERICSSON": """100.0 * (
                                COALESCE(SUM(CAST(CNRELCONG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(TFNRELCONG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(TFNRELCONGSUB AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(THNRELCONG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(THNRELCONGSUB AS DOUBLE)), 0)
                              ) / (COALESCE(SUM(CAST(TASSALL AS DOUBLE)), 0) + 1e-7)""",
                            "CDR_ERICSSON": """100.0 * (
                                COALESCE(SUM(CAST(THNDROP AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(THNDROPSUB AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(TFNDROP AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(TFNDROPSUB AS DOUBLE)), 0)
                              ) / (
                                COALESCE(SUM(CAST(TFCASSALL AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(TFCASSALLSUB AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(THCASSALL AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(THCASSALLSUB AS DOUBLE)), 0)
                                + 1e-7
                              )""",
                            "CELL_AVAILABILITY_RATE_ERICSSON": """100.0 - (100.0 * COALESCE(SUM(CAST(TDWNACC AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(TDWNSCAN AS DOUBLE)), 0) + 1e-7))""",
                            "TCH_AVAILABILITY_RATE_ERICSSON": """100.0 * (COALESCE(SUM(CAST(TAVAACC AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(TAVASCAN AS DOUBLE)), 0) + 1e-7)) /
                              (COALESCE(AVG(CAST(AVG_TNUCHCNT AS DOUBLE)), 1))""",
                            "DOWNTIME_MANUAL": """COALESCE(SUM(CAST(HDWNACC AS DOUBLE)), 0)""",
                            "TCH_CONGESTION_RATE_ERICSSON": """100.0 * (
                                COALESCE(SUM(CAST(CNRELCONG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(TFNRELCONG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(TFNRELCONGSUB AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(THNRELCONG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(THNRELCONGSUB AS DOUBLE)), 0)
                              ) / (COALESCE(SUM(CAST(TASSALL AS DOUBLE)), 0) + 1e-7)""",
                            "SDCCH_DROP_RATE_ERICSSON": """100.0 * COALESCE(SUM(CAST(CNDROP AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(CMSESTAB AS DOUBLE)), 0) + 1e-7)""",
                            "SDCCH_TRAFFIC_ERICSSON": """COALESCE(SUM(CAST(CTRALACC AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(CNSCAN AS DOUBLE)), 0) + 1e-7)""",
                            "SDCCH_BLOCKING_RATE_ERICSSON": """100.0 * COALESCE(SUM(CAST(CCONGS AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(CCALLS AS DOUBLE)), 0) + 1e-7)""",
                            "SDCCH_CONGESTION_RATE_ERICSSON": """100.0 * (COALESCE(SUM(CAST(CCONGS AS DOUBLE)), 0) + COALESCE(SUM(CAST(CCONGSSUB AS DOUBLE)), 0)) /
                              (COALESCE(SUM(CAST(CCALLS AS DOUBLE)), 0) + 1e-7)"""
                        },
                        "sql_template": """
                            SELECT {fields}
                            FROM {table} {alias}
                            WHERE {where_conditions}
                            {group_by}
                        """
                    }
                ],

                # How to join the CTEs
                "cte_joins": [
                    {
                        "type": "LEFT JOIN",
                        "left": "HUAWEI_2G_KPI",
                        "right": "ERICSSON_2G_KPI",
                        "on": ["HUAWEI_2G_KPI.date = ERICSSON_2G_KPI.DATE", "HUAWEI_2G_KPI.time = ERICSSON_2G_KPI.TIME"]
                    }
                ],

                # Outer query configuration (what the client controls)
                "time_granularities": {
                    "HOURLY": {"is_available": True, "date_field": "HUAWEI_2G_KPI.date", "time_field": "HUAWEI_2G_KPI.time"},
                    "DAILY": {"is_available": True, "date_field": "HUAWEI_2G_KPI.date", "time_field": None},
                    "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(HUAWEI_2G_KPI.date, 1)", "time_field": None},
                    "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(HUAWEI_2G_KPI.date, '%Y-%m')", "time_field": None}
                },

                "aggregation_levels": {
                },

                "kpi_fields": {
                    "CSSR_HUAWEI": "HUAWEI_2G_KPI.CSSR_HUAWEI",
                    "CDR_HUAWEI": "HUAWEI_2G_KPI.CDR_HUAWEI",
                    "CBR_HUAWEI": "HUAWEI_2G_KPI.CBR_HUAWEI",
                    "TCH_CONGESTION_RATE_HUAWEI": "HUAWEI_2G_KPI.TCH_CONGESTION_RATE_HUAWEI",
                    "SDCCH_CONGESTION_RATE_HUAWEI": "HUAWEI_2G_KPI.SDCCH_CONGESTION_RATE_HUAWEI",
                    "SDCCH_DROP_RATE_HUAWEI": "HUAWEI_2G_KPI.SDCCH_DROP_RATE_HUAWEI",
                    "TCH_ASSIGNMENT_SUCCESS_RATE_HUAWEI": "HUAWEI_2G_KPI.TCH_ASSIGNMENT_SUCCESS_RATE_HUAWEI",
                    "SDCCH_TRAFFIC_HUAWEI": "HUAWEI_2G_KPI.SDCCH_TRAFFIC_HUAWEI",
                    "CELL_AVAILABILITY_RATE_HUAWEI": "HUAWEI_2G_KPI.CELL_AVAILABILITY_RATE_HUAWEI",

                    "CSSR_ERICSSON": "ERICSSON_2G_KPI.CSSR_ERICSSON",
                    "CDR_ERICSSON": "ERICSSON_2G_KPI.CDR_ERICSSON",
                    "CBR_ERICSSON": "ERICSSON_2G_KPI.CBR_ERICSSON",
                    "CELL_AVAILABILITY_RATE_ERICSSON": "ERICSSON_2G_KPI.CELL_AVAILABILITY_RATE_ERICSSON",
                    "TCH_AVAILABILITY_RATE_ERICSSON": "ERICSSON_2G_KPI.TCH_AVAILABILITY_RATE_ERICSSON",
                    "DOWNTIME_MANUAL": "ERICSSON_2G_KPI.DOWNTIME_MANUAL",
                    "TCH_CONGESTION_RATE_ERICSSON": "ERICSSON_2G_KPI.TCH_CONGESTION_RATE_ERICSSON",
                    "SDCCH_DROP_RATE_ERICSSON": "ERICSSON_2G_KPI.SDCCH_DROP_RATE_ERICSSON",
                    "SDCCH_TRAFFIC_ERICSSON": "ERICSSON_2G_KPI.SDCCH_TRAFFIC_ERICSSON",
                    "SDCCH_BLOCKING_RATE_ERICSSON": "ERICSSON_2G_KPI.SDCCH_BLOCKING_RATE_ERICSSON",
                    "SDCCH_CONGESTION_RATE_ERICSSON": "ERICSSON_2G_KPI.SDCCH_CONGESTION_RATE_ERICSSON",

                    "CSSR_NETWORK": "(COALESCE(HUAWEI_2G_KPI.CSSR_HUAWEI, 0) * 0.22) + (COALESCE(ERICSSON_2G_KPI.CSSR_ERICSSON, 0) * 0.78)",
                    "CDR_NETWORK": "(COALESCE(HUAWEI_2G_KPI.CDR_HUAWEI, 0) * 0.22) + (COALESCE(ERICSSON_2G_KPI.CDR_ERICSSON, 0) * 0.78)",
                    "CBR_NETWORK": "(COALESCE(HUAWEI_2G_KPI.CBR_HUAWEI, 0) * 0.22) + (COALESCE(ERICSSON_2G_KPI.CBR_ERICSSON, 0) * 0.78)",
                    "TCH_CONGESTION_RATE_NETWORK": "(COALESCE(HUAWEI_2G_KPI.TCH_CONGESTION_RATE_HUAWEI, 0) * 0.22) + (COALESCE(ERICSSON_2G_KPI.TCH_CONGESTION_RATE_ERICSSON, 0) * 0.78)",
                    "SDCCH_CONGESTION_RATE_NETWORK": "(COALESCE(HUAWEI_2G_KPI.SDCCH_CONGESTION_RATE_HUAWEI, 0) * 0.22) + (COALESCE(ERICSSON_2G_KPI.SDCCH_CONGESTION_RATE_ERICSSON, 0) * 0.78)",
                    "SDCCH_DROP_RATE_NETWORK": "(COALESCE(HUAWEI_2G_KPI.SDCCH_DROP_RATE_HUAWEI, 0) * 0.22) + (COALESCE(ERICSSON_2G_KPI.SDCCH_DROP_RATE_ERICSSON, 0) * 0.78)",
                    "SDCCH_TRAFFIC_NETWORK": "(HUAWEI_2G_KPI.SDCCH_TRAFFIC_HUAWEI) + (ERICSSON_2G_KPI.SDCCH_TRAFFIC_ERICSSON)",
                    "CELL_AVAILABILITY_RATE_NETWORK": "(COALESCE(HUAWEI_2G_KPI.CELL_AVAILABILITY_RATE_HUAWEI, 0) * 0.22) + (COALESCE(ERICSSON_2G_KPI.CELL_AVAILABILITY_RATE_ERICSSON, 0) * 0.78)"
                },

                "chart_configs": {
                    "CSSR_ALL": {
                        "KPIS": ["CSSR_NETWORK", "CSSR_HUAWEI", "CSSR_ERICSSON"],
                        "title": "2G CSSR - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "CSSR (%)",
                        "threshold": 99
                    },
                    "CDR_ALL": {
                        "KPIS": ["CDR_NETWORK", "CDR_HUAWEI", "CDR_ERICSSON"],
                        "title": "2G CDR - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "CDR (%)"
                    },
                    "CBR_ALL": {
                        "KPIS": ["CBR_NETWORK", "CBR_HUAWEI", "CBR_ERICSSON"],
                        "title": "2G CBR - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "CBR (%)",
                        "threshold": 3
                    },
                    "TCH_CONGESTION_RATE_ALL": {
                        "KPIS": ["TCH_CONGESTION_RATE_NETWORK", "TCH_CONGESTION_RATE_HUAWEI", "TCH_CONGESTION_RATE_ERICSSON"],
                        "title": "2G TCH Congestion - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "TCH Congestion (%)",
                        "threshold": 2
                    },
                    "SDCCH_CONGESTION_RATE_ALL": {
                        "KPIS": ["SDCCH_CONGESTION_RATE_NETWORK", "SDCCH_CONGESTION_RATE_HUAWEI", "SDCCH_CONGESTION_RATE_ERICSSON"],
                        "title": "2G SDCCH Congestion - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "SDCCH Congestion (%)",
                        "threshold": 1
                    },
                    "SDCCH_DROP_RATE_ALL": {
                        "KPIS": ["SDCCH_DROP_RATE_NETWORK", "SDCCH_DROP_RATE_HUAWEI", "SDCCH_DROP_RATE_ERICSSON"],
                        "title": "2G SDCCH Drop Rate - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "SDCCH Drop Rate (%)"
                    },
                    "CELL_AVAILABILITY_ALL": {
                        "KPIS": ["CELL_AVAILABILITY_RATE_NETWORK", "CELL_AVAILABILITY_RATE_HUAWEI", "CELL_AVAILABILITY_RATE_ERICSSON"],
                        "title": "2G Cell Availability - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "Availability (%)",
                        "threshold": 99
                    },
                    "SDCCH_TRAFFIC_ALL": {
                        "KPIS": ["SDCCH_TRAFFIC_NETWORK", "SDCCH_TRAFFIC_HUAWEI", "SDCCH_TRAFFIC_ERICSSON"],
                        "title": "2G SDCCH Traffic - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "SDCCH Traffic (Erl)"
                    },

                    "CSSR_NETWORK_vs_CDR": {
                        "KPIS": ["CSSR_NETWORK", "CDR_NETWORK"],
                        "title": "2G Network CSSR vs CDR",
                        "default_type_1": "line",
                        "default_type_2": "bar",
                        "is_dual_axis": True,
                        "y_axis_titles": ["CSSR (%)", "CDR (%)"]
                    }

                },

                "sql_template": """
                    SELECT {fields}
                    FROM {cte_joins}
                    {group_by}
                    {order_by}
                """
            },

        
            "KPI_MONITORING_2": {
                "query_type": "multi_cte",
                "description": "3G Network KPI - Combined Vendors (Hourly with All KPIs)",
                "date_time_filters": ["start_date", "end_date", "multiple_date", "start_hour", "end_hour", "multiple_hour"],

                # Subqueries configuration
                "subqueries": [
                    {
                        "name": "HUAWEI_2G_KPI",
                        "table": "hourly_huawei_2g_all_counters",
                        "alias": "h",
                        "time_granularities": {
                            "HOURLY": {"is_available": True, "date_field": "date", "time_field": "TIME_FORMAT(time, '%H:%i:%s')"},
                            "DAILY": {"is_available": True, "date_field": "date", "time_field": None},
                            "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(date, 1)", "time_field": None},
                            "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(date, '%Y-%m')", "time_field": None}
                        },
                        "aggregation_levels": {},
                        "kpi_fields": {
                            "CSSR_HUAWEI": """CASE
                                WHEN COALESCE(SUM(CAST(CELL_KPI_SD_REQ AS DOUBLE)), 0) = 0
                                  OR COALESCE(SUM(CAST(CELL_KPI_SD_SUCC AS DOUBLE)), 0) = 0
                                  OR (COALESCE(SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE)), 0)
                                       + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
                                       + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0)) = 0
                                  OR (COALESCE(SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE)), 0)
                                       + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE)), 0)
                                       + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)), 0)) = 0
                                THEN 0
                                ELSE
                                  100.0
                                  * (SUM(CAST(CELL_KPI_SD_SUCC AS DOUBLE)) / SUM(CAST(CELL_KPI_SD_REQ AS DOUBLE)))
                                  * (1.0 - (SUM(CAST(CELL_SD_CALL_DROPS AS DOUBLE)) / SUM(CAST(CELL_KPI_SD_SUCC AS DOUBLE))))
                                  * ((SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE))
                                       + SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE))
                                       + SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)))
                                      / (SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE))
                                          + SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE))
                                          + SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE))))
                                  * (1.0 - ((SUM(CAST(CELL_KPI_TCH_DROPS_SIG AS DOUBLE))
                                              + SUM(CAST(CELL_TRAF_CH_CALL_DROPS AS DOUBLE))
                                              + SUM(CAST(CELL_KPI_TCH_HO_DROPS_TRAF AS DOUBLE)))
                                            / (SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE))
                                                + SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE))
                                                + SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)))))
                              END""",
                            "CDR_HUAWEI": """100.0 * (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_DROPS_SIG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_TRAF_CH_CALL_DROPS AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_DROPS_TRAF AS DOUBLE)), 0)
                              ) / (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)), 0)
                                + 1e-7
                              )""",
                            "CBR_HUAWEI": """100.0 * (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_CONG_SIG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_CONG_TRAF AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_CONGEST_TRAF AS DOUBLE)), 0)
                              ) / (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0)
                                + 1e-7
                              )""",
                            "TCH_CONGESTION_RATE_HUAWEI": """100.0 * (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_CONG_TRAF AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_CONGEST_TRAF AS DOUBLE)), 0)
                              ) / (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0)
                                + 1e-7
                              )""",
                            "SDCCH_CONGESTION_RATE_HUAWEI": """100.0 * COALESCE(SUM(CAST(CELL_KPI_SD_CONGEST AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(CELL_KPI_SD_REQ AS DOUBLE)), 0) + 1e-7)""",
                            "SDCCH_DROP_RATE_HUAWEI": """100.0 * COALESCE(SUM(CAST(CELL_SD_CALL_DROPS AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(CELL_IMM_ASS_SUCC_SD AS DOUBLE)), 0) + 1e-7)""",
                            "TCH_ASSIGNMENT_SUCCESS_RATE_HUAWEI": """100.0 * (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)), 0)
                              ) / (
                                COALESCE(SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0)
                                + 1e-7
                              )""",
                            "SDCCH_TRAFFIC_HUAWEI": """COALESCE(SUM(CAST(CELL_KPI_SD_TRAF_ERL AS DOUBLE)), 0)""",
                            "CELL_AVAILABILITY_RATE_HUAWEI": """100.0 * COALESCE(SUM(CAST(CELL_KPI_TCH_AVAIL_NUM AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(CELL_KPI_TCH_CFG_NUM AS DOUBLE)), 0) + 1e-7)"""
                        },
                        
                        "sql_template": """
                            SELECT {fields}
                            FROM {table} {alias}
                            WHERE {where_conditions}
                            {group_by}
                        """
                    },
                    {
                        "name": "ERICSSON_2G_KPI",
                        "table": "hourly_ericsson_arcep_2g_counters",
                        "alias": "e",
                        "time_granularities": {
                            "HOURLY": {"is_available": True, "date_field": "DATE", "time_field": "TIME"},
                            "DAILY": {"is_available": True, "date_field": "DATE", "time_field": None},
                            "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(DATE, 1)", "time_field": None},
                            "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(DATE, '%Y-%m')", "time_field": None}
                        },
                        "aggregation_levels": {},
                        "kpi_fields": {
                            "CSSR_ERICSSON": """CASE
                                WHEN COALESCE(SUM(CAST(CCALLS AS DOUBLE)), 0) = 0
                                  OR COALESCE(SUM(CAST(CMSESTAB AS DOUBLE)), 0) = 0
                                  OR COALESCE(SUM(CAST(TASSALL AS DOUBLE)), 0) = 0
                                  OR (COALESCE(SUM(CAST(TFCASSALL AS DOUBLE)), 0)
                                       + COALESCE(SUM(CAST(THCASSALL AS DOUBLE)), 0)
                                       + COALESCE(SUM(CAST(TFCASSALLSUB AS DOUBLE)), 0)
                                       + COALESCE(SUM(CAST(THCASSALLSUB AS DOUBLE)), 0)) = 0
                                THEN 0
                                ELSE
                                  100.0
                                  * (1.0 - (SUM(CAST(CCONGS AS DOUBLE)) / SUM(CAST(CCALLS AS DOUBLE))))
                                  * (1.0 - ((SUM(CAST(CNDROP AS DOUBLE)) - SUM(CAST(CNRELCONG AS DOUBLE))) / SUM(CAST(CMSESTAB AS DOUBLE))))
                                  * (SUM(CAST(TCASSALL AS DOUBLE)) / SUM(CAST(TASSALL AS DOUBLE)))
                                  * (1.0 - ((SUM(CAST(TFNDROP AS DOUBLE)) + SUM(CAST(THNDROP AS DOUBLE))
                                              + SUM(CAST(TFNDROPSUB AS DOUBLE)) + SUM(CAST(THNDROPSUB AS DOUBLE)))
                                            / (SUM(CAST(TFCASSALL AS DOUBLE)) + SUM(CAST(THCASSALL AS DOUBLE))
                                                + SUM(CAST(TFCASSALLSUB AS DOUBLE)) + SUM(CAST(THCASSALLSUB AS DOUBLE)))))
                              END""",
                            "CBR_ERICSSON": """100.0 * (
                                COALESCE(SUM(CAST(CNRELCONG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(TFNRELCONG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(TFNRELCONGSUB AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(THNRELCONG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(THNRELCONGSUB AS DOUBLE)), 0)
                              ) / (COALESCE(SUM(CAST(TASSALL AS DOUBLE)), 0) + 1e-7)""",
                            "CDR_ERICSSON": """100.0 * (
                                COALESCE(SUM(CAST(THNDROP AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(THNDROPSUB AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(TFNDROP AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(TFNDROPSUB AS DOUBLE)), 0)
                              ) / (
                                COALESCE(SUM(CAST(TFCASSALL AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(TFCASSALLSUB AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(THCASSALL AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(THCASSALLSUB AS DOUBLE)), 0)
                                + 1e-7
                              )""",
                            "CELL_AVAILABILITY_RATE_ERICSSON": """100.0 - (100.0 * COALESCE(SUM(CAST(TDWNACC AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(TDWNSCAN AS DOUBLE)), 0) + 1e-7))""",
                            "TCH_AVAILABILITY_RATE_ERICSSON": """100.0 * (COALESCE(SUM(CAST(TAVAACC AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(TAVASCAN AS DOUBLE)), 0) + 1e-7)) /
                              (COALESCE(AVG(CAST(AVG_TNUCHCNT AS DOUBLE)), 1))""",
                            "DOWNTIME_MANUAL": """COALESCE(SUM(CAST(HDWNACC AS DOUBLE)), 0)""",
                            "TCH_CONGESTION_RATE_ERICSSON": """100.0 * (
                                COALESCE(SUM(CAST(CNRELCONG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(TFNRELCONG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(TFNRELCONGSUB AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(THNRELCONG AS DOUBLE)), 0)
                                + COALESCE(SUM(CAST(THNRELCONGSUB AS DOUBLE)), 0)
                              ) / (COALESCE(SUM(CAST(TASSALL AS DOUBLE)), 0) + 1e-7)""",
                            "SDCCH_DROP_RATE_ERICSSON": """100.0 * COALESCE(SUM(CAST(CNDROP AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(CMSESTAB AS DOUBLE)), 0) + 1e-7)""",
                            "SDCCH_TRAFFIC_ERICSSON": """COALESCE(SUM(CAST(CTRALACC AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(CNSCAN AS DOUBLE)), 0) + 1e-7)""",
                            "SDCCH_BLOCKING_RATE_ERICSSON": """100.0 * COALESCE(SUM(CAST(CCONGS AS DOUBLE)), 0) /
                              (COALESCE(SUM(CAST(CCALLS AS DOUBLE)), 0) + 1e-7)""",
                            "SDCCH_CONGESTION_RATE_ERICSSON": """100.0 * (COALESCE(SUM(CAST(CCONGS AS DOUBLE)), 0) + COALESCE(SUM(CAST(CCONGSSUB AS DOUBLE)), 0)) /
                              (COALESCE(SUM(CAST(CCALLS AS DOUBLE)), 0) + 1e-7)"""
                        },
                        "sql_template": """
                            SELECT {fields}
                            FROM {table} {alias}
                            WHERE {where_conditions}
                            {group_by}
                        """
                    }
                ],

                # How to join the CTEs
                "cte_joins": [
                    {
                        "type": "LEFT JOIN",
                        "left": "HUAWEI_2G_KPI",
                        "right": "ERICSSON_2G_KPI",
                        "on": ["HUAWEI_2G_KPI.date = ERICSSON_2G_KPI.DATE", "HUAWEI_2G_KPI.time = ERICSSON_2G_KPI.TIME"]
                    }
                ],

                # Outer query configuration (what the client controls)
                "time_granularities": {
                    "HOURLY": {"is_available": True, "date_field": "HUAWEI_2G_KPI.date", "time_field": "HUAWEI_2G_KPI.time"},
                    "DAILY": {"is_available": True, "date_field": "HUAWEI_2G_KPI.date", "time_field": None},
                    "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(HUAWEI_2G_KPI.date, 1)", "time_field": None},
                    "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(HUAWEI_2G_KPI.date, '%Y-%m')", "time_field": None}
                },

                "aggregation_levels": {
                },

                "kpi_fields": {
                    "CSSR_HUAWEI": "HUAWEI_2G_KPI.CSSR_HUAWEI",
                    "CDR_HUAWEI": "HUAWEI_2G_KPI.CDR_HUAWEI",
                    "CBR_HUAWEI": "HUAWEI_2G_KPI.CBR_HUAWEI",
                    "TCH_CONGESTION_RATE_HUAWEI": "HUAWEI_2G_KPI.TCH_CONGESTION_RATE_HUAWEI",
                    "SDCCH_CONGESTION_RATE_HUAWEI": "HUAWEI_2G_KPI.SDCCH_CONGESTION_RATE_HUAWEI",
                    "SDCCH_DROP_RATE_HUAWEI": "HUAWEI_2G_KPI.SDCCH_DROP_RATE_HUAWEI",
                    "TCH_ASSIGNMENT_SUCCESS_RATE_HUAWEI": "HUAWEI_2G_KPI.TCH_ASSIGNMENT_SUCCESS_RATE_HUAWEI",
                    "SDCCH_TRAFFIC_HUAWEI": "HUAWEI_2G_KPI.SDCCH_TRAFFIC_HUAWEI",
                    "CELL_AVAILABILITY_RATE_HUAWEI": "HUAWEI_2G_KPI.CELL_AVAILABILITY_RATE_HUAWEI",

                    "CSSR_ERICSSON": "ERICSSON_2G_KPI.CSSR_ERICSSON",
                    "CDR_ERICSSON": "ERICSSON_2G_KPI.CDR_ERICSSON",
                    "CBR_ERICSSON": "ERICSSON_2G_KPI.CBR_ERICSSON",
                    "CELL_AVAILABILITY_RATE_ERICSSON": "ERICSSON_2G_KPI.CELL_AVAILABILITY_RATE_ERICSSON",
                    "TCH_AVAILABILITY_RATE_ERICSSON": "ERICSSON_2G_KPI.TCH_AVAILABILITY_RATE_ERICSSON",
                    "DOWNTIME_MANUAL": "ERICSSON_2G_KPI.DOWNTIME_MANUAL",
                    "TCH_CONGESTION_RATE_ERICSSON": "ERICSSON_2G_KPI.TCH_CONGESTION_RATE_ERICSSON",
                    "SDCCH_DROP_RATE_ERICSSON": "ERICSSON_2G_KPI.SDCCH_DROP_RATE_ERICSSON",
                    "SDCCH_TRAFFIC_ERICSSON": "ERICSSON_2G_KPI.SDCCH_TRAFFIC_ERICSSON",
                    "SDCCH_BLOCKING_RATE_ERICSSON": "ERICSSON_2G_KPI.SDCCH_BLOCKING_RATE_ERICSSON",
                    "SDCCH_CONGESTION_RATE_ERICSSON": "ERICSSON_2G_KPI.SDCCH_CONGESTION_RATE_ERICSSON",

                    "CSSR_NETWORK": "(COALESCE(HUAWEI_2G_KPI.CSSR_HUAWEI, 0) * 0.22) + (COALESCE(ERICSSON_2G_KPI.CSSR_ERICSSON, 0) * 0.78)",
                    "CDR_NETWORK": "(COALESCE(HUAWEI_2G_KPI.CDR_HUAWEI, 0) * 0.22) + (COALESCE(ERICSSON_2G_KPI.CDR_ERICSSON, 0) * 0.78)",
                    "CBR_NETWORK": "(COALESCE(HUAWEI_2G_KPI.CBR_HUAWEI, 0) * 0.22) + (COALESCE(ERICSSON_2G_KPI.CBR_ERICSSON, 0) * 0.78)",
                    "TCH_CONGESTION_RATE_NETWORK": "(COALESCE(HUAWEI_2G_KPI.TCH_CONGESTION_RATE_HUAWEI, 0) * 0.22) + (COALESCE(ERICSSON_2G_KPI.TCH_CONGESTION_RATE_ERICSSON, 0) * 0.78)",
                    "SDCCH_CONGESTION_RATE_NETWORK": "(COALESCE(HUAWEI_2G_KPI.SDCCH_CONGESTION_RATE_HUAWEI, 0) * 0.22) + (COALESCE(ERICSSON_2G_KPI.SDCCH_CONGESTION_RATE_ERICSSON, 0) * 0.78)",
                    "SDCCH_DROP_RATE_NETWORK": "(COALESCE(HUAWEI_2G_KPI.SDCCH_DROP_RATE_HUAWEI, 0) * 0.22) + (COALESCE(ERICSSON_2G_KPI.SDCCH_DROP_RATE_ERICSSON, 0) * 0.78)",
                    "SDCCH_TRAFFIC_NETWORK": "(HUAWEI_2G_KPI.SDCCH_TRAFFIC_HUAWEI) + (ERICSSON_2G_KPI.SDCCH_TRAFFIC_ERICSSON)",
                    "CELL_AVAILABILITY_RATE_NETWORK": "(COALESCE(HUAWEI_2G_KPI.CELL_AVAILABILITY_RATE_HUAWEI, 0) * 0.22) + (COALESCE(ERICSSON_2G_KPI.CELL_AVAILABILITY_RATE_ERICSSON, 0) * 0.78)"
                },

                "chart_configs": {
                    "CSSR_ALL": {
                        "KPIS": ["CSSR_NETWORK", "CSSR_HUAWEI", "CSSR_ERICSSON"],
                        "title": "2G CSSR - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "CSSR (%)",
                        "threshold": 99
                    },
                    "CDR_ALL": {
                        "KPIS": ["CDR_NETWORK", "CDR_HUAWEI", "CDR_ERICSSON"],
                        "title": "2G CDR - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "CDR (%)"
                    },
                    "CBR_ALL": {
                        "KPIS": ["CBR_NETWORK", "CBR_HUAWEI", "CBR_ERICSSON"],
                        "title": "2G CBR - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "CBR (%)",
                        "threshold": 3
                    },
                    "TCH_CONGESTION_RATE_ALL": {
                        "KPIS": ["TCH_CONGESTION_RATE_NETWORK", "TCH_CONGESTION_RATE_HUAWEI", "TCH_CONGESTION_RATE_ERICSSON"],
                        "title": "2G TCH Congestion - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "TCH Congestion (%)",
                        "threshold": 2
                    },
                    "SDCCH_CONGESTION_RATE_ALL": {
                        "KPIS": ["SDCCH_CONGESTION_RATE_NETWORK", "SDCCH_CONGESTION_RATE_HUAWEI", "SDCCH_CONGESTION_RATE_ERICSSON"],
                        "title": "2G SDCCH Congestion - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "SDCCH Congestion (%)",
                        "threshold": 1
                    },
                    "SDCCH_DROP_RATE_ALL": {
                        "KPIS": ["SDCCH_DROP_RATE_NETWORK", "SDCCH_DROP_RATE_HUAWEI", "SDCCH_DROP_RATE_ERICSSON"],
                        "title": "2G SDCCH Drop Rate - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "SDCCH Drop Rate (%)"
                    },
                    "CELL_AVAILABILITY_ALL": {
                        "KPIS": ["CELL_AVAILABILITY_RATE_NETWORK", "CELL_AVAILABILITY_RATE_HUAWEI", "CELL_AVAILABILITY_RATE_ERICSSON"],
                        "title": "2G Cell Availability - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "Availability (%)",
                        "threshold": 99
                    },
                    "SDCCH_TRAFFIC_ALL": {
                        "KPIS": ["SDCCH_TRAFFIC_NETWORK", "SDCCH_TRAFFIC_HUAWEI", "SDCCH_TRAFFIC_ERICSSON"],
                        "title": "2G SDCCH Traffic - Network & Vendors",
                        "default_type": "line",
                        "y_axis_title": "SDCCH Traffic (Erl)"
                    },

                    "CSSR_NETWORK_vs_CDR": {
                        "KPIS": ["CSSR_NETWORK", "CDR_NETWORK"],
                        "title": "2G Network CSSR vs CDR",
                        "default_type_1": "line",
                        "default_type_2": "bar",
                        "is_dual_axis": True,
                        "y_axis_titles": ["CSSR (%)", "CDR (%)"]
                    }

                },

                "sql_template": """
                    SELECT {fields}
                    FROM {cte_joins}
                    {group_by}
                    {order_by}
                """
            },

     

        }
        ,

            "KPI_MONITORING2": {
                "description": "2G HUAWEI KPI Monitoring",
                "date_time_filters": ["start_date", "end_date", "multiple_date", "start_hour", "end_hour", "multiple_hour"],
                "parameters": [],
                "aggregation_levels": {
                   
                },

                "time_granularities": {
                    "HOURLY": {"is_available": True, "date_field":{"eric" "e.DATE"}, 
                               "time_field":{"e_time": "TIME_FORMAT(TIME, '%H:%i:%s') Time" } },

                    "DAILY": {"is_available": True, "date_field": "DATE", "time_field": None},
                    "WEEKLY": {"is_available": True, "date_field": "YEARWEEK(DATE, 1)", "time_field": None},
                    "MONTHLY": {"is_available": True, "date_field": "DATE_FORMAT(DATE, '%Y-%m')", "time_field": None}
                },

                # "kpi_fields": {
                #      "CSSR_HUAWEI": """100 * (SUM(Suc_SDCCH_Seiz) / NULLIF(SUM(SDCCH_Seiz_Req), 0)) *
                #         (1 - (SUM(SDCCHDrop) / NULLIF(SUM(Suc_SDCCH_Seiz), 0))) *
                #         (SUM(Suc_TCH_Seiz_SG) + SUM(Suc_TCH_Seiz_TrafChann) + SUM(Suc_TCH_Seiz_handTrafChan)) /
                #         NULLIF((SUM(TCH_Seizure_Req_SC) + SUM(TCH_Seiz_Req_TrafChan) + SUM(TCH_Seiz_Req_HandTrafChan)), 0) *
                #         (1 - (SUM(TCHDrop_Nume_ARCEP) / NULLIF(SUM(TCHDrop_Deno_ARCEP), 0)))""",

                #     "CDR_HUAWEI": "(100 * SUM(TCHDrop_Nume_ARCEP) / NULLIF(SUM(TCHDrop_Deno_ARCEP), 0))",
                #     "CBR_HUAWEI": "100 * (SUM(TCHCONG_Nume_ARCEP) / NULLIF(SUM(TCHCONG_Deno_ARCEP), 0))"
                # },

                "chart_configs":{
                    
                    "CSSR_HUAWEI":{
                        "KPIS":['CSSR_HUAWEI'],
                        "title":"2G Huawei CSSR ",
                        "default_type": "line",
                        "y_axis_title":"CSSR (%)",
                        "threshold":99
                    },
                    "CSSR_CDR_HUAWEI":{
                        "KPIS":["CSSR_HUAWEI","CDR_HUAWEI"],
                        "title":"2G Huawei CSSR vs CDR",
                        "default_type_1": "line",
                        "default_type_2": "bar",
                        "is_dual_axis":True,
                        "y_axis_titles":["CSSR (%)", "CDR (%)"]
                    },
                    "CBR_HUAWEI":{
                        "KPIS":["CBR_HUAWEI"],
                        "title":"2G Huawei CBR ",
                        "default_type": "line",
                        "y_axis_title":"CBR (%)",
                        "threshold":3
                    }

                },

                "sql_template": """

                WITH
  HUAWEI_2G_KPI AS (
    SELECT
      h.date,
      h.time
      -- h.CELL_NAME,
      -- ept.SITE_NAME,
      -- ept.ARRONDISSEMENT,
      -- ept.COMMUNE,
      -- ept.DEPARTEMENT,

      
      -- CASE
      --   WHEN COALESCE(SUM(CAST(CELL_KPI_SD_REQ AS DOUBLE)), 0) = 0
      --     OR COALESCE(SUM(CAST(CELL_KPI_SD_SUCC AS DOUBLE)), 0) = 0
      --     OR (COALESCE(SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE)), 0)
      --          + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
      --          + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0)) = 0
      --     OR (COALESCE(SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE)), 0)
      --          + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE)), 0)
      --          + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)), 0)) = 0
      --   THEN 0
      --   ELSE
      --     100.0
      --     * (SUM(CAST(CELL_KPI_SD_SUCC AS DOUBLE))
      --         / SUM(CAST(CELL_KPI_SD_REQ AS DOUBLE)))
      --     * (1.0 - (SUM(CAST(CELL_SD_CALL_DROPS AS DOUBLE))
      --                / SUM(CAST(CELL_KPI_SD_SUCC AS DOUBLE))))
      --     * ((SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE))
      --          + SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE))
      --          + SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)))
      --         / (SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE))
      --             + SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE))
      --             + SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE))))
      --     * (1.0 - ((SUM(CAST(CELL_KPI_TCH_DROPS_SIG AS DOUBLE))
      --                 + SUM(CAST(CELL_TRAF_CH_CALL_DROPS AS DOUBLE))
      --                 + SUM(CAST(CELL_KPI_TCH_HO_DROPS_TRAF AS DOUBLE)))
      --                / (SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE))
      --                    + SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE))
      --                    + SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)))))
      -- END AS CSSR_HUAWEI
      

      ,100 * (SUM(CELL_KPI_SD_SUCC
      ) / SUM(CELL_KPI_SD_REQ)) * 
  (1 - (SUM(CELL_SD_CALL_DROPS) / SUM(CELL_KPI_SD_SUCC))) * 
  (SUM(CELL_KPI_TCH_SUCC_SIG) + SUM(CELL_KPI_TCH_ASS_SUCC_TRAF) + SUM(CELL_KPI_TCH_HO_SUCC_TRAF)) /
  (SUM(CELL_KPI_TCH_REQ_SIG) + SUM(CELL_KPI_TCH_ASS_REQ_TRAF) + SUM(CELL_KPI_TCH_HO_REQ_TRAF)) *
  (1 - (((sum(CELL_KPI_TCH_DROPS_SIG)+sum(CELL_KPI_TCH_STATIC_DROPS_TRAF)+sum(CELL_KPI_TCH_HO_DROPS_TRAF))) /
    ((sum(CELL_KPI_TCH_SUCC_SIG)+sum(CELL_KPI_TCH_ASS_SUCC_TRAF)+sum(CELL_KPI_TCH_HO_SUCC_TRAF))))) AS  CSSR_HUAWEI,

      -- ARCEP 2G CALL DROP RATE - HUAWEI
      100.0 * (
        COALESCE(SUM(CAST(CELL_KPI_TCH_DROPS_SIG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_TRAF_CH_CALL_DROPS AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_DROPS_TRAF AS DOUBLE)), 0)
      ) / NULLIF(
        COALESCE(SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)), 0),
        0
      ) AS CDR_HUAWEI,

      -- ARCEP 2G CALL BLOCKING RATE - HUAWEI
      100.0 * (
        COALESCE(SUM(CAST(CELL_KPI_TCH_CONG_SIG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_CONG_TRAF AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_CONGEST_TRAF AS DOUBLE)), 0)
      ) / NULLIF(
        COALESCE(SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0),
        0
      ) AS CBR_HUAWEI,

      -- TCH Congestion Rate - HUAWEI
      100.0 * (
        COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_CONG_TRAF AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_CONGEST_TRAF AS DOUBLE)), 0)
      ) / NULLIF(
        COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0),
        0
      ) AS TCH_CONGESTION_RATE_HUAWEI,

      -- SDCCH Congestion Rate - HUAWEI
      100.0 * COALESCE(SUM(CAST(CELL_KPI_SD_CONGEST AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(CELL_KPI_SD_REQ AS DOUBLE)), 0), 0) AS SDCCH_CONGESTION_RATE_HUAWEI,

      -- SDCCH Drop Rate - HUAWEI
      100.0 * COALESCE(SUM(CAST(CELL_SD_CALL_DROPS AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(CELL_IMM_ASS_SUCC_SD AS DOUBLE)), 0), 0) AS SDCCH_DROP_RATE_HUAWEI,

      -- ARCEP TCH Assignment Success Rate - HUAWEI (Vendor-specific)
      100.0 * (
        COALESCE(SUM(CAST(CELL_KPI_TCH_SUCC_SIG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_SUCC_TRAF AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_SUCC_TRAF AS DOUBLE)), 0)
      ) / NULLIF(
        COALESCE(SUM(CAST(CELL_KPI_TCH_REQ_SIG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_ASS_REQ_TRAF AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(CELL_KPI_TCH_HO_REQ_TRAF AS DOUBLE)), 0),
        0
      ) AS TCH_ASSIGNMENT_SUCCESS_RATE_HUAWEI,

      -- SDCCH Traffic (Erlangs) - HUAWEI
      COALESCE(SUM(CAST(CELL_KPI_SD_TRAF_ERL AS DOUBLE)), 0) AS SDCCH_TRAFFIC_HUAWEI,

      -- 2G Availability Rate - HUAWEI
      100.0 * COALESCE(SUM(CAST(CELL_KPI_TCH_AVAIL_NUM AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(CELL_KPI_TCH_CFG_NUM AS DOUBLE)), 0), 0) AS CELL_AVAILABILITY_RATE_HUAWEI,


        COALESCE(SUM(CAST(CELL_KPI_TCH_TRAF_ERL_TRAF AS DOUBLE)), 0) TRAFFIC_VOIX_HUAWEI,
        -- handover sr
        100.0 * (
  (COALESCE(SUM(CAST(CELL_INTRABSC_OUTCELL_HO_SUCC AS DOUBLE)), 0)
   + COALESCE(SUM(CAST(CELL_INTERBSC_OUTCELL_HO_SUCC AS DOUBLE)), 0))
  / NULLIF(
      COALESCE(SUM(CAST(CELL_INTRABSC_OUTCELL_HO_CMD AS DOUBLE)), 0)
      + COALESCE(SUM(CAST(CELL_INTERBSC_OUTCELL_HO_CMD AS DOUBLE)), 0),
      0
    )
) AS HANDOVER_SUCCESS_RATE_HUAWEI



    FROM
      hourly_huawei_2g_all_counters h
    -- LEFT JOIN
    --   EPT_2G ept ON h.CELL_NAME = ept.CELL_NAME AND ept.VENDOR = 'HUAWEI'

    WHERE
      h.date BETWEEN '2026-01-26' AND '2026-01-26'  -- Adjust date range as needed

    GROUP BY
      h.date,
      h.time
      -- h.CELL_NAME,
      -- ept.SITE_NAME,
      -- ept.ARRONDISSEMENT,
      -- ept.COMMUNE,
      -- ept.DEPARTEMENT
  ),

  ERICSSON_2G_KPI AS (
    SELECT
      e.DATE,
      e.TIME,
      -- e.CELL_NAME,
      -- ept.SITE_NAME,
      -- ept.ARRONDISSEMENT,
      -- ept.COMMUNE,
      -- ept.DEPARTEMENT,

      -- CSSR (Call Setup Success Rate) - ERICSSON
      CASE
        WHEN COALESCE(SUM(CAST(CCALLS AS DOUBLE)), 0) = 0
          OR COALESCE(SUM(CAST(CMSESTAB AS DOUBLE)), 0) = 0
          OR COALESCE(SUM(CAST(TASSALL AS DOUBLE)), 0) = 0
          OR (COALESCE(SUM(CAST(TFCASSALL AS DOUBLE)), 0)
               + COALESCE(SUM(CAST(THCASSALL AS DOUBLE)), 0)
               + COALESCE(SUM(CAST(TFCASSALLSUB AS DOUBLE)), 0)
               + COALESCE(SUM(CAST(THCASSALLSUB AS DOUBLE)), 0)
            ) = 0
        THEN 0
        ELSE
          100.0
          * (1.0 - (SUM(CAST(CCONGS AS DOUBLE))
                     / SUM(CAST(CCALLS AS DOUBLE))))
          * (1.0 - ((SUM(CAST(CNDROP AS DOUBLE))
                      - SUM(CAST(CNRELCONG AS DOUBLE)))
                     / SUM(CAST(CMSESTAB AS DOUBLE))))
          * (SUM(CAST(TCASSALL AS DOUBLE))
              / SUM(CAST(TASSALL AS DOUBLE)))
          * (1.0 - (
              (SUM(CAST(TFNDROP AS DOUBLE))
                + SUM(CAST(THNDROP AS DOUBLE))
                + SUM(CAST(TFNDROPSUB AS DOUBLE))
                + SUM(CAST(THNDROPSUB AS DOUBLE))
              )
              / (
                SUM(CAST(TFCASSALL AS DOUBLE))
                + SUM(CAST(THCASSALL AS DOUBLE))
                + SUM(CAST(TFCASSALLSUB AS DOUBLE))
                + SUM(CAST(THCASSALLSUB AS DOUBLE))
              )
            ))
      END AS CSSR_ERICSSON,

      -- Call Blocking Rate - ERICSSON (ARCEP)
      100.0 * (
        COALESCE(SUM(CAST(CNRELCONG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(TFNRELCONG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(TFNRELCONGSUB AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(THNRELCONG AS DOUBLE)), 0)
        + COALESCE(SUM(CAST(THNRELCONGSUB AS DOUBLE)), 0)
      ) / NULLIF(COALESCE(SUM(CAST(TASSALL AS DOUBLE)), 0), 0) AS CBR_ERICSSON,

      -- Call Drop Rate - ERICSSON (ARCEP)
      100.0 * (
        COALESCE(SUM(CAST(THNDROP AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(THNDROPSUB AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(TFNDROP AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(TFNDROPSUB AS DOUBLE)), 0)
      ) / NULLIF(
        COALESCE(SUM(CAST(TFCASSALL AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(TFCASSALLSUB AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(THCASSALL AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(THCASSALLSUB AS DOUBLE)), 0),
        0
      ) AS CDR_ERICSSON,

      -- 2G Cell Availability Rate - ERICSSON
      100.0 - (100.0 * COALESCE(SUM(CAST(TDWNACC AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(TDWNSCAN AS DOUBLE)), 0), 0)) AS CELL_AVAILABILITY_RATE_ERICSSON,

      -- TCH Channel Availability Rate - ERICSSON (Vendor-specific)
      100.0 * (COALESCE(SUM(CAST(TAVAACC AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(TAVASCAN AS DOUBLE)), 0), 0)) /
        (COALESCE(AVG(CAST(AVG_TNUCHCNT AS DOUBLE)), 1)) AS TCH_AVAILABILITY_RATE_ERICSSON,

      -- Downtime Manual - ERICSSON (Vendor-specific)
      COALESCE(SUM(CAST(HDWNACC AS DOUBLE)), 0) AS DOWNTIME_MANUAL,

      -- TCH Congestion Rate - ERICSSON
      100.0 * (
        COALESCE(SUM(CAST(CNRELCONG AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(TFNRELCONG AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(TFNRELCONGSUB AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(THNRELCONG AS DOUBLE)), 0) +
        COALESCE(SUM(CAST(THNRELCONGSUB AS DOUBLE)), 0)
      ) / NULLIF(COALESCE(SUM(CAST(TASSALL AS DOUBLE)), 0), 0) AS TCH_CONGESTION_RATE_ERICSSON,

      -- SDCCH Drop Rate - ERICSSON
      100.0 * COALESCE(SUM(CAST(CNDROP AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(CMSESTAB AS DOUBLE)), 0), 0) AS SDCCH_DROP_RATE_ERICSSON,

      -- SDCCH Traffic - ERICSSON (Erlangs)
      COALESCE(SUM(CAST(CTRALACC AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(CNSCAN AS DOUBLE)), 0), 0) AS SDCCH_TRAFFIC_ERICSSON,

      -- SDCCH Blocking Rate - ERICSSON (Vendor-specific)
      100.0 * COALESCE(SUM(CAST(CCONGS AS DOUBLE)), 0) /
        NULLIF(COALESCE(SUM(CAST(CCALLS AS DOUBLE)), 0), 0) AS SDCCH_BLOCKING_RATE_ERICSSON,

      -- SDCCH Congestion Rate - ERICSSON
      100.0 * (COALESCE(SUM(CAST(CCONGS AS DOUBLE)), 0) + COALESCE(SUM(CAST(CCONGSSUB AS DOUBLE)), 0)) /
        NULLIF(COALESCE(SUM(CAST(CCALLS AS DOUBLE)), 0), 0) AS SDCCH_CONGESTION_RATE_ERICSSON,

      COALESCE(SUM(CAST(TRAFFIC_DATA_GB AS DOUBLE)), 0) TRAFFIC_DATA_GB_ERICSSON,
      COALESCE(SUM(CAST(TRAFFIC_VOIX AS DOUBLE)), 0) TRAFFIC_VOIX_ERICSSON
    FROM
      hourly_ericsson_arcep_2g_counters e
    -- LEFT JOIN
    --   EPT_2G ept ON e.CELL_NAME = ept.CELL_NAME AND ept.VENDOR = 'ERICSSON'

    WHERE
      e.DATE BETWEEN '2026-01-28' AND '2026-01-29'  -- Adjust date range as needed

    GROUP BY
      e.DATE,
      e.TIME
      -- e.CELL_NAME,
      -- ept.SITE_NAME,
      -- ept.ARRONDISSEMENT,
      -- ept.COMMUNE,
      -- ept.DEPARTEMENT
  )



  

SELECT
  COALESCE(h.date, e.DATE) AS date,
  COALESCE(h.time, e.TIME) AS time,

  -- Location Info
  -- COALESCE(h.SITE_NAME, e.SITE_NAME) AS SITE_NAME,
  -- COALESCE(h.ARRONDISSEMENT, e.ARRONDISSEMENT) AS ARRONDISSEMENT,
  -- COALESCE(h.COMMUNE, e.COMMUNE) AS COMMUNE,
  -- COALESCE(h.DEPARTEMENT, e.DEPARTEMENT) AS DEPARTEMENT,

  -- Vendor-specific KPIs
  h.CSSR_HUAWEI,
  -- h.CSSR_HUAWEI_2,
  e.CSSR_ERICSSON,
  h.CDR_HUAWEI,
  e.CDR_ERICSSON,
  h.CBR_HUAWEI,
  e.CBR_ERICSSON,
  h.TCH_CONGESTION_RATE_HUAWEI,
  e.TCH_CONGESTION_RATE_ERICSSON,
  h.SDCCH_CONGESTION_RATE_HUAWEI,
  e.SDCCH_CONGESTION_RATE_ERICSSON,
  h.SDCCH_DROP_RATE_HUAWEI,
  e.SDCCH_DROP_RATE_ERICSSON,
  h.SDCCH_TRAFFIC_HUAWEI,
  e.SDCCH_TRAFFIC_ERICSSON,
  h.CELL_AVAILABILITY_RATE_HUAWEI,
  e.CELL_AVAILABILITY_RATE_ERICSSON,

  h.TRAFFIC_VOIX_HUAWEI,
  e.TRAFFIC_VOIX_ERICSSON,

  -- Vendor-specific only
  h.TCH_ASSIGNMENT_SUCCESS_RATE_HUAWEI,
  e.TCH_AVAILABILITY_RATE_ERICSSON,
  e.DOWNTIME_MANUAL,
  e.SDCCH_BLOCKING_RATE_ERICSSON,

  h.HANDOVER_SUCCESS_RATE_HUAWEI,

  -- Network-level weighted averages (0.22 Huawei + 0.78 Ericsson)
  (COALESCE(h.CSSR_HUAWEI, 0) * 0.22) + (COALESCE(e.CSSR_ERICSSON, 0) * 0.78) AS CSSR_NETWORK,
  (COALESCE(h.CDR_HUAWEI, 0) * 0.22) + (COALESCE(e.CDR_ERICSSON, 0) * 0.78) AS CDR_NETWORK,
  (COALESCE(h.CBR_HUAWEI, 0) * 0.22) + (COALESCE(e.CBR_ERICSSON, 0) * 0.78) AS CBR_NETWORK,
  (COALESCE(h.TCH_CONGESTION_RATE_HUAWEI, 0) * 0.22) + (COALESCE(e.TCH_CONGESTION_RATE_ERICSSON, 0) * 0.78) AS TCH_CONGESTION_RATE_NETWORK,
  (COALESCE(h.SDCCH_CONGESTION_RATE_HUAWEI, 0) * 0.22) + (COALESCE(e.SDCCH_CONGESTION_RATE_ERICSSON, 0) * 0.78) AS SDCCH_CONGESTION_RATE_NETWORK,
  (COALESCE(h.SDCCH_DROP_RATE_HUAWEI, 0) * 0.22) + (COALESCE(e.SDCCH_DROP_RATE_ERICSSON, 0) * 0.78) AS SDCCH_DROP_RATE_NETWORK,
  (COALESCE(h.SDCCH_TRAFFIC_HUAWEI, 0) * 0.22) + (COALESCE(e.SDCCH_TRAFFIC_ERICSSON, 0) * 0.78) AS SDCCH_TRAFFIC_NETWORK,
  (COALESCE(h.CELL_AVAILABILITY_RATE_HUAWEI, 0) * 0.22) + (COALESCE(e.CELL_AVAILABILITY_RATE_ERICSSON, 0) * 0.78) AS CELL_AVAILABILITY_RATE_NETWORK,
  h.TRAFFIC_VOIX_HUAWEI + e.TRAFFIC_VOIX_ERICSSON AS TRAFFIC_NETWORK

FROM HUAWEI_2G_KPI h
LEFT JOIN ERICSSON_2G_KPI e
  ON h.date = e.DATE
  AND h.time = e.TIME

ORDER BY date, time



                """
            }


    },

    # =========================================================================
    # SAMPLE RAW_SQL QUERY (for testing the new query type)
    # =========================================================================
    "TEST_RAW_SQL": {

        "SAMPLE": {
            "KPI_TEST": {
                "query_type": "raw_sql",
                "description": "Sample raw_sql query - 2G Network KPIs",
                "date_time_filters": ["start_date", "end_date", "multiple_date", "start_hour", "end_hour", "multiple_hour"],

                "time_granularities": {
                    "HOURLY": {"is_available": True},
                    "DAILY": {"is_available": True},
                    "WEEKLY": {"is_available": True},
                    "MONTHLY": {"is_available": True}
                },

                "sources": {
                    "eric": {
                        "date_col": "e.DATE",
                        "time_col": "e.TIME",
                        "aggregations": {
                            "cell_name": "e.CELL_NAME",
                            "site_name": "e.SITE_NAME",
                            "commune": "e.COMMUNE"
                        }
                    },
                    "huawei": {
                        "date_col": "h.date",
                        "time_col": "h.time",
                        "aggregations": {
                            "cell_name": "h.cell_name",
                            "site_name": "h.site_name",
                            "commune": "h.COMMUNE"
                        }
                    }
                },

                # Joins config: auto-generate JOIN ON conditions based on granularity/aggregation
                "joins": [
                    {
                        "left": "eric",
                        "right": "huawei",
                        "include_date": True,       # Always include date
                        "include_time": True,       # Only if HOURLY granularity
                        "include_aggregation": True, # Only if aggregation is selected
                        "custom": []                 # Extra static conditions if needed
                    }
                ],

                "sql_template": """
WITH ERICSSON_2G AS (
    SELECT
        {eric__select_fields},
        SUM(e.counter1) AS kpi_eric_1,
        SUM(e.counter2) AS kpi_eric_2
    FROM ericsson_2g_table e
    WHERE {eric__where_clause}
    GROUP BY {eric__group_by}
),
HUAWEI_2G AS (
    SELECT
        {huawei__select_fields},
        SUM(h.counter1) AS kpi_huawei_1,
        SUM(h.counter2) AS kpi_huawei_2
    FROM huawei_2g_table h
    WHERE {huawei__where_clause}
    GROUP BY {huawei__group_by}
)
SELECT
    COALESCE(e.date, h.date) AS date,
    COALESCE(e.time, h.time) AS time,
    e.kpi_eric_1,
    e.kpi_eric_2,
    h.kpi_huawei_1,
    h.kpi_huawei_2,
    (COALESCE(e.kpi_eric_1, 0) * 0.78) + (COALESCE(h.kpi_huawei_1, 0) * 0.22) AS kpi_network_1
FROM ERICSSON_2G e
LEFT JOIN HUAWEI_2G h
    ON {eric__huawei__join_on}
ORDER BY {eric__order_by}
                """
            }
        }

    }

}

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def get_query_config(category: str, subcategory: Optional[str], query_name: str) -> Dict:
    """Get query configuration"""
    try:
        if subcategory:
            return QUERY_CONFIG[category][subcategory][query_name]
        else:
            return QUERY_CONFIG[category][query_name]
    except KeyError:
        raise HTTPException(
            status_code=404,
            detail=f"Query not found: {category}/{subcategory}/{query_name}"
        )

def get_aggregation_hierarchy(aggregation_levels: Dict, selected_level: str) -> List[tuple]:
    """
    Get aggregation levels from selected level onwards (hierarchical filtering)
    Returns list of (display_name, sql_expression) tuples
    """
    levels_order = list(aggregation_levels.keys())

    if selected_level not in levels_order:
        # If no level or invalid, return all levels
        return [(name, expr) for name, expr in aggregation_levels.items()]

    # Get index of selected level
    start_index = levels_order.index(selected_level)

    # Return from selected level onwards
    return [(name, aggregation_levels[name]) for name in levels_order[start_index:]]


def validate_chart_configs(config: Dict) -> bool:
    """
    Validate chart_configs to ensure all referenced KPIs exist in kpi_fields.
    Raises ValueError if validation fails.
    """
    chart_configs = config.get("chart_configs", {})
    available_kpis = list(config.get("kpi_fields", {}).keys())

    for chart_name, chart_config in chart_configs.items():
        kpis = chart_config.get("KPIS", [])

        # Validate each KPI exists
        for kpi in kpis:
            if kpi not in available_kpis:
                raise ValueError(
                    f"Chart '{chart_name}': KPI '{kpi}' not found in available KPIs. "
                    f"Available: {available_kpis}"
                )

        # Validate dual-axis requirements
        if chart_config.get("is_dual_axis", False):
            if len(kpis) != 2:
                raise ValueError(
                    f"Chart '{chart_name}': Dual-axis charts must have exactly 2 KPIs. "
                    f"Found {len(kpis)}: {kpis}"
                )

            # If y_axis_titles provided, must match KPI count
            y_axis_titles = chart_config.get("y_axis_titles")
            if y_axis_titles and len(y_axis_titles) != len(kpis):
                raise ValueError(
                    f"Chart '{chart_name}': y_axis_titles count ({len(y_axis_titles)}) "
                    f"must match KPIS count ({len(kpis)})"
                )

    return True


# ============================================================================
# API ENDPOINTS
# ============================================================================

@app.get("/")
def root():
    return {
        "status": "RAN KPI Data Service - Universal Query System",
        "version": "3.0",
        "database": "MySQL"
    }

@app.get("/api/query-categories")
def get_query_categories():
    """Get all query categories (2G, 3G, 4G, GENERAL)"""
    return {
        "categories": list(QUERY_CONFIG.keys())
    }

@app.get("/api/query-subcategories/{category}")
def get_query_subcategories(category: str):
    """Get subcategories for a category (e.g., vendors for 2G)"""
    if category not in QUERY_CONFIG:
        raise HTTPException(status_code=404, detail=f"Category '{category}' not found")
    
    config = QUERY_CONFIG[category]
    
    # Check if this has subcategories (vendors) or direct queries
    first_key = next(iter(config.keys()))
    if isinstance(config[first_key], dict) and "description" in config[first_key]:
        # Direct queries (like GENERAL)
        return {
            "category": category,
            "has_subcategories": False,
            "queries": list(config.keys())
        }
    else:
        # Has subcategories (like 2G -> HUAWEI, ERICSSON)
        return {
            "category": category,
            "has_subcategories": True,
            "subcategories": list(config.keys())
        }

@app.get("/api/queries/{category}")
def get_queries_in_category(category: str, subcategory: Optional[str] = None):
    """Get available queries in a category/subcategory"""
    try:
        if subcategory:
            queries = QUERY_CONFIG[category][subcategory]
        else:
            queries = QUERY_CONFIG[category]
        
        return {
            "category": category,
            "subcategory": subcategory,
            "queries": [
                {
                    "name": name,
                    "description": config["description"]
                }
                for name, config in queries.items()
            ]
        }
    except KeyError:
        raise HTTPException(status_code=404, detail="Category/subcategory not found")

@app.get("/api/query-details/{category}/{query_name}")
def get_query_details(category: str, query_name: str, subcategory: Optional[str] = None):
    """Get detailed information about a specific query"""
    config = get_query_config(category, subcategory, query_name)
    query_type = config.get("query_type", "simple")

    # Validate chart_configs if present
    try:
        if "chart_configs" in config:
            validate_chart_configs(config)
    except ValueError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Invalid chart configuration: {str(e)}"
        )

    # Build response based on query type
    response = {
        "category": category,
        "subcategory": subcategory,
        "query_name": query_name,
        "query_type": query_type,
        "description": config["description"],
        "time_granularities": list(config["time_granularities"].keys()),
        "chart_configs": config.get("chart_configs", {})
    }

    # Add fields based on query type
    if query_type == "raw_sql":
        # For raw_sql, extract aggregation levels from sources
        sources = config.get("sources", {})
        # Get aggregation levels from first source (they should be consistent)
        first_source = next(iter(sources.values()), {})
        response["aggregation_levels"] = list(first_source.get("aggregations", {}).keys())
        response["sources"] = list(sources.keys())
        response["available_kpis"] = []  # raw_sql doesn't have kpi_fields
        response["date_time_filters"] = config.get("date_time_filters", ["start_date", "end_date"])
    else:
        # For other query types
        response["aggregation_levels"] = list(config.get("aggregation_levels", {}).keys())
        response["available_kpis"] = list(config.get("kpi_fields", {}).keys())
        response["date_time_filters"] = config.get("date_time_filters", [])

    return response

@app.post("/api/data/query")
def execute_query(request: QueryRequest):
    """Execute a query with dynamic parameters"""
    try:
        # Get query configuration
        config = get_query_config(
            request.query_category,
            request.query_subcategory,
            request.query_name
        )

        # Detect query type and route to appropriate builder
        query_type = config.get("query_type", "simple")

        if query_type == "multi_cte":
            sql_query = build_multi_cte_query(request, config)
        elif query_type == "raw_sql":
            sql_query = build_raw_sql_query(request, config)
            return
        else:
            # Default to simple query builder
            sql_query = build_dynamic_query(request, config)

        # Log the generated SQL for debugging
        print("\n" + "=" * 80)
        print("EXECUTING QUERY:")
        print(sql_query)
        print("=" * 80 + "\n")

        # Execute query
        results = db_manager.execute_query(sql_query)

        # Build chart metadata based on chart_configs and selected KPIs
        chart_metadata = {}
        chart_configs = config.get("chart_configs", {})

        if chart_configs:
            # Find chart configs that match the selected KPIs
            for chart_name, chart_config in chart_configs.items():
                chart_kpis = set(chart_config.get("KPIS", []))
                selected_kpis = set(request.selected_kpis)

                # Include chart if all its KPIs are in selected_kpis
                if chart_kpis.issubset(selected_kpis):
                    chart_metadata[chart_name] = chart_config

        return {
            "success": True,
            "query_category": request.query_category,
            "query_subcategory": request.query_subcategory,
            "query_name": request.query_name,
            "query_type": query_type,
            "row_count": len(results),
            "data": results,
            "metadata": {
                "chart_configs": chart_metadata,
                "aggregation_level": request.aggregation_level,
                "time_granularity": request.time_granularity
            },
            "query": sql_query  # Include query in response for debugging
        }

    except Exception as e:
        import traceback
        error_detail = {
            "error": str(e),
            "type": type(e).__name__,
            "traceback": traceback.format_exc()
        }
        print("\n" + "!" * 80)
        print("ERROR IN QUERY EXECUTION:")
        print(traceback.format_exc())
        print("!" * 80 + "\n")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/health")
def health_check():
    """Check database connection health"""
    try:
        db_manager.connect()
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        return {"status": "unhealthy", "error": str(e)}

@app.post("/api/export/excel")
def export_to_excel(request: QueryRequest):
    """Export query results to Excel file"""
    try:
        # Get query configuration
        config = get_query_config(
            request.query_category,
            request.query_subcategory,
            request.query_name
        )

        # Detect query type and route to appropriate builder
        query_type = config.get("query_type", "simple")

        if query_type == "multi_cte":
            sql_query = build_multi_cte_query(request, config)
        elif query_type == "raw_sql":
            sql_query = build_raw_sql_query(request, config)
        else:
            sql_query = build_dynamic_query(request, config)

        # Execute query
        results = db_manager.execute_query(sql_query)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "KPI Data"

        # Style for header
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        if results:
            headers = list(results[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Write data rows
            for row_num, row_data in enumerate(results, 2):
                for col_num, header in enumerate(headers, 1):
                    value = row_data.get(header)
                    ws.cell(row=row_num, column=col_num, value=value)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"KPI_Report_{request.query_category}_{timestamp}.xlsx"

        # Return as streaming response
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================================
# DYNAMIC SQL QUERY BUILDER
# ============================================================================

def build_subquery_cte(request: QueryRequest, subquery_config: Dict) -> str:
    """Build a single CTE (subquery) for multi-CTE queries"""
    cte_name = subquery_config["name"]
    table = subquery_config["table"]
    alias = subquery_config.get("alias", "m")

    # Build SELECT fields for this subquery
    fields = []

    # Get time granularity config for this subquery
    time_granularity = request.time_granularity
    time_config = subquery_config["time_granularities"].get(time_granularity, {})

    # Validate that this subquery supports the requested time granularity
    if not time_config or not time_config.get("is_available"):
        raise HTTPException(
            status_code=400,
            detail=f"Subquery '{cte_name}' does not support time granularity '{time_granularity}'. Available granularities: {list(subquery_config['time_granularities'].keys())}"
        )

    date_field = time_config.get("date_field")
    time_field = time_config.get("time_field")

    # Add date field
    if request.include_date and date_field:
        fields.append(f"{date_field} AS date")

    # Add time field if requested
    if request.include_time and time_field:
        fields.append(f"{time_field} AS time")

    # Add aggregation level fields
    if request.aggregation_level and subquery_config["aggregation_levels"]:
        agg_levels = subquery_config["aggregation_levels"]
        if request.aggregation_level in agg_levels:
            agg_expr = agg_levels[request.aggregation_level]
            fields.append(f"{agg_expr} AS {request.aggregation_level}")

    # Add KPI fields (all of them for this subquery)
    kpi_fields = subquery_config["kpi_fields"]
    for kpi_name, kpi_expr in kpi_fields.items():
        fields.append(f"{kpi_expr} AS {kpi_name}")

    select_clause = ",\n        ".join(fields)

    # Build WHERE conditions for this subquery
    where_conditions = []

    # Date filters
    if date_field:
        if request.multiple_date:
            dates_str = "', '".join(request.multiple_date)
            where_conditions.append(f"{date_field} IN ('{dates_str}')")
        elif request.start_date or request.end_date:
            if request.start_date:
                where_conditions.append(f"{date_field} >= '{request.start_date}'")
            if request.end_date:
                where_conditions.append(f"{date_field} <= '{request.end_date}'")

    # Time/Hour filters
    if time_field:
        if request.multiple_hour:
            hours_str = ", ".join(map(str, request.multiple_hour))
            where_conditions.append(f"HOUR({time_field}) IN ({hours_str})")
        elif request.start_hour is not None or request.end_hour is not None:
            if request.start_hour is not None:
                where_conditions.append(f"HOUR({time_field}) >= {request.start_hour}")
            if request.end_hour is not None:
                where_conditions.append(f"HOUR({time_field}) <= {request.end_hour}")

    where_clause = "\n        AND ".join(where_conditions) if where_conditions else "1=1"

    # Build GROUP BY for this subquery
    group_fields = []
    if request.include_date and date_field:
        group_fields.append(date_field)
    if request.include_time and time_field:
        group_fields.append(time_field)
    if request.aggregation_level and subquery_config["aggregation_levels"]:
        agg_levels = subquery_config["aggregation_levels"]
        if request.aggregation_level in agg_levels:
            group_fields.append(agg_levels[request.aggregation_level])

    group_by_clause = "GROUP BY " + ", ".join(group_fields) if group_fields else ""

    # Build the CTE SQL using the subquery's template
    sql_template = subquery_config["sql_template"]
    cte_sql = sql_template.format(
        fields=select_clause,
        table=table,
        alias=alias,
        where_conditions=where_clause,
        group_by=group_by_clause
    )

    return f"{cte_name} AS (\n    {cte_sql.strip()}\n)"


def build_cte_joins(cte_joins_config: List[Dict], first_cte: str, include_time: bool = True) -> str:
    """Build JOIN clauses between CTEs, filtering out time conditions if not needed"""
    # Start with the first CTE
    join_sql = first_cte

    for join_config in cte_joins_config:
        join_type = join_config["type"]
        right_cte = join_config["right"]
        on_conditions = join_config["on"]

        # Handle both string and list formats for ON conditions
        if isinstance(on_conditions, str):
            on_clause = on_conditions
        elif isinstance(on_conditions, list):
            # Filter out time-related conditions if time is not included
            filtered_conditions = []
            for condition in on_conditions:
                # Skip time conditions if include_time is False
                if not include_time and '.time' in condition.lower():
                    continue
                filtered_conditions.append(condition)

            on_clause = " AND ".join(filtered_conditions) if filtered_conditions else "1=1"
        else:
            on_clause = str(on_conditions)

        join_sql += f"\n    {join_type} {right_cte} ON {on_clause}"

    return join_sql


def build_multi_cte_query(request: QueryRequest, config: Dict) -> str:
    """Build multi-CTE SQL query dynamically"""

    # 1. Build all CTEs
    ctes = []
    for subquery_config in config["subqueries"]:
        cte_sql = build_subquery_cte(request, subquery_config)
        ctes.append(cte_sql)

    # 2. Build CTE joins (pass include_time to filter time conditions)
    first_cte = config["subqueries"][0]["name"]
    cte_joins_sql = build_cte_joins(config["cte_joins"], first_cte, request.include_time)

    # 3. Build outer SELECT fields
    outer_fields = []

    # Get time granularity for outer query
    time_config = config["time_granularities"].get(request.time_granularity, {})
    date_field = time_config.get("date_field")
    time_field = time_config.get("time_field")

    # Add date field if requested
    if request.include_date and date_field:
        outer_fields.append(f"{date_field} AS date")

    # Add time field if requested
    if request.include_time and time_field:
        outer_fields.append(f"{time_field} AS time")

    # Add aggregation level fields
    if request.aggregation_level:
        agg_levels = config.get("aggregation_levels", {})
        if request.aggregation_level in agg_levels:
            agg_expr = agg_levels[request.aggregation_level]
            if agg_expr:  # Only if not None
                outer_fields.append(f"{agg_expr} AS {request.aggregation_level}")

    # Add selected KPI fields
    kpi_fields = config["kpi_fields"]
    if request.selected_kpis:
        for kpi in request.selected_kpis:
            if kpi in kpi_fields:
                outer_fields.append(f"{kpi_fields[kpi]} AS {kpi}")
    else:
        # Add all KPIs if none selected
        for kpi_name, kpi_expr in kpi_fields.items():
            outer_fields.append(f"{kpi_expr} AS {kpi_name}")

    select_clause = ",\n    ".join(outer_fields)

    # 4. Build outer GROUP BY
    group_fields = []
    if request.include_date and date_field:
        group_fields.append(date_field)
    if request.include_time and time_field:
        group_fields.append(time_field)
    if request.aggregation_level:
        agg_levels = config.get("aggregation_levels", {})
        if request.aggregation_level in agg_levels:
            agg_expr = agg_levels[request.aggregation_level]
            if agg_expr:
                group_fields.append(agg_expr)

    group_by_clause = "GROUP BY " + ", ".join(group_fields) if group_fields else ""

    # 5. Build outer ORDER BY
    order_fields = []
    if request.include_date and date_field:
        order_fields.append(date_field)
    if request.include_time and time_field:
        order_fields.append(time_field)
    if not order_fields:
        order_fields.append("1")

    order_by_clause = "ORDER BY " + ", ".join(order_fields)

    # 6. Combine everything
    sql_template = config["sql_template"]
    with_clause = ",\n".join(ctes)

    final_query = f"""WITH {with_clause}
    {sql_template.format(
        fields=select_clause,
        cte_joins=cte_joins_sql,
        group_by=group_by_clause,
        order_by=order_by_clause
    ).strip()}"""

    return final_query


def build_dynamic_query(request: QueryRequest, config: Dict) -> str:
    """Build SQL query dynamically based on request and configuration"""
    
    # Get the SQL template
    sql_template = config["sql_template"]
    
    # 1. Build SELECT fields
    select_fields = build_select_fields(request, config)
    
    # 2. Build WHERE conditions
    where_conditions = build_where_clause(request, config)
    
    # 3. Build GROUP BY clause
    group_by_clause = build_group_by_clause(request, config)
    
    # 4. Build ORDER BY clause
    order_by_clause = build_order_by(request, config)
    
    # 5. Replace placeholders in SQL template
    sql_query = sql_template.format(
        fields=select_fields,
        where_conditions=where_conditions if where_conditions else "1=1",
        group_by=group_by_clause,
        order_by=order_by_clause
    )
    
    return sql_query.strip()

def build_select_fields(request: QueryRequest, config: Dict) -> str:
    """Build SELECT field list dynamically"""
    fields = []
    
    # Get time granularity config
    time_config = config["time_granularities"].get(request.time_granularity, {})
    date_field = time_config.get("date_field")
    time_field = time_config.get("time_field")
    
    # 1. Add date field if requested
    if request.include_date and date_field:
        fields.append(f"{date_field} AS date")
    
    # 2. Add time field if requested and available
    if request.include_time and time_field:
        fields.append(f"{time_field} AS time")
    
    # 3. Add aggregation level fields (hierarchical)
    if request.aggregation_level:
        agg_hierarchy = get_aggregation_hierarchy(
            config["aggregation_levels"],
            request.aggregation_level
        )
        for level_name, level_expr in agg_hierarchy:
            fields.append(f"{level_expr} AS {level_name}")
    
    # 4. Add selected KPI fields
    kpi_fields = config["kpi_fields"]
    if request.selected_kpis:
        for kpi in request.selected_kpis:
            if kpi in kpi_fields:
                fields.append(f"{kpi_fields[kpi]} AS {kpi}")
    else:
        # Add all KPIs if none selected
        for kpi_name, kpi_expr in kpi_fields.items():
            fields.append(f"{kpi_expr} AS {kpi_name}")
    
    return ",\n    ".join(fields)


def build_where_clause(request: QueryRequest, config: Dict) -> str:
    """Build WHERE conditions with date/time and additional filters"""
    conditions = []
    
    # Get time granularity config
    time_config = config["time_granularities"].get(request.time_granularity, {})
    date_field = time_config.get("date_field")
    time_field = time_config.get("time_field")
    
    # Date filters
    if date_field:
        if request.multiple_date:
            # Multiple specific dates
            dates_str = "', '".join(request.multiple_date)
            conditions.append(f"{date_field} IN ('{dates_str}')")
        elif request.start_date or request.end_date:
            # Date range
            if request.start_date:
                conditions.append(f"{date_field} >= '{request.start_date}'")
            if request.end_date:
                conditions.append(f"{date_field} <= '{request.end_date}'")
    
    # Time/Hour filters (if time field exists)
    if time_field:
        if request.multiple_hour:
            # Multiple specific hours
            hours_str = ", ".join(map(str, request.multiple_hour))
            conditions.append(f"HOUR({time_field}) IN ({hours_str})")
        elif request.start_hour is not None or request.end_hour is not None:
            # Hour range
            if request.start_hour is not None:
                conditions.append(f"HOUR({time_field}) >= {request.start_hour}")
            if request.end_hour is not None:
                conditions.append(f"HOUR({time_field}) <= {request.end_hour}")
    
    # Additional filters
    if request.filters:
        for filter_key, filter_values in request.filters.items():
            if filter_values:
                # Find the filter column in aggregation levels or use as-is
                agg_levels = config["aggregation_levels"]
                if filter_key in agg_levels:
                    filter_column = agg_levels[filter_key]
                else:
                    filter_column = filter_key
                
                escaped_values = [val.replace("'", "''") for val in filter_values]
                values_str = "', '".join(escaped_values)
                conditions.append(f"{filter_column} IN ('{values_str}')")
    
    if conditions:
        return "\n    AND ".join(conditions)
    return "1=1"  # Return default condition if no filters

def build_group_by_clause(request: QueryRequest, config: Dict) -> str:
    """Build GROUP BY clause matching SELECT fields"""
    group_fields = []
    
    # Get time granularity config
    time_config = config["time_granularities"].get(request.time_granularity, {})
    date_field = time_config.get("date_field")
    time_field = time_config.get("time_field")
    
    # 1. Add date field if included
    if request.include_date and date_field:
        group_fields.append(date_field)
    
    # 2. Add time field if included
    if request.include_time and time_field:
        group_fields.append(time_field)
    
    # 3. Add aggregation level fields (hierarchical)
    if request.aggregation_level:
        agg_hierarchy = get_aggregation_hierarchy(
            config["aggregation_levels"],
            request.aggregation_level
        )
        for level_name, level_expr in agg_hierarchy:
            group_fields.append(level_expr)
    
    # Only add GROUP BY if we have aggregation or date/time grouping with KPIs
    if group_fields and config["kpi_fields"]:
        return "GROUP BY " + ", ".join(group_fields)
    
    return ""

def build_order_by(request: QueryRequest, config: Dict) -> str:
    """Build ORDER BY clause"""
    order_fields = []
    
    # Get time granularity config
    time_config = config["time_granularities"].get(request.time_granularity, {})
    date_field = time_config.get("date_field")
    time_field = time_config.get("time_field")
    
    if request.include_date and date_field:
        order_fields.append(date_field)
    
    if request.include_time and time_field:
        order_fields.append(time_field)
    
    if not order_fields:
        order_fields.append("1")  # Order by first column
    
    return ", ".join(order_fields)


# ============================================================================
# RAW SQL QUERY BUILDER (new flexible query type)
# ============================================================================

def build_raw_sql_query(request: QueryRequest, config: Dict) -> str:
    """
    Build SQL query from raw template with source-based placeholders.
    Uses double underscore (__) for compatibility with str.format().

    Supports placeholders like:
    - {source__select_fields} - Date + Time + Aggregation fields
    - {source__where_clause} - Date/time filter conditions
    - {source__group_by} - GROUP BY fields
    - {source__order_by} - ORDER BY fields
    - {source__date} - Just the date column (transformed by granularity)
    - {source__time} - Just the time column (or empty if not HOURLY)
    - {source__aggregation} - Just the aggregation column
    - {source__time_join_condition} - For conditional JOIN clauses

    Example: {eric__select_fields}, {huawei__where_clause}
    """

    sources = config.get("sources", {})
    sql_template = config["sql_template"]
    granularity = request.time_granularity

    # Validate granularity is available
    time_granularities = config.get("time_granularities", {})
    if granularity not in time_granularities:
        raise HTTPException(
            status_code=400,
            detail=f"Time granularity '{granularity}' not available. Options: {list(time_granularities.keys())}"
        )

    if not time_granularities[granularity].get("is_available", True):
        raise HTTPException(
            status_code=400,
            detail=f"Time granularity '{granularity}' is not available for this query"
        )

    # Get global transform config
    transform_config = GRANULARITY_TRANSFORMS.get(granularity, GRANULARITY_TRANSFORMS["DAILY"])
    include_time = transform_config["include_time"]

    # --- Cross-source consistency validation ---
    # If something is needed, it must be defined in ALL sources — no silent skips.

    # date_col: required in all sources when include_date is True
    # (unless source has no_date: True)
    if request.include_date:
        missing_date = [n for n, c in sources.items() if not c.get("no_date") and not c.get("date_col")]
        if missing_date:
            raise HTTPException(
                status_code=400,
                detail=f"'include_date' is True but source(s) {missing_date} have no 'date_col' defined."
            )

    # time_col: required in all sources when granularity is HOURLY
    # (unless source has no_time: True)
    if include_time:
        missing_time = [n for n, c in sources.items() if not c.get("no_time") and not c.get("time_col")]
        if missing_time:
            raise HTTPException(
                status_code=400,
                detail=f"Granularity 'HOURLY' requested but source(s) {missing_time} have no 'time_col' defined."
            )

    # aggregation_level: if requested, must be defined in all sources
    # (unless source has no_aggregation: True)
    if request.aggregation_level:
        missing_agg = [
            n for n, c in sources.items()
            if not c.get("no_aggregation") and request.aggregation_level not in c.get("aggregations", {})
        ]
        if missing_agg:
            raise HTTPException(
                status_code=400,
                detail=f"Aggregation level '{request.aggregation_level}' requested but source(s) {missing_agg} don't have it defined in 'aggregations'."
            )

    # Build all placeholder values for each source
    placeholders = {}

    for source_name, source_config in sources.items():
        source_placeholders = _build_source_placeholders(
            source_name=source_name,
            source_config=source_config,
            request=request,
            transform_config=transform_config,
            include_time=include_time
        )
        placeholders.update(source_placeholders)

    # Build JOIN ON placeholders for each join pair
    joins = config.get("joins", [])
    for join_config in joins:
        join_placeholders = _build_join_on_placeholder(
            join_config=join_config,
            sources=sources,
            request=request,
            transform_config=transform_config,
            include_time=include_time
        )
        placeholders.update(join_placeholders)

    # Replace all placeholders using str.format()
    # Placeholders use double underscore: {eric__select_fields}, {huawei__where_clause}
    try:
        final_sql = sql_template.format(**placeholders)
    except KeyError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Missing placeholder in SQL template: {e}"
        )
    print(final_sql.strip())
    return final_sql.strip()


def _build_source_placeholders(
    source_name: str,
    source_config: Dict,
    request: QueryRequest,
    transform_config: Dict,
    include_time: bool
) -> Dict[str, str]:
    """Build all placeholder values for a single source."""

    placeholders = {}

    # Get base columns
    date_col = source_config.get("date_col")
    time_col = source_config.get("time_col")
    aggregations = source_config.get("aggregations", {})

    # Apply date transform
    date_transform = transform_config["date_transform"]
    transformed_date = date_transform.format(col=date_col) if date_col else ""

    # Time column (only for HOURLY)
    transformed_time = time_col if include_time and time_col else ""

    # Get aggregation column if requested
    agg_col = ""
    if request.aggregation_level and request.aggregation_level in aggregations:
        agg_col = aggregations[request.aggregation_level]

    # === Build individual placeholders (using __ instead of . for str.format compatibility) ===

    # {source__date} - transformed date column
    placeholders[f"{source_name}__date"] = transformed_date

    # {source__time} - time column or empty
    placeholders[f"{source_name}__time"] = transformed_time

    # {source__aggregation} - aggregation column or empty
    placeholders[f"{source_name}__aggregation"] = agg_col

    # {source__select_fields} - date, time, aggregation as SELECT fields
    select_parts = []
    if request.include_date and transformed_date:
        select_parts.append(f"{transformed_date} AS date")
    if request.include_time and transformed_time:
        select_parts.append(f"{transformed_time} AS time")
    if agg_col:
        select_parts.append(f"{agg_col} AS {request.aggregation_level}")
    placeholders[f"{source_name}__select_fields"] = ", ".join(select_parts) if select_parts else "1"

    # {source__where_clause} - date/time filter conditions
    where_parts = []
    if date_col:
        if request.multiple_date:
            dates_str = "', '".join(request.multiple_date)
            where_parts.append(f"{date_col} IN ('{dates_str}')")
        else:
            if request.start_date:
                where_parts.append(f"{date_col} >= '{request.start_date}'")
            if request.end_date:
                where_parts.append(f"{date_col} <= '{request.end_date}'")

    if include_time and time_col:
        if request.multiple_hour:
            hours_str = ", ".join(map(str, request.multiple_hour))
            where_parts.append(f"HOUR({time_col}) IN ({hours_str})")
        else:
            if request.start_hour is not None:
                where_parts.append(f"HOUR({time_col}) >= {request.start_hour}")
            if request.end_hour is not None:
                where_parts.append(f"HOUR({time_col}) <= {request.end_hour}")

    # Additional filters - IN (include these values)
    if request.filters:
        for filter_key, filter_values in request.filters.items():
            if filter_values and filter_key in aggregations:
                filter_col = aggregations[filter_key]
                escaped_values = [val.replace("'", "''") for val in filter_values]
                values_str = "', '".join(escaped_values)
                where_parts.append(f"{filter_col} IN ('{values_str}')")

    # Additional filters - NOT IN (exclude these values)
    if request.not_in_filters:
        for filter_key, filter_values in request.not_in_filters.items():
            if filter_values and filter_key in aggregations:
                filter_col = aggregations[filter_key]
                escaped_values = [val.replace("'", "''") for val in filter_values]
                values_str = "', '".join(escaped_values)
                where_parts.append(f"{filter_col} NOT IN ('{values_str}')")

    placeholders[f"{source_name}__where_clause"] = " AND ".join(where_parts) if where_parts else "1=1"

    # {source__group_by} - GROUP BY fields (without "GROUP BY" keyword)
    group_parts = []
    if request.include_date and transformed_date:
        group_parts.append(transformed_date)
    if request.include_time and transformed_time:
        group_parts.append(transformed_time)
    if agg_col:
        group_parts.append(agg_col)
    placeholders[f"{source_name}__group_by"] = ", ".join(group_parts) if group_parts else "1"

    # {source__order_by} - ORDER BY fields (without "ORDER BY" keyword)
    order_parts = []
    if request.include_date and transformed_date:
        order_parts.append(transformed_date)
    if request.include_time and transformed_time:
        order_parts.append(transformed_time)
    placeholders[f"{source_name}__order_by"] = ", ".join(order_parts) if order_parts else "1"

    # {source__time_join_condition} - for JOIN ON clauses, empty if not HOURLY
    placeholders[f"{source_name}__time_join_condition"] = ""

    return placeholders


def _build_join_on_placeholder(
    join_config: Dict,
    sources: Dict,
    request: QueryRequest,
    transform_config: Dict,
    include_time: bool
) -> Dict[str, str]:
    """
    Build JOIN ON placeholder for a pair of sources.

    Config example:
    {
        "left": "eric",
        "right": "huawei",
        "include_date": True,       # Always include date
        "include_time": True,       # Only if HOURLY granularity
        "include_aggregation": True, # Only if aggregation is selected
        "custom": []                 # Extra custom conditions
    }

    Generates placeholder: {left__right__join_on}
    Example: {eric__huawei__join_on} = "e.DATE = h.date AND e.TIME = h.time AND e.CELL_NAME = h.cell_name"
    """

    left_name = join_config.get("left")
    right_name = join_config.get("right")

    if not left_name or not right_name:
        return {}

    left_config = sources.get(left_name, {})
    right_config = sources.get(right_name, {})

    if not left_config or not right_config:
        return {}

    conditions = []

    # Date condition
    if join_config.get("include_date", True):
        left_date = left_config.get("date_col")
        right_date = right_config.get("date_col")

        if left_date and right_date:
            # Apply date transform for both
            date_transform = transform_config["date_transform"]
            left_date_transformed = date_transform.format(col=left_date)
            right_date_transformed = date_transform.format(col=right_date)
            conditions.append(f"{left_date_transformed} = {right_date_transformed}")

    # Time condition (only if HOURLY and configured)
    if join_config.get("include_time", True) and include_time:
        left_time = left_config.get("time_col")
        right_time = right_config.get("time_col")

        if left_time and right_time:
            conditions.append(f"{left_time} = {right_time}")

    # Aggregation condition (only if aggregation is selected and configured)
    if join_config.get("include_aggregation", True) and request.aggregation_level:
        left_aggs = left_config.get("aggregations", {})
        right_aggs = right_config.get("aggregations", {})

        left_agg_col = left_aggs.get(request.aggregation_level)
        right_agg_col = right_aggs.get(request.aggregation_level)

        if left_agg_col and right_agg_col:
            conditions.append(f"{left_agg_col} = {right_agg_col}")

    # Custom conditions (static additional conditions)
    custom_conditions = join_config.get("custom", [])
    for custom_cond in custom_conditions:
        conditions.append(custom_cond)

    # Build the placeholder key: {left__right__join_on}
    placeholder_key = f"{left_name}__{right_name}__join_on"

    # Join all conditions with AND
    join_on_value = " AND ".join(conditions) if conditions else "1=1"

    return {placeholder_key: join_on_value}


# ============================================================================
# RUN SERVER
# ============================================================================

if __name__ == "__main__":
    import uvicorn
    print("=" * 70)
    print("RAN KPI Data Service - Universal Query System")
    print("=" * 70)
    print(f"Database: {DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}")
    print(f"Query Categories: {list(QUERY_CONFIG.keys())}")
    print("=" * 70)
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=False)