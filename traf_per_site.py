import mysql.connector
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from utils import log_info , log_error

# Database configuration
db_config = {
    'host': "10.22.33.116",
    'user': "root",
    'password': "performance",
    'database': "performanceroute"
}

# Calculate date range (2 weeks ago to today)
end_date = datetime.now().strftime('%Y-%m-%d')
start_date = (datetime.now() - timedelta(weeks=2)).strftime('%Y-%m-%d')

start_date = datetime(2025,5,5).strftime('%Y-%m-%d')

log_info(f"Extracting data from {start_date} to {end_date}")

# # Define all queries with their sheet names
# queries = {
#     "HW_2G_Voice": f"""
#          SELECT 
#             t.date, 
#             COALESCE(s.site, 'UNKNOWN') AS SITE_NAME,
#             SUM(t.Total_Voice) AS Total_Voice
#         FROM hourly_huawei_trafic_per_cell_site_2g t
#         LEFT JOIN (
#             SELECT CELL, MIN(site) as site
#             FROM sites_and_info_2g_huawei
#             GROUP BY CELL
#         ) s ON t.CELL_NAME = s.CELL
#         WHERE t.date BETWEEN '{start_date}' AND '{end_date}'
#         GROUP BY t.date, s.site
#         ORDER BY t.date, s.site
#     """,
    
#     "HW_2G_Data": f"""
#         SELECT 
#             t.date, 
#             COALESCE(s.site, 'UNKNOWN') AS SITE_NAME,
#             SUM(t.Total_Data) AS Total_Data
#         FROM hourly_huawei_trafic_per_cell_site_2g t
#         LEFT JOIN (
#             SELECT CELL, MIN(site) as site
#             FROM sites_and_info_2g_huawei
#             GROUP BY CELL
#         ) s ON t.CELL_NAME = s.CELL
#         WHERE t.date BETWEEN '{start_date}' AND '{end_date}'
#         GROUP BY t.date, s.site
#         ORDER BY t.date, s.site
#     """,
    
#     "HW_3G_Voice": f"""
#         SELECT 
#             t.date, 
#             COALESCE(s.SITE, 'UNKNOWN') AS SITE_NAME,
#             SUM(t.Total_Voice) AS Total_Voice
#         FROM hourly_huawei_trafic_per_cell_site_3g t
#         LEFT JOIN (
#             SELECT ucell, MIN(SITE) as SITE
#             FROM sites_and_info_3g_huawei
#             GROUP BY ucell
#         ) s ON t.CELL_NAME = s.ucell
#         WHERE t.date BETWEEN '{start_date}' AND '{end_date}'
#         GROUP BY t.date, s.SITE
#         ORDER BY t.date, s.SITE
#     """,
    
#     "HW_3G_Data": f"""
#         SELECT 
#             t.date, 
#             COALESCE(s.SITE, 'UNKNOWN') AS SITE_NAME,
#             SUM(t.Total_data) AS Total_data
#         FROM hourly_huawei_trafic_per_cell_site_3g t
#         LEFT JOIN (
#             SELECT ucell, MIN(SITE) as SITE
#             FROM sites_and_info_3g_huawei
#             GROUP BY ucell
#         ) s ON t.CELL_NAME = s.ucell
#         WHERE t.date BETWEEN '{start_date}' AND '{end_date}'
#         GROUP BY t.date, s.SITE
#         ORDER BY t.date, s.SITE
#     """,
    
#     "HW_4G_Data": f"""
#         SELECT 
#             t.date, 
#             COALESCE(s.SITE, 'UNKNOWN') AS SITE_NAME,
#             SUM(t.Total_Data) AS Total_Data
#         FROM hourly_huawei_trafic_per_cell_site_4g t
#         LEFT JOIN (
#             SELECT cell, MIN(SITE) as SITE
#             FROM sites_and_info_4g_huawei
#             GROUP BY cell
#         ) s ON t.CELL_NAME = s.cell
#         WHERE t.date BETWEEN '{start_date}' AND '{end_date}'
#         GROUP BY t.date, s.SITE
#         ORDER BY t.date, s.SITE
#     """,
    
#     "Ericsson_2G_Data": f"""
#         SELECT 
#             date_id AS date, 
#             SITE_NAME,
#             SUM(Total_Data_Volume) AS Total_Data_Volume
#         FROM hourly_ericsson_trafic_per_site_2g
#         WHERE date_id BETWEEN '{start_date}' AND '{end_date}'
#         GROUP BY date_id, SITE_NAME
#         ORDER BY date_id, SITE_NAME
#     """,
    
#     "Ericsson_2G_Voice": f"""
#         SELECT 
#             date, 
#             SITE_NAME,
#             SUM(trafic2gVoix) AS trafic2gVoix
#         FROM hourly_ericsson_trafic_per_site_2g_voix
#         WHERE date BETWEEN '{start_date}' AND '{end_date}'
#         GROUP BY date, SITE_NAME
#         ORDER BY date, SITE_NAME
#     """,
    
#     "Ericsson_3G_Data": f"""
#         SELECT 
#             date_id AS date, 
#             RBS_Id AS SITE_NAME,
#             SUM(Total_Data_Volume) AS Total_Data_Volume
#         FROM hourly_ericsson_trafic_per_site_3g 
#         WHERE date_id BETWEEN '{start_date}' AND '{end_date}'
#         GROUP BY date_id, RBS_Id
#         ORDER BY date_id, RBS_Id
#     """,
    
#     "Ericsson_3G_Voice": f"""
#         SELECT 
#             date, 
#             RBS AS site_name, 
#             SUM(trafic_3g_voix) AS trafic_3g_voix
#         FROM hourly_ericsson_trafic_per_site_3g_voix 
#         WHERE date BETWEEN '{start_date}' AND '{end_date}'
#         GROUP BY date, RBS
#         ORDER BY date, RBS
#     """,
    
#     "Ericsson_4G_Data": f"""
#         SELECT 
#             date, 
#             ERBS_Id AS SITE_NAME,
#             SUM(valeur) AS total
#         FROM ericsson4g_data 
#         WHERE date BETWEEN '{start_date}' AND '{end_date}'
#         GROUP BY date, ERBS_Id
#         ORDER BY date, ERBS_Id
#     """
# }


queries = {
    "HW_2G_Voice": f"""
        SELECT 
            DATE_FORMAT(t.date, '%Y-%m') AS month, 
            COALESCE(s.site, 'UNKNOWN') AS SITE_NAME,
            SUM(t.Total_Voice) AS Total_Voice
        FROM hourly_huawei_trafic_per_cell_site_2g t
        LEFT JOIN (
            SELECT CELL, MIN(site) as site
            FROM sites_and_info_2g_huawei
            GROUP BY CELL
        ) s ON t.CELL_NAME = s.CELL
        WHERE t.date BETWEEN '{start_date}' AND '{end_date}'
        GROUP BY DATE_FORMAT(t.date, '%Y-%m'), s.site
        ORDER BY month, s.site
    """,
    
    "HW_2G_Data": f"""
        SELECT 
            DATE_FORMAT(t.date, '%Y-%m') AS month, 
            COALESCE(s.site, 'UNKNOWN') AS SITE_NAME,
            SUM(t.Total_Data) AS Total_Data
        FROM hourly_huawei_trafic_per_cell_site_2g t
        LEFT JOIN (
            SELECT CELL, MIN(site) as site
            FROM sites_and_info_2g_huawei
            GROUP BY CELL
        ) s ON t.CELL_NAME = s.CELL
        WHERE t.date BETWEEN '{start_date}' AND '{end_date}'
        GROUP BY DATE_FORMAT(t.date, '%Y-%m'), s.site
        ORDER BY month, s.site
    """,
    
    "HW_3G_Voice": f"""
        SELECT 
            DATE_FORMAT(t.date, '%Y-%m') AS month, 
            COALESCE(s.SITE, 'UNKNOWN') AS SITE_NAME,
            SUM(t.Total_Voice) AS Total_Voice
        FROM hourly_huawei_trafic_per_cell_site_3g t
        LEFT JOIN (
            SELECT ucell, MIN(SITE) as SITE
            FROM sites_and_info_3g_huawei
            GROUP BY ucell
        ) s ON t.CELL_NAME = s.ucell
        WHERE t.date BETWEEN '{start_date}' AND '{end_date}'
        GROUP BY DATE_FORMAT(t.date, '%Y-%m'), s.SITE
        ORDER BY month, s.SITE
    """,
    
    "HW_3G_Data": f"""
        SELECT 
            DATE_FORMAT(t.date, '%Y-%m') AS month, 
            COALESCE(s.SITE, 'UNKNOWN') AS SITE_NAME,
            SUM(t.Total_Data) AS Total_Data
        FROM hourly_huawei_trafic_per_cell_site_3g t
        LEFT JOIN (
            SELECT ucell, MIN(SITE) as SITE
            FROM sites_and_info_3g_huawei
            GROUP BY ucell
        ) s ON t.CELL_NAME = s.ucell
        WHERE t.date BETWEEN '{start_date}' AND '{end_date}'
        GROUP BY DATE_FORMAT(t.date, '%Y-%m'), s.SITE
        ORDER BY month, s.SITE
    """,
    
    "HW_4G_Data": f"""
        SELECT 
            DATE_FORMAT(t.date, '%Y-%m') AS month, 
            COALESCE(s.SITE, 'UNKNOWN') AS SITE_NAME,
            SUM(t.Total_Data) AS Total_Data
        FROM hourly_huawei_trafic_per_cell_site_4g t
        LEFT JOIN (
            SELECT cell, MIN(SITE) as SITE
            FROM sites_and_info_4g_huawei
            GROUP BY cell
        ) s ON t.CELL_NAME = s.cell
        WHERE t.date BETWEEN '{start_date}' AND '{end_date}'
        GROUP BY DATE_FORMAT(t.date, '%Y-%m'), s.SITE
        ORDER BY month, s.SITE
    """,
    
    "Ericsson_2G_Data": f"""
        SELECT 
            DATE_FORMAT(date_id, '%Y-%m') AS month, 
            SITE_NAME,
            SUM(Total_Data_Volume) AS Total_Data_Volume
        FROM hourly_ericsson_trafic_per_site_2g
        WHERE date_id BETWEEN '{start_date}' AND '{end_date}'
        GROUP BY DATE_FORMAT(date_id, '%Y-%m'), SITE_NAME
        ORDER BY month, SITE_NAME
    """,
    
    "Ericsson_2G_Voice": f"""
        SELECT 
            DATE_FORMAT(date, '%Y-%m') AS month, 
            SITE_NAME,
            SUM(trafic2gVoix) AS trafic2gVoix
        FROM hourly_ericsson_trafic_per_site_2g_voix
        WHERE date BETWEEN '{start_date}' AND '{end_date}'
        GROUP BY DATE_FORMAT(date, '%Y-%m'), SITE_NAME
        ORDER BY month, SITE_NAME
    """,
    
    "Ericsson_3G_Data": f"""
        SELECT 
            DATE_FORMAT(date_id, '%Y-%m') AS month, 
            RBS_Id AS SITE_NAME,
            SUM(Total_Data_Volume) AS Total_Data_Volume
        FROM hourly_ericsson_trafic_per_site_3g 
        WHERE date_id BETWEEN '{start_date}' AND '{end_date}'
        GROUP BY DATE_FORMAT(date_id, '%Y-%m'), RBS_Id
        ORDER BY month, RBS_Id
    """,
    
    "Ericsson_3G_Voice": f"""
        SELECT 
            DATE_FORMAT(date, '%Y-%m') AS month, 
            RBS AS site_name, 
            SUM(trafic_3g_voix) AS trafic_3g_voix
        FROM hourly_ericsson_trafic_per_site_3g_voix 
        WHERE date BETWEEN '{start_date}' AND '{end_date}'
        GROUP BY DATE_FORMAT(date, '%Y-%m'), RBS
        ORDER BY month, RBS
    """,
    
    "Ericsson_4G_Data": f"""
        SELECT 
            DATE_FORMAT(date, '%Y-%m') AS month, 
            ERBS_Id AS SITE_NAME,
            SUM(valeur) AS total
        FROM ericsson4g_data 
        WHERE date BETWEEN '{start_date}' AND '{end_date}'
        GROUP BY DATE_FORMAT(date, '%Y-%m'), ERBS_Id
        ORDER BY month, ERBS_Id
    """
}

def extract_and_export():
    """Extract data from database and export to Excel with multiple sheets"""
    
    # Connect to database
    try:
        conn = mysql.connector.connect(**db_config)
        log_info("Connected to database successfully!")
    except Exception as e:
        log_info(f"Error connecting to database: {e}")
        return
    
    output_file = f"Traffic_Data_{start_date}_to_{end_date}.xlsx"
    
    try:
        # Try to load existing workbook or create new one
        try:
            from openpyxl import load_workbook
            wb = load_workbook(output_file)
            log_info(f"Continuing with existing file: {output_file}")
        except FileNotFoundError:
            from openpyxl import Workbook
            wb = Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            log_info(f"Creating new file: {output_file}")
        
        for sheet_name, query in queries.items():
            log_info(f"Processing: {sheet_name}...")
            
            try:
                # Execute query and load into DataFrame
                df = pd.read_sql(query, conn)
                
                # Remove existing sheet if it exists
                if sheet_name in wb.sheetnames:
                    wb.remove(wb[sheet_name])
                
                # Create new sheet
                ws = wb.create_sheet(sheet_name)
                
                # Write DataFrame to sheet
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Save after each successful query
                wb.save(output_file)
                
                log_info(f"  ✓ {sheet_name}: {len(df)} rows exported and saved")
                
            except Exception as e:
                log_info(f"  ✗ Error processing {sheet_name}: {e}")
                log_info(f"  → Continuing with next query...")
                continue
        
        log_info(f"\n✓ All data exported successfully to: {output_file}")
        log_info(f"✓ Total sheets created: {len(wb.sheetnames)}")
        
    except Exception as e:
        log_error(f"Error with Excel file: {e}")
    
    finally:
        # Close database connection
        conn.close()
        log_info("Database connection closed.")

if __name__ == "__main__":
    try:
        extract_and_export()
    except KeyboardInterrupt:
        log_error("\n\n⚠️  Process interrupted by user (Ctrl+C)")
        log_error("Exiting...")
        exit(0)