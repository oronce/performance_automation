"""
Batch Alarm Impact Analysis

Reads alarm logs and automatically analyzes traffic impact for each SOFTWARE ERROR alarm.
Creates one Excel file with multiple sheets - one per alarm.
"""

import pandas as pd
import mysql.connector
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# Database configuration
DB_CONFIG = {
    'host': '10.22.33.116',
    'user': 'root',
    'password': 'performance',
    'database': 'performanceroute'
}


def get_controller_traffic(alarm_date, start_hour, controllers):
    """Retrieve controller traffic for alarm day and 4 previous weeks."""

    alarm_dt = datetime.strptime(alarm_date, '%Y-%m-%d')
    week1_dt = alarm_dt - timedelta(days=7)
    week2_dt = alarm_dt - timedelta(days=14)
    week3_dt = alarm_dt - timedelta(days=21)
    week4_dt = alarm_dt - timedelta(days=28)

    dates = [
        alarm_dt.strftime('%Y-%m-%d'),
        week1_dt.strftime('%Y-%m-%d'),
        week2_dt.strftime('%Y-%m-%d'),
        week3_dt.strftime('%Y-%m-%d'),
        week4_dt.strftime('%Y-%m-%d')
    ]

    controller_list_str = "', '".join(controllers)
    controller_filter = f"AND controler IN ('{controller_list_str}')"
    dates_str = "', '".join(dates)

    query = f"""
    WITH controller_raw AS (
        SELECT
            DATE,
            HOUR(STR_TO_DATE(heur, '%H:%i:%s')) AS heur,
            bsc AS controler,
            ROUND(SUM(CAST(valeur AS DECIMAL(10,2))), 2) AS total_voice_controller,
            'BSC' AS node_type
        FROM ericsson2g_voix
        WHERE date IN ('{dates_str}')
          AND HOUR(STR_TO_DATE(heur, '%H:%i:%s')) BETWEEN {start_hour} AND 23
        GROUP BY DATE, HOUR(STR_TO_DATE(heur, '%H:%i:%s')), bsc

        UNION ALL

        SELECT
            DATE,
            HOUR(STR_TO_DATE(heur, '%H:%i:%s')),
            rnc AS controler,
            ROUND(SUM(CAST(valeur AS DECIMAL(10,2))), 2),
            'RNC' AS node_type
        FROM ericsson3g_voix
        WHERE date IN ('{dates_str}')
          AND HOUR(STR_TO_DATE(heur, '%H:%i:%s')) BETWEEN {start_hour} AND 23
        GROUP BY DATE, HOUR(STR_TO_DATE(heur, '%H:%i:%s')), rnc

        UNION ALL

        SELECT
            date,
            HOUR(STR_TO_DATE(time, '%H:%i:%s')),
            controller AS controler,
            ROUND(SUM(Total_Voice), 2),
            'BSC' AS node_type
        FROM hourly_huawei_trafic_per_cell_site_2g
        WHERE date IN ('{dates_str}')
          AND HOUR(STR_TO_DATE(time, '%H:%i:%s')) BETWEEN {start_hour} AND 23
        GROUP BY date, HOUR(STR_TO_DATE(time, '%H:%i:%s')), controller

        UNION ALL

        SELECT
            date,
            HOUR(STR_TO_DATE(time, '%H:%i:%s')),
            controller AS controler,
            ROUND(SUM(Total_Voice), 2),
            'RNC' AS node_type
        FROM hourly_huawei_trafic_per_cell_site_3g
        WHERE date IN ('{dates_str}')
          AND HOUR(STR_TO_DATE(time, '%H:%i:%s')) BETWEEN {start_hour} AND 23
        GROUP BY date, HOUR(STR_TO_DATE(time, '%H:%i:%s')), controller
    )
    SELECT
        DATE,
        heur,
        controler,
        node_type,
        SUM(total_voice_controller) AS traffic
    FROM controller_raw
    WHERE 1=1
      {controller_filter}
    GROUP BY DATE, heur, controler, node_type
    ORDER BY controler, DATE, heur
    """

    conn = mysql.connector.connect(**DB_CONFIG)
    df = pd.read_sql(query, conn)
    conn.close()

    if not df.empty:
        df['DATE'] = df['DATE'].astype(str)

    return df, dates


def analyze_impact(df, alarm_date, dates):
    """Calculate 4-week average and compare with alarm day."""

    alarm_data = df[df['DATE'] == alarm_date].copy()
    historical_data = df[df['DATE'] != alarm_date].copy()

    avg_data = historical_data.groupby(['heur', 'controler', 'node_type'])['traffic'].mean().reset_index()
    avg_data.rename(columns={'traffic': 'avg_traffic_4weeks'}, inplace=True)

    week1_data = df[df['DATE'] == dates[1]][['heur', 'controler', 'node_type', 'traffic']].copy()
    week1_data.rename(columns={'traffic': 'week1_traffic'}, inplace=True)

    week2_data = df[df['DATE'] == dates[2]][['heur', 'controler', 'node_type', 'traffic']].copy()
    week2_data.rename(columns={'traffic': 'week2_traffic'}, inplace=True)

    week3_data = df[df['DATE'] == dates[3]][['heur', 'controler', 'node_type', 'traffic']].copy()
    week3_data.rename(columns={'traffic': 'week3_traffic'}, inplace=True)

    week4_data = df[df['DATE'] == dates[4]][['heur', 'controler', 'node_type', 'traffic']].copy()
    week4_data.rename(columns={'traffic': 'week4_traffic'}, inplace=True)

    comparison = alarm_data.merge(avg_data, on=['heur', 'controler', 'node_type'], how='left')
    comparison = comparison.merge(week1_data, on=['heur', 'controler', 'node_type'], how='left')
    comparison = comparison.merge(week2_data, on=['heur', 'controler', 'node_type'], how='left')
    comparison = comparison.merge(week3_data, on=['heur', 'controler', 'node_type'], how='left')
    comparison = comparison.merge(week4_data, on=['heur', 'controler', 'node_type'], how='left')

    comparison['difference'] = comparison['traffic'] - comparison['avg_traffic_4weeks']
    comparison['pct_change'] = ((comparison['traffic'] - comparison['avg_traffic_4weeks']) /
                                 comparison['avg_traffic_4weeks'] * 100)

    comparison['avg_traffic_4weeks'] = comparison['avg_traffic_4weeks'].round(2)
    comparison['week1_traffic'] = comparison['week1_traffic'].round(2)
    comparison['week2_traffic'] = comparison['week2_traffic'].round(2)
    comparison['week3_traffic'] = comparison['week3_traffic'].round(2)
    comparison['week4_traffic'] = comparison['week4_traffic'].round(2)
    comparison['difference'] = comparison['difference'].round(2)
    comparison['pct_change'] = comparison['pct_change'].round(2)

    comparison.rename(columns={
        'DATE': 'alarm_date',
        'heur': 'hour',
        'traffic': 'alarm_traffic'
    }, inplace=True)

    comparison = comparison[['alarm_date', 'hour', 'controler', 'node_type',
                            'alarm_traffic', 'avg_traffic_4weeks',
                            'week1_traffic', 'week2_traffic', 'week3_traffic', 'week4_traffic',
                            'difference', 'pct_change']]

    comparison = comparison.sort_values(['controler', 'hour']).reset_index(drop=True)

    return comparison


def create_chart(ws, data_start_row, data_end_row):
    """Create a line chart for alarm traffic vs avg traffic."""

    chart = LineChart()
    chart.title = "Traffic Comparison"
    chart.y_axis.title = "Traffic (Erlang)"
    chart.x_axis.title = "Hour"

    # Set Y-axis to start at 0
    chart.y_axis.scaling.min = 0

    # Remove gridlines
    chart.y_axis.majorGridlines = None
    chart.x_axis.majorGridlines = None

    # Alarm traffic
    alarm_data = Reference(ws, min_col=5, min_row=data_start_row, max_row=data_end_row)
    # Avg traffic
    avg_data = Reference(ws, min_col=6, min_row=data_start_row, max_row=data_end_row)
    # Hours (categories)
    hours = Reference(ws, min_col=2, min_row=data_start_row+1, max_row=data_end_row)

    chart.add_data(alarm_data, titles_from_data=True)
    chart.add_data(avg_data, titles_from_data=True)
    chart.set_categories(hours)

    return chart


def process_alarms():
    """Process all SOFTWARE ERROR alarms and create Excel with multiple sheets."""

    # Read alarm file
    alarm_file = r'c:\Users\o84432318\Desktop\project\nt_report_perf\performance_automation\assets\Copy of SOFTWARE ERROR Alarms logs.xlsx'
    df_alarms = pd.read_excel(alarm_file, sheet_name=0)

    # Filter SOFTWARE ERROR only (exclude APPLICATION DETECTED SOFTWARE ERROR)
    df_alarms = df_alarms[df_alarms['Alarm Name'] == 'SOFTWARE ERROR'].copy()

    # Exclude HAMBC01 and BHMBC02 (not BSC)
    df_alarms = df_alarms[~df_alarms['Alarm Source'].isin(['HAMBC01', 'BHMBC02'])]

    print(f"Found {len(df_alarms)} SOFTWARE ERROR alarms to process")

    # Create Excel workbook
    output_dir = 'excel_output'
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for idx, row in df_alarms.iterrows():
        controller = row['Alarm Source']
        first_occurred = pd.to_datetime(row['First Occurred On'])
        alarm_date = first_occurred.strftime('%Y-%m-%d')
        start_hour = first_occurred.hour

        sheet_name = f"{controller}_{alarm_date}_{start_hour}h"[:31]  # Excel sheet name limit

        print(f"\nProcessing: {controller} - {alarm_date} at {start_hour}:00")

        try:
            # Get traffic data
            df_traffic, dates = get_controller_traffic(alarm_date, start_hour, [controller])

            if df_traffic.empty:
                print(f"  No data found for {controller}")
                continue

            # Analyze impact
            comparison = analyze_impact(df_traffic, alarm_date, dates)

            # Create sheet
            ws = wb.create_sheet(title=sheet_name)

            # Write data to sheet
            for r_idx, row_data in enumerate(dataframe_to_rows(comparison, index=False, header=True), 1):
                for c_idx, value in enumerate(row_data, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Create chart
            if len(comparison) > 0:
                chart = create_chart(ws, 1, len(comparison) + 1)
                ws.add_chart(chart, f"N2")

            print(f"  [OK] Added sheet: {sheet_name}")

        except Exception as e:
            print(f"  [ERROR] Error processing {controller}: {str(e)}")
            continue

    # Save workbook
    output_file = os.path.join(output_dir, f'batch_alarm_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    wb.save(output_file)

    print(f"\n{'='*80}")
    print(f"Analysis complete!")
    print(f"Output file: {output_file}")
    print(f"{'='*80}")


if __name__ == "__main__":
    process_alarms()
