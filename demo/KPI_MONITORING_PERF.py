import pandas as pd
import mysql.connector
from openpyxl import load_workbook

import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from nt_report_perf.abstract_class.base_report.Excel_report import utils


def fill_to_excel_native3(file_path=None, sheet_name=None,
        df=None, start_row=1, start_col=1, start_coordinate_cell_letter=None,
        withIndex=False, withColumn=True, worksheet=None,
        str_message_in_case_of_empty_df="No Data"):
    """
    fill data to excel version3 — preserves original columns when df is empty.
    """
    # in the case df is empty: keep original columns, write message in first cell
    if df.empty:
        empty_row = {col: np.NaN for col in df.columns}
        if len(df.columns) > 0:
            empty_row[df.columns[0]] = str_message_in_case_of_empty_df
        df = pd.DataFrame([empty_row])

    # get row and column number from letter_coordinate if provided (e.g. "A2")
    if start_coordinate_cell_letter:
        col_and_row = coordinate_from_string(start_coordinate_cell_letter)
        start_col = column_index_from_string(col_and_row[0])
        start_row = col_and_row[1]

    # if worksheet is not defined, load it from workbook
    book = None
    if not worksheet:
        book = load_workbook(file_path)
        worksheet = book[sheet_name]

    df_array_rows = utils.convert_pandas_dataframe_to_array(df, with_index=withIndex, with_columns=withColumn)
    print(df_array_rows)

    reference_for_chart = utils.get_dataframe_reference_for_chart_creating(
        df, with_index=withIndex, start_row=start_row, start_col=start_col, worksheet=worksheet
    )

    # fill excel cell by cell
    actual_row = start_row
    for row in df_array_rows:
        actual_col = start_col
        for element in row:
            coordinate = f"{get_column_letter(actual_col)}{actual_row}"
            worksheet[coordinate] = element
            actual_col += 1
        actual_row += 1

    if book:
        book.save(file_path)

    return reference_for_chart

# Query generator functions

def get_2g_query(granularity='HOURLY'):
    """Generate 2G query based on granularity (HOURLY or DAILY)"""
    if granularity == 'HOURLY':
        time_field = "T1.TIME,"
        time_group = "T.TIME,"
        time_group2 = "H.TIME,"
        time_join = "AND T1.TIME = T2.TIME"
    else:
        time_field = ""
        time_group = ""
        time_group2 = ""
        time_join = ""

    return f"""
        SELECT
            T1.DATE,
            {time_field}
            T1.SITE_NAME,
            T1.CELL_NAME,
            ROUND(T1.CSSR_HUAWEI,2) `2G CSSR`,
            ROUND(T1.CDR_HUAWEI,2) `2G Call Drop_rate`,
            ROUND(T1.CBR_HUAWEI,2) `2G Call Block_rate`,
            T2.voix_2g_huawei VOICE_TRAFIC_2G,
            T2.DATA2g DATA_TRAFIC_2G
        FROM
            (
                SELECT
                    T.DATE,
                    {time_group.replace('T.TIME,', 'T.TIME,')}
                    S.site AS SITE_NAME,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(T.UCELL, '=', -3), ',', 1) AS CELL_NAME,
                    100 * (SUM(T.Suc_SDCCH_Seiz) / SUM(T.SDCCH_Seiz_Req)) *
                    (1 - (SUM(T.SDCCHDrop) / SUM(T.Suc_SDCCH_Seiz))) *
                    (SUM(T.Suc_TCH_Seiz_SG) + SUM(T.Suc_TCH_Seiz_TrafChann) + SUM(T.Suc_TCH_Seiz_handTrafChan)) /
                    (SUM(T.TCH_Seizure_Req_SC) + SUM(T.TCH_Seiz_Req_TrafChan) + SUM(T.TCH_Seiz_Req_HandTrafChan)) *
                    (1 - (SUM(T.TCHDrop_Nume_ARCEP) / SUM(T.TCHDrop_Deno_ARCEP))) AS CSSR_HUAWEI,
                    (100 * SUM(T.TCHDrop_Nume_ARCEP) / SUM(T.TCHDrop_Deno_ARCEP)) AS CDR_HUAWEI,
                    100 * (SUM(T.TCHCONG_Nume_ARCEP) / SUM(T.TCHCONG_Deno_ARCEP)) AS CBR_HUAWEI
                FROM
                    hourly_arcep_huawei_2g AS T
                INNER JOIN
                    sites_and_info_2g_huawei AS S ON S.Gcell = T.UCELL
                WHERE
                    T.DATE BETWEEN '{{start_date}}' AND '{{end_date}}'
                    AND FIND_IN_SET(S.site, '{{sites}}')
                GROUP BY
                    T.DATE, {time_group}S.site, CELL_NAME
            ) AS T1
        INNER JOIN
            (
                SELECT
                    H.DATE AS DATE,
                    {time_group2.replace('H.TIME,', 'H.TIME,')}
                    S.SITE AS SITE_NAME,
                    H.CELL_NAME,
                    ROUND(SUM(H.Total_Voice), 2) AS voix_2g_huawei,
                    ROUND(SUM(H.Total_Data), 2) AS DATA2g
                FROM
                    hourly_huawei_trafic_per_cell_site_2g AS H
                INNER JOIN
                    sites_and_info_2g_huawei AS S ON S.cell = H.CELL_NAME
                WHERE
                    H.DATE BETWEEN '{{start_date}}' AND '{{end_date}}'
                    AND FIND_IN_SET(S.site, '{{sites}}')
                GROUP BY
                    H.DATE, {time_group2}S.SITE, H.CELL_NAME
            ) AS T2
        ON
            T1.DATE = T2.DATE
            {time_join}
            AND T1.SITE_NAME = T2.SITE_NAME
            AND T1.CELL_NAME = T2.CELL_NAME
    """


def get_3g_query(granularity='HOURLY'):
    """Generate 3G query based on granularity (HOURLY or DAILY)"""
    if granularity == 'HOURLY':
        time_field = "T1.time,"
        time_group_t1 = "h1.time,"
        time_group_t2 = "h2.time,"
        time_join = "AND T1.time = T2.time"
    else:
        time_field = ""
        time_group_t1 = ""
        time_group_t2 = ""
        time_join = ""

    return f"""
        SELECT
            T1.date,
            {time_field}
            T1.SITE_NAME,
            T1.`3G PS_DROP`,
            T1.`3G CS_CSSR`,
            T1.`3G CALL BLOCK RATE`,
            T1.`Total_PS_Traffic_GB`  AS `3G Traffic Data`,
            T2.`3G PS_CSSR`,
            T2.`3G Downlink Throughput (kbps)`,
            T2.`3G Uplink Throughput (kbps)`,
            T2.`3G Traffic_Voix`
        FROM
            (
                -- T1: From hourly_huawei_3g_all_counters_1
                SELECT
                    date,
                    {time_group_t1}
                    s.SITE AS SITE_NAME,
                    CASE
                        WHEN (COALESCE(SUM(VS_RAB_NormRel_PS),0) + COALESCE(SUM(VS_RAB_AbnormRel_PS),0) -
                              COALESCE(SUM(VS_RAB_AbnormRel_PS_PCH),0) - COALESCE(SUM(VS_RAB_NormRel_PS_PCH),0) +
                              COALESCE(SUM(VS_DCCC_D2P_Succ),0) + COALESCE(SUM(VS_DCCC_Succ_F2P),0)) = 0
                        THEN NULL
                        ELSE ROUND((100 *
                            (COALESCE(SUM(VS_RAB_AbnormRel_PS),0) - COALESCE(SUM(VS_RAB_AbnormRel_PS_PCH),0) -
                             COALESCE(SUM(VS_RAB_AbnormRel_PS_D2P),0) - COALESCE(SUM(VS_RAB_AbnormRel_PS_F2P),0))
                        ) /
                        (
                            COALESCE(SUM(VS_RAB_NormRel_PS),0) + COALESCE(SUM(VS_RAB_AbnormRel_PS),0) -
                            COALESCE(SUM(VS_RAB_AbnormRel_PS_PCH),0) - COALESCE(SUM(VS_RAB_NormRel_PS_PCH),0) +
                            COALESCE(SUM(VS_DCCC_D2P_Succ),0) + COALESCE(SUM(VS_DCCC_Succ_F2P),0)
                        ), 3)
                    END AS `3G PS_DROP`,
                    100 * (
                        (SUM(RRC_SuccConnEstab_CallReEst) + SUM(RRC_SuccConnEstab_EmgCall) +
                         SUM(RRC_SuccConnEstab_OrgConvCall) + SUM(RRC_SuccConnEstab_OrgSubCall) +
                         SUM(RRC_SuccConnEstab_TmConvCall) + SUM(RRC_SuccConnEstab_Unkown)) /
                        NULLIF((SUM(RRC_AttConnEstab_CallReEst) + SUM(RRC_AttConnEstab_EmgCall) +
                         SUM(RRC_AttConnEstab_OrgSubCall) + SUM(RRC_AttConnEstab_OrgConvCall) +
                         SUM(RRC_AttConnEstab_TmConvCall) + SUM(RRC_AttConnEstab_Unknown)), 0)
                    ) * (
                        (SUM(VS_HSPA_RAB_SuccEstab_CS_Conv) + SUM(VS_RAB_SuccEstabCS_Conv) +
                         SUM(VS_RAB_SuccEstabCS_Str) + SUM(VS_RAB_SuccEstabCS_AMR) +
                         SUM(VS_RAB_SuccEstabCS_AMRWB)) /
                        NULLIF((SUM(VS_HSPA_RAB_AttEstab_CS_Conv) + SUM(VS_RAB_AttEstabCS_Conv) +
                         SUM(VS_RAB_AttEstabCS_Str) + SUM(VS_RAB_AttEstab_AMR) +
                         SUM(VS_RAB_AttEstabCS_AMRWB)), 0)
                    ) * (
                        1 - (SUM(VS_RAB_AbnormRel_AMR) + SUM(VS_RAB_AbnormRel_AMRWB) +
                             SUM(VS_RAB_AbnormRel_CS_HSPA_Conv)) /
                        NULLIF((SUM(VS_RAB_AbnormRel_AMR) + SUM(VS_RAB_AbnormRel_AMRWB) +
                         SUM(VS_RAB_AbnormRel_CS_HSPA_Conv) + SUM(VS_RAB_NormRel_CS) +
                         SUM(VS_RAB_NormRel_CS_HSPA_Conv) + SUM(VS_RAB_NormRel_AMRWB)), 0)
                    ) AS `3G CS_CSSR`,
                    100 * (
                        (SUM(VS_RAB_FailEstabCS_DLIUBBand_Cong) + SUM(VS_RAB_FailEstabCS_ULIUBBand_Cong) +
                         SUM(VS_RAB_FailEstabCS_ULCE_Cong) + SUM(VS_RAB_FailEstabCS_DLCE_Cong) +
                         SUM(VS_RAB_FailEstabCS_Code_Cong) + SUM(VS_RAB_FailEstabCS_ULPower_Cong) +
                         SUM(VS_RAB_FailEstabCS_DLPower_Cong)) /
                        NULLIF((SUM(VS_RAB_AttEstabCS_Conv) + SUM(VS_RAB_AttEstabCS_Str)), 0)
                    ) AS `3G CALL BLOCK RATE`,
                    SUM(Total_PS_Traffic_GB) AS Total_PS_Traffic_GB
                FROM hourly_huawei_3g_all_counters_1 h1
                INNER JOIN sites_and_info_3g_huawei AS s
                    ON s.UCELL = h1.cell_name
                WHERE
                    h1.DATE BETWEEN '{{start_date}}' AND '{{end_date}}'
                    AND FIND_IN_SET(s.site, '{{sites}}')
                GROUP BY
                    h1.date, {time_group_t1}s.SITE
            ) AS T1
        INNER JOIN
            (
                -- T2: From hourly_huawei_3g_all_counters_2
                SELECT
                    date,
                    {time_group_t2}
                    s.SITE AS SITE_NAME,
                    100 * (
                        (SUM(RRC_SuccConnEstab_OrgInterCall) + SUM(RRC_SuccConnEstab_TmItrCall) +
                         SUM(RRC_SuccConnEstab_OrgBkgCall) + SUM(RRC_SuccConnEstab_TmBkgCall) +
                         SUM(RRC_SuccConnEstab_OrgSubCall) + SUM(RRC_SuccConnEstab_OrgStrCall) +
                         SUM(RRC_SuccConnEstab_TmStrCall) + SUM(VS_RRC_AttConnEstab_EDCH) +
                         SUM(VS_RRC_AttConnEstab_HSDSCH)) /
                        NULLIF((SUM(RRC_AttConnEstab_OrgInterCall) + SUM(RRC_AttConnEstab_TmInterCall) +
                         SUM(RRC_AttConnEstab_TmBkgCall) + SUM(RRC_AttConnEstab_OrgBkgCall) +
                         SUM(RRC_AttConnEstab_OrgSubCall) + SUM(RRC_AttConnEstab_OrgStrCall) +
                         SUM(RRC_AttConnEstab_TmStrCall) + SUM(VS_RRC_SuccConnEstab_EDCH) +
                         SUM(VS_RRC_SuccConnEstab_HSDSCH)), 0)
                    ) * (
                        (SUM(VS_RAB_SuccEstabPS_Conv) + SUM(VS_RAB_SuccEstabPS_Str) +
                         SUM(VS_RAB_SuccEstabPS_Int) + SUM(VS_RAB_SuccEstabPS_Bkg)) /
                        NULLIF((SUM(VS_RAB_AttEstabPS_Conv) + SUM(VS_RAB_AttEstabPS_Str) +
                         SUM(VS_RAB_AttEstabPS_Int) + SUM(VS_RAB_AttEstabPS_Bkg)), 0)
                    ) AS `3G PS_CSSR`,
                    AVG(VS_HSDPA_MeanChThroughput) AS `3G Downlink Throughput (kbps)`,
                    AVG(VS_HSUPA_MeanChThroughput) AS `3G Uplink Throughput (kbps)`,
                    SUM(VS_AMR_Erlang_BestCell) AS `3G Traffic_Voix`
                FROM hourly_huawei_3g_all_counters_2 h2
                INNER JOIN sites_and_info_3g_huawei AS s
                    ON s.UCELL = h2.cell_name
                WHERE
                    h2.DATE BETWEEN '{{start_date}}' AND '{{end_date}}'
                    AND FIND_IN_SET(s.site, '{{sites}}')
                GROUP BY
                    h2.date, {time_group_t2}s.SITE
            ) AS T2
        ON
            T1.date = T2.date
            {time_join}
            AND T1.SITE_NAME = T2.SITE_NAME
        ORDER BY T1.date, {time_field} T1.SITE_NAME
    """


def get_4g_query(granularity='HOURLY'):
    """Generate 4G query based on granularity (HOURLY or DAILY)"""
    if granularity == 'HOURLY':
        time_field = "T1.TIME,"
        time_group = "TIME,"
        time_join = "AND T1.TIME = T2.TIME"
    else:
        time_field = ""
        time_group = ""
        time_join = ""

    return f"""
        SELECT
            T1.DATE,
            {time_field}
            T1.SITE_NAME,
            T1.CELL_NAME,
            ROUND(T1.PS_CSSR,2) `LTE PS_CSSR`,
            ROUND(T1.DEBIT_UL,2) `LTE Uplink Throughput (mbps)`,
            ROUND(T1.DEBIT_DL,2) `LTE Downlink Throughput (mbps)`,
            ROUND(T1.ERAB_DROP,2) ERAB_DROP,
            T2.total_4g_data 'LTE DataTrafic'
        FROM
            (
                SELECT
                    DATE,
                    {time_group}
                    NE_Name AS SITE_NAME,
                    SUBSTRING_INDEX(SUBSTRING_INDEX(UCELL, 'Cell Name=', -1), ',', 1) AS CELL_NAME,
                    100 * (SUM(RRCReqEmc) + SUM(RRCReqHiPri) + SUM(RRCReqModa) + SUM(RRCReqMT)) / (SUM(AttEMC) + SUM(AttHiPri) + SUM(AttModa) + SUM(AttMT)) * (SUM(ERABSucEst) / SUM(ERABAttEst)) AS PS_CSSR,
                    (SUM(ULThrp) / SUM(ULHiPre)) / 1000 AS DEBIT_UL,
                    (SUM(DLThrp) / SUM(DLHiPre)) / 1000 AS DEBIT_DL,
                    100 * (SUM(RABAbn) / (SUM(RABAbn) + SUM(RABNorrel))) AS ERAB_DROP
                FROM
                    hourly_arcep_huawei_4g
                WHERE
                    DATE BETWEEN '{{start_date}}' AND '{{end_date}}'
                    AND FIND_IN_SET(NE_Name, '{{sites}}')
                GROUP BY
                    DATE, {time_group}NE_Name, CELL_NAME
            ) AS T1
        INNER JOIN
            (
                SELECT
                    DATE,
                    {time_group}
                    SITE_NAME,
                    CELL_NAME,
                    ROUND(SUM(Total_Data), 2) AS total_4g_data
                FROM
                    hourly_huawei_trafic_per_cell_site_4g
                WHERE
                    DATE BETWEEN '{{start_date}}' AND '{{end_date}}'
                    AND FIND_IN_SET(SITE_NAME, '{{sites}}')
                GROUP BY
                    DATE, {time_group}SITE_NAME, CELL_NAME
            ) AS T2
        ON
            T1.DATE = T2.DATE
            {time_join}
            AND T1.SITE_NAME = T2.SITE_NAME
            AND T1.CELL_NAME = T2.CELL_NAME
    """


# Configuration dictionary
QUERIES_CONFIG = {
    "2G_KPIS": {
        "query_func": get_2g_query,
        "sheet_name": "2G_RAW_KPIS",
        "start_row": 1,
        "start_col": 1
    },
    "3G_KPIS": {
        "query_func": get_3g_query,
        "sheet_name": "3G_RAW_KPIS",
        "start_row": 1,
        "start_col": 1
    },
    "4G_KPIS": {
        "query_func": get_4g_query,
        "sheet_name": "4G_RAW_KPIS",
        "start_row": 1,
        "start_col": 1
    }
}


def clear_sheet_data(worksheet, start_row=2, max_row=1000, max_col=50):
    """Clear data from worksheet"""
    for row in worksheet.iter_rows(min_row=start_row, max_row=max_row,
                                    min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None


def execute_query_and_fill_excel(excel_file_path, query_key, start_date, end_date, sites_list, db_config, granularity='DAILY'):
    """
    Execute query and fill data into Excel

    Args:
        excel_file_path: Path to Excel file
        query_key: Key for query config (e.g., '2G_KPIS', '3G_KPIS', '4G_KPIS')
        start_date: Start date for query
        end_date: End date for query
        sites_list: List of sites to filter
        db_config: Database configuration dict
        granularity: 'HOURLY' or 'DAILY' (default: 'DAILY')
    """
    config = QUERIES_CONFIG[query_key]

    # Format sites
    sites_formatted = ",".join(sites_list)

    # Generate query based on granularity
    query_template = config['query_func'](granularity=granularity)

    # Build query with parameters
    query = query_template.format(
        start_date=start_date,
        end_date=end_date,
        sites=sites_formatted
    )
    print(query)
    # Execute query
    conn = mysql.connector.connect(**db_config)
    df = pd.read_sql(query, conn)
    conn.close()

    # Load workbook
    book = load_workbook(excel_file_path, keep_vba=True)
    worksheet = book[config['sheet_name']]

    # Clear existing data
    clear_sheet_data(worksheet, start_row=config.get('start_row', 2))

    # Fill data to Excel
    fill_to_excel_native3(
        file_path=excel_file_path,
        sheet_name=config['sheet_name'],
        df=df,
        start_row=config.get('start_row', 2),
        start_col=config.get('start_col', 1),
        withIndex=False,
        withColumn=True,
        worksheet=worksheet
    )

    book.save(excel_file_path)
    print(f"Data written to sheet: {config['sheet_name']} ({granularity})")


# Example usage
if __name__ == "__main__":
    db_config = {
        'host': '10.22.33.116',
        'user': 'root',
        'password': 'performance',
        'database': 'performanceroute'
    }

    excel_file = r"performance_automation\assets\vba_automation_kpi_monitoring.xlsm"
    start_date = "2026-02-11"
    end_date = "2026-02-25"
    #sites = [ "R_KAKI_KOKA" , "R_KPANE_BOUGNEROU" , "R_GBABIRE"  , "R_TABEROU" , "R_OUERE_SONKEROU" , "R_KAKOUROGOU"]
    sites = [ "KANDI_8","KANDI_9"]

    # Choose granularity: 'HOURLY' or 'DAILY'
   # granularity = "HOURLY"  # Change to 'HOURLY' for hourly data
    granularity = "DAILY"  # Change to 'HOURLY' for hourly data

    # Execute for all technologies with chosen granularity
    queries_to_run = ["2G_KPIS", "3G_KPIS", "4G_KPIS"]
    #queries_to_run = ["3G_KPIS",]

    for query_key in queries_to_run:
        execute_query_and_fill_excel(
            excel_file_path=excel_file,
            query_key=query_key,
            start_date=start_date,
            end_date=end_date,
            sites_list=sites,
            db_config=db_config,
            granularity=granularity  # 'HOURLY' or 'DAILY'
        )














