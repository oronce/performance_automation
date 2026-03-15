#!/usr/bin/env python3
"""
SMSC Stats Processor
Extracts SMS attempt statistics from SMSC failed stats sheets

Usage:
    1. Edit the FILE_PATHS list below with your file paths
    2. Run: python3 process_smsc_stats.py
"""

import openpyxl
import re
import sys
import os

# ============================================================================
# CONFIGURATION - EDIT YOUR FILE PATHS HERE
# ============================================================================

FILE_PATHS = [
    #r"C:\Users\o84432318\Downloads\MonthlyStat_\Moov Benin Combined SMSC Stats - April 2025.xlsx",
    r"C:\Users\o84432318\Downloads\MonthlyStat_\Moov Benin Combined SMSC Stats - July 2025.xlsx",
    # r"C:\Users\o84432318\Downloads\MonthlyStat_\Moov Benin Combined SMSC Stats - August 2025.xlsx",
    # r"C:\Users\o84432318\Downloads\MonthlyStat_\Moov Benin Combined SMSC Stats - September 2025.xlsx",
    # r"C:\Users\o84432318\Downloads\MonthlyStat_\Moov Benin Combined SMSC Stats - November 2025.xlsx",
    # r"C:\Users\o84432318\Downloads\MonthlyStat_\Moov Benin Combined SMSC Stats - December 2025.xlsx",
    # r"C:\Users\o84432318\Downloads\MonthlyStat_\Moov Benin Combined SMSC Stats - October 2025.xlsx",
    # r"C:\Users\o84432318\Downloads\MonthlyStat_\Moov Benin Combined SMSC Stats - July 2025.xlsx",
    # r"C:\Users\o84432318\Downloads\MonthlyStat_\Moov Benin Combined SMSC Stats - June 2025.xlsx",
    # r"C:\Users\o84432318\Downloads\MonthlyStat_\Moov Benin Combined SMSC Stats - May 2025.xlsx",
    # r"C:\Users\o84432318\Downloads\MonthlyStat_\Moov Benin Combined SMSC Stats - April 2025.xlsx",
    # r"C:\Users\o84432318\Downloads\MonthlyStat_\Moov Benin Combined SMSC Stats - March 2025.xlsx",
    # r"C:\Users\o84432318\Downloads\MonthlyStat_\Moov Benin Combined SMSC Stats - February 2025.xlsx",
    # r"C:\Users\o84432318\Downloads\MonthlyStat_\Moov Benin Combined SMSC Stats - January 2025.xlsx",
    # Add more files below (uncomment and edit):
    # "/mnt/user-data/uploads/file2.xlsx",
    # "/mnt/user-data/uploads/file3.xlsx",
]

# ============================================================================
# DO NOT EDIT BELOW THIS LINE
# ============================================================================

def process_smsc_file(file_path):
    """
    Process SMSC stats file following the defined workflow
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        dict: Results containing averages for each SMSC and total
    """
    
    print("\n" + "="*80)
    print(f"Processing file: {os.path.basename(file_path)}")
    print(f"Full path: {file_path}")
    print("="*80)
    
    # Step 1: Load workbook
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print("✓ File loaded successfully")
    except FileNotFoundError:
        print(f"✗ ERROR: File not found: {file_path}")
        sys.exit(1)
    except Exception as e:
        print(f"✗ ERROR: Failed to load file: {e}")
        sys.exit(1)
    
    # Step 2: Check required sheets exist
    required_sheets = ['SMSSC1 Failed Stats', 'SMSSC2 Failed Stats', 'SMSSC3 Failed Stats']
    available_sheets = wb.sheetnames
    
    print(f"\nAvailable sheets: {', '.join(available_sheets)}")
    
    for sheet_name in required_sheets:
        if sheet_name not in available_sheets:
            print(f"✗ ERROR: Required sheet '{sheet_name}' not found!")
            print(f"   Available sheets: {available_sheets}")
            wb.close()
            sys.exit(1)
    
    print("✓ All required sheets found")
    
    # Step 3: Process each SMSC sheet
    results = {}
    
    for sheet_name in required_sheets:
        print(f"\n{'-'*80}")
        print(f"Processing: {sheet_name}")
        print(f"{'-'*80}")
        
        ws = wb[sheet_name]
        
        # Step 3a: Check line 35 has "All data"
        line_35_value = ws.cell(row=35, column=1).value
        
        if line_35_value is None:
            print(f"✗ ERROR: Line 35 is empty!")
            wb.close()
            sys.exit(1)
        
        if not isinstance(line_35_value, str) or 'All data' not in line_35_value:
            print(f"✗ ERROR: Line 35 does not contain 'All data'")
            print(f"   Found: {line_35_value}")
            wb.close()
            sys.exit(1)
        
        print(f"✓ Line 35 verified: '{line_35_value}'")
        
        # Step 3b: Check line 36 format
        line_36_value = ws.cell(row=36, column=1).value
        
        if line_36_value is None:
            print(f"✗ ERROR: Line 36 is empty!")
            wb.close()
            sys.exit(1)
        
        if not isinstance(line_36_value, str):
            print(f"✗ ERROR: Line 36 is not a string")
            print(f"   Found type: {type(line_36_value)}")
            wb.close()
            sys.exit(1)
        
        # Check format: should contain "Stats: SPU" and "from X attempts"
        if 'Stats: SPU' not in line_36_value or 'from' not in line_36_value or 'attempts' not in line_36_value:
            print(f"✗ ERROR: Line 36 does not follow expected format")
            print(f"   Expected: 'Dec X XX:XX:XX.XXX : Stats: SPU 0 Previous minute fail percent (from XXXX attempts)'")
            print(f"   Found: {line_36_value}")
            wb.close()
            sys.exit(1)
        
        print(f"✓ Line 36 format verified")
        print(f"   Sample: {line_36_value[:80]}...")
        
        # Step 4: Extract attempts from line 36 onwards
        attempts_list = []
        error_lines = []
        
        for row_num in range(36, ws.max_row + 1):
            cell_value = ws.cell(row=row_num, column=1).value
            
            # Stop if we hit empty cells
            if cell_value is None:
                break
            
            if isinstance(cell_value, str):
                # Use regex to extract the number from "(from XXXX attempts)"
                match = re.search(r'from (\d+) attempts', cell_value)
                if match:
                    attempts = int(match.group(1))
                    attempts_list.append(attempts)
                else:
                    # Track lines that don't match pattern (might be end of data)
                    error_lines.append(row_num)
        
        # Step 5: Calculate statistics
        if len(attempts_list) == 0:
            print(f"✗ ERROR: No valid attempt data found!")
            wb.close()
            sys.exit(1)
        
        avg_attempts = sum(attempts_list) / len(attempts_list)
        
        print(f"✓ Data extraction complete")
        print(f"   Total lines processed: {len(attempts_list)}")
        print(f"   Average attempts: {avg_attempts:.2f}")
        print(f"   Min attempts: {min(attempts_list)}")
        print(f"   Max attempts: {max(attempts_list)}")
        print(f"   Total sum: {sum(attempts_list):,}")
        
        if error_lines and len(error_lines) < 10:
            print(f"   Note: {len(error_lines)} lines didn't match pattern (possibly end of data)")
        
        # Store results
        results[sheet_name] = {
            'average': avg_attempts,
            'count': len(attempts_list),
            'min': min(attempts_list),
            'max': max(attempts_list),
            'total': sum(attempts_list)
        }
    
    wb.close()
    
    # Step 6: Calculate total
    print(f"\n{'='*80}")
    print("FINAL RESULTS - AVERAGE ATTEMPTS SUMMARY")
    print(f"{'='*80}")
    
    total_average = 0
    for sheet_name in required_sheets:
        avg = results[sheet_name]['average']
        total_average += avg
        print(f"{sheet_name:25s}  {avg:>12,.2f}")
    
    print(f"{'-'*80}")
    print(f"{'TOTAL':25s}  {total_average:>12,.2f}")
    print(f"{'='*80}")
    
    results['total_average'] = total_average
    
    return results


if __name__ == "__main__":
    
    print("="*80)
    print("SMSC STATS PROCESSOR")
    print("="*80)
    
    # Check if FILE_PATHS is configured
    if not FILE_PATHS:
        print("\n✗ ERROR: No file paths configured!")
        print("\nPlease edit the FILE_PATHS list at the top of the script:")
        print("  FILE_PATHS = [")
        print('      "/path/to/your/file1.xlsx",')
        print('      "/path/to/your/file2.xlsx",')
        print("  ]")
        print("\nThen run: python3 process_smsc_stats.py")
        print("="*80)
        sys.exit(1)
    
    # Get file paths from configuration
    file_paths = FILE_PATHS
    
    # Check all files exist before processing
    print("\n" + "="*80)
    print("FILE VALIDATION")
    print("="*80)
    print(f"Files to process: {len(file_paths)}\n")
    
    missing_files = []
    for file_path in file_paths:
        if os.path.exists(file_path):
            print(f"✓ Found: {os.path.basename(file_path)}")
            print(f"  Path: {file_path}")
        else:
            print(f"✗ NOT FOUND: {file_path}")
            missing_files.append(file_path)
    
    if missing_files:
        print(f"\n✗ ERROR: {len(missing_files)} file(s) not found!")
        print("\nPlease check the file paths in the FILE_PATHS list at the top of the script.")
        sys.exit(1)
    
    print(f"\n✓ All {len(file_paths)} file(s) found!")
    
    # Process all files
    all_results = {}
    successful = 0
    failed = 0
    
    for file_path in file_paths:
        try:
            results = process_smsc_file(file_path)
            all_results[os.path.basename(file_path)] = results
            successful += 1
        except SystemExit:
            failed += 1
            print(f"\n✗ Failed to process: {file_path}")
        except Exception as e:
            failed += 1
            print(f"\n✗ Unexpected error processing {file_path}: {e}")
    
    # Print overall summary if multiple files
    if len(file_paths) > 1:
        print("\n\n" + "="*80)
        print("OVERALL SUMMARY - ALL FILES")
        print("="*80)
        print(f"Total files processed: {len(file_paths)}")
        print(f"Successful: {successful}")
        print(f"Failed: {failed}")
        
        if all_results:
            print("\n" + "-"*80)
            print(f"{'File Name':<40s} {'Total Average':>15s}")
            print("-"*80)
            
            grand_total = 0
            for filename, results in all_results.items():
                total_avg = results['total_average']
                grand_total += total_avg
                print(f"{filename:<40s} {total_avg:>15,.2f}")
            
            print("-"*80)
            print(f"{'GRAND TOTAL (All Files):':<40s} {grand_total:>15,.2f}")
            print("="*80)
    
    print("\n✓ Processing completed!")
    
    if successful == len(file_paths):
        print(f"✓ All {successful} file(s) processed successfully!")
    elif successful > 0:
        print(f"⚠ {successful} file(s) processed successfully, {failed} failed")